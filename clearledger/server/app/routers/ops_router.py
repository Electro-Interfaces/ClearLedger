"""
Управленческий кокпит ЭЗС (раздел «Управленческий», energy) — обзор ситуации,
конкретные расхождения и рабочие списки для управленческих решений.
Изолированный слой поверх L2 (station_energy_periods + charge_sessions +
station_contract_settlements) — см. services/ops_dashboard.py.
"""
import io
import uuid
from datetime import date

from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth import assert_company_member, get_current_user
from app.database import get_db
from app.models import User
from app.services import ops_closing, ops_expectations, ops_payments, ops_terms
from app.services.ops_dashboard import (
    ops_balance, ops_completeness, ops_overview, ops_station,
)

router = APIRouter(prefix="/ops", tags=["Управленческий кокпит (ЭЗС)"])


@router.get("/overview")
async def get_ops_overview(
    company_id: str = Query(...),
    region: str | None = Query(None, description="фильтр: регион (federalSubject)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Обзор: помесячная серия вход/отпуск/СН/сверхнорматив/деньги + светофор
    проблем с рабочими списками (нет данных · небаланс · простой · не оплачено ·
    договор истёк · нет тарифа · сироты)."""
    cid = await assert_company_member(company_id, current_user, db)
    return await ops_overview(db, cid, region=region)


@router.get("/balance")
async def get_ops_balance(
    company_id: str = Query(...),
    period: str | None = Query(None, description="месяц 'YYYY-MM-01'; пусто = последний полный"),
    region: str | None = Query(None, description="фильтр: регион (federalSubject)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Пообъектный энергобаланс за месяц: вход · отпуск · собственные нужды
    (оценка) · сверхнормативный небаланс · стоимость · выручка · маржа."""
    cid = await assert_company_member(company_id, current_user, db)
    return await ops_balance(db, cid, period, region=region)


@router.get("/completeness")
async def get_ops_completeness(
    company_id: str = Query(...),
    date_from: str | None = Query(None, alias="from", description="месяц 'YYYY-MM-01'"),
    date_to: str | None = Query(None, alias="to", description="месяц 'YYYY-MM-01'"),
    region: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """«Полнота данных»: каких документов/данных не хватает за период анализа —
    матрица «вид данных × месяц» + списки недостающего (станция → месяцы)."""
    cid = await assert_company_member(company_id, current_user, db)
    return await ops_completeness(db, cid, date_from, date_to, region=region)


@router.get("/station/{location_id}")
async def get_ops_station(
    location_id: str,
    company_id: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Карточка объекта (drill-down): паспорт + помесячный энергобаланс +
    договорные контуры (э/э / аренда / сервис) с контрагентами и оплатами."""
    cid = await assert_company_member(company_id, current_user, db)
    return await ops_station(db, cid, location_id)


# ===========================================================================
# Закрытие месяца по затратам: ожидания, документы, расчётные суммы
# ===========================================================================

def _period(value: str | None) -> str:
    """Месяц 'YYYY-MM-01'; пусто — предыдущий, он и есть закрываемый."""
    if value:
        return f"{value[:7]}-01"
    return ops_expectations.shift_month(date.today().isoformat()[:7] + "-01", -1)


@router.get("/closing")
async def get_ops_closing(
    company_id: str = Query(...),
    period: str | None = Query(None, description="месяц 'YYYY-MM-01'; пусто = предыдущий"),
    scope: str = Query("location", description="location | company | all"),
    status: str | None = Query(None, description="фильтр по статусу строки"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Рабочий стол закрытия месяца: что ждали, что пришло, чем закрыли.

    Ожидания разворачиваются при чтении: условие, заведённое сегодня, попадает
    в реестр текущего месяца сразу, а не после чьего-то отдельного действия.
    """
    cid = await assert_company_member(company_id, current_user, db)
    p = _period(period)
    expanded = await ops_expectations.expand(db, cid, p)
    period_row = await ops_closing.get_or_create(db, cid, p)
    await db.commit()

    stats = await ops_closing.counters(db, cid, p)
    return {
        "period": p,
        "status": (period_row.status if period_row.status != "open"
                   else ops_closing.auto_status(p, date.today().isoformat())),
        "closedAt": period_row.closed_at.isoformat() if period_row.closed_at else None,
        "counters": stats,
        "charges": await ops_closing.charges(db, cid, p, scope=scope, status=status),
        "blocked": expanded["blocked"],
        "contractsWithoutTerms": await ops_expectations.contracts_without_terms(db, cid, p),
    }


# ===========================================================================
# Условия начисления: из чего разворачивается ожидание месяца
# ===========================================================================
# Ведём их в карточке договора, а не отдельным реестром приложения: условие —
# это и есть договор, прочитанный учётом («5000 ₽ в месяц до 10-го числа»).
# Разносить договор и его условие по разным экранам значит заставлять человека
# держать в голове связь, которую система знает сама.

@router.get("/terms")
async def get_ops_terms(
    company_id: str = Query(...),
    contract_id: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Условия начисления — все или по одному договору."""
    cid = await assert_company_member(company_id, current_user, db)
    return await ops_terms.list_terms(db, cid, contract_id=contract_id)


@router.post("/terms")
async def post_ops_term(
    company_id: str = Query(...),
    payload: dict = Body(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Завести условие. Дальше ожидания месяцев появятся сами."""
    cid = await assert_company_member(company_id, current_user, db)
    try:
        row = await ops_terms.create_term(db, cid, payload)
    except ValueError as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    await db.commit()
    return row


@router.patch("/terms/{term_id}")
async def patch_ops_term(
    term_id: uuid.UUID,
    company_id: str = Query(...),
    payload: dict = Body(...),
    new_version: bool = Query(False, description="закрыть текущую версию и завести новую"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Изменить условие.

    `new_version=true` — правильный способ поднять ставку: старая версия
    закрывается датой, новая начинает действовать со следующего периода.
    Правка на месте переписала бы историю, включая уже закрытые месяцы.
    """
    cid = await assert_company_member(company_id, current_user, db)
    try:
        row = await ops_terms.update_term(db, cid, term_id, payload,
                                          new_version=new_version)
    except ValueError as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    await db.commit()
    return row


@router.delete("/terms/{term_id}")
async def delete_ops_term(
    term_id: uuid.UUID,
    company_id: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Удалить условие. Начисления закрытых периодов остаются нетронутыми."""
    cid = await assert_company_member(company_id, current_user, db)
    try:
        result = await ops_terms.delete_term(db, cid, term_id)
    except ValueError as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    await db.commit()
    return result


# ===========================================================================
# Документы контрагентов
# ===========================================================================

@router.get("/docs")
async def get_ops_docs(
    company_id: str = Query(...),
    match_status: str | None = Query(None, description="unmatched | auto | manual | rejected"),
    counterparty_id: str | None = Query(None),
    limit: int = Query(200, ge=1, le=1000),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Входящая первичка: что пришло, что не разобрано, что не привязано."""
    cid = await assert_company_member(company_id, current_user, db)
    return await ops_terms.list_docs(db, cid, match_status=match_status,
                                     counterparty_id=counterparty_id, limit=limit)


@router.post("/docs")
async def post_ops_doc(
    company_id: str = Query(...),
    payload: dict = Body(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Завести документ руками: акт, УПД, счёт, расшифровку.

    Пока приём почтой и ЭДО не подключены, это единственный вход первички —
    и он остаётся нужным и после: бумажный оригинал, привезённый курьером,
    ниоткуда сам не появится.
    """
    cid = await assert_company_member(company_id, current_user, db)
    try:
        row = await ops_terms.create_doc(db, cid, payload, user_id=current_user.id)
    except ValueError as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    await db.commit()
    return row


@router.post("/docs/upload", status_code=201)
async def upload_ops_doc(
    company_id: str = Query(...),
    doc_type: str = Query("act"),
    number: str | None = Query(None),
    doc_date: str | None = Query(None),
    period: str | None = Query(None),
    amount_gross: float | None = Query(None),
    counterparty_id: str | None = Query(None),
    contract_id: str | None = Query(None),
    charge_id: uuid.UUID | None = Query(None, description="сразу привязать к ожиданию"),
    note: str | None = Query(None),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Загрузить скан документа и завести карточку.

    Файл ложится в то же общее хранилище (`source_files`), что и документы
    проектов, поэтому скачивается существующей ручкой `/api/files/{id}` с её
    проверкой владельца — своего файлового контура контур затрат не заводит.

    `charge_id` позволяет закрыть ожидание тем же действием: человек, который
    держит в руках акт, не должен делать два шага там, где смысл один.
    """
    cid = await assert_company_member(company_id, current_user, db)
    content = await file.read()
    if not content:
        raise HTTPException(400, "Пустой файл")
    try:
        file_id = await ops_terms.store_file(db, cid, file.filename, file.content_type,
                                             content)
        row = await ops_terms.create_doc(db, cid, {
            "docType": doc_type, "number": number, "docDate": doc_date,
            "period": period, "amountGross": amount_gross,
            "counterpartyId": counterparty_id, "contractId": contract_id,
            "fileId": str(file_id), "note": note, "channel": "manual",
        }, user_id=current_user.id)
        if charge_id is not None:
            link = await ops_closing.attach_doc(
                db, cid, charge_id, uuid.UUID(row["id"]),
                amount_gross=amount_gross, user_id=current_user.id)
            row["attach"] = link
    except ValueError as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    await db.commit()
    return row


@router.post("/docs/{doc_id}/file", status_code=201)
async def attach_ops_doc_file(
    doc_id: uuid.UUID,
    company_id: str = Query(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Прикрепить скан к уже заведённому документу.

    Обычный порядок в жизни обратный тому, что удобно программисту: сначала
    сумму вбили с телефона, потом дошли до сканера.
    """
    cid = await assert_company_member(company_id, current_user, db)
    content = await file.read()
    if not content:
        raise HTTPException(400, "Пустой файл")
    try:
        result = await ops_terms.attach_file(db, cid, doc_id, file.filename,
                                             file.content_type, content)
    except ValueError as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    await db.commit()
    return result


@router.get("/periods")
async def get_ops_periods(
    company_id: str = Query(...),
    months: int = Query(12, ge=1, le=36, description="сколько месяцев показать назад"),
    ahead: int = Query(2, ge=0, le=12, description="сколько месяцев вперёд развернуть"),
    expand_all: bool = Query(True, description="развернуть ожидания по всему диапазону"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Состояние каждого отчётного периода: сколько ждали и сколько закрыли.

    Договоры переходящие, поэтому ожидания разворачиваются сразу по всему
    диапазону, включая месяцы вперёд: сроки известны заранее, и собирать
    документы надо не в последний день, а по календарю.
    """
    cid = await assert_company_member(company_id, current_user, db)
    to = ops_expectations.shift_month(date.today().isoformat()[:7] + "-01", ahead)
    frm = ops_expectations.shift_month(to, -(months + ahead - 1))
    if expand_all:
        await ops_expectations.expand_range(db, cid, frm, to)
        await db.commit()
    return await ops_closing.periods_scale(db, cid, frm, to)


@router.get("/counterparties")
async def get_ops_counterparties(
    company_id: str = Query(...),
    date_from: str | None = Query(None, alias="from"),
    date_to: str | None = Query(None, alias="to"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Кто как с нами работает: вовремя, с опозданием или молча — и контакты.

    Без этой таблицы «контрагент не присылает документы» остаётся ощущением.
    С ней это список имён с телефонами, по которому можно звонить.
    """
    cid = await assert_company_member(company_id, current_user, db)
    to = _period(date_to)
    frm = _period(date_from) if date_from else ops_expectations.shift_month(to, -11)
    return await ops_closing.counterparty_discipline(db, cid, frm, to)


@router.get("/calendar")
async def get_ops_calendar(
    company_id: str = Query(...),
    horizon: int = Query(45, ge=7, le=180, description="горизонт в днях"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Что и от кого ждём в ближайшие недели и что уже просрочено."""
    cid = await assert_company_member(company_id, current_user, db)
    return await ops_closing.calendar(db, cid, horizon_days=horizon)


@router.get("/charges")
async def get_ops_charges(
    company_id: str = Query(...),
    date_from: str | None = Query(None, alias="from", description="месяц 'YYYY-MM-01'"),
    date_to: str | None = Query(None, alias="to", description="месяц 'YYYY-MM-01'"),
    scope: str = Query("location", description="location | company | all"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Затраты объектов: матрица объект × месяц × статья за диапазон.

    Квартал и год своего закрытия не имеют — это свёртка закрытых месяцев,
    поэтому итоги считаются здесь, а не хранятся отдельной сущностью.
    """
    cid = await assert_company_member(company_id, current_user, db)
    to = _period(date_to)
    frm = _period(date_from) if date_from else ops_expectations.shift_month(to, -11)
    return await ops_closing.matrix(db, cid, frm, to, scope=scope)


@router.post("/closing/{period}/close")
async def post_ops_close(
    period: str,
    company_id: str = Query(...),
    force: bool = Query(False, description="закрыть, даже если часть строк нечем закрыть"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Закрыть месяц: незакрытое доводится расчётом, итоги замораживаются.

    Без `force` закрытие останавливается, если есть строки, у которых нет ни
    договорной суммы, ни истории: закрыть их нулём — значит записать в отчёт
    «платили ноль» вместо «не из чего считать».
    """
    cid = await assert_company_member(company_id, current_user, db)
    try:
        result = await ops_closing.close_period(db, cid, _period(period),
                                                user_id=current_user.id, force=force)
    except ValueError as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    await db.commit()
    return result


@router.post("/closing/{period}/reopen")
async def post_ops_reopen(
    period: str,
    company_id: str = Query(...),
    reason: str = Body(..., embed=True),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Открыть закрытый месяц заново. Причина обязательна и остаётся навсегда."""
    cid = await assert_company_member(company_id, current_user, db)
    try:
        result = await ops_closing.reopen_period(db, cid, _period(period), reason,
                                                 user_id=current_user.id)
    except ValueError as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    await db.commit()
    return result


@router.post("/charges/{charge_id}/doc")
async def post_charge_doc(
    charge_id: uuid.UUID,
    company_id: str = Query(...),
    payload: dict = Body(default_factory=dict),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Привязать документ контрагента к ожиданию и посчитать расхождение.

    Сумму можно передать явно (`amountGross`): один документ закрывает
    несколько объектов, и доля объекта — не сумма всего документа.
    """
    cid = await assert_company_member(company_id, current_user, db)
    doc_id = payload.get("docId")
    if not doc_id:
        raise HTTPException(status_code=400, detail="Не указан документ (docId)")
    try:
        result = await ops_closing.attach_doc(
            db, cid, charge_id, uuid.UUID(str(doc_id)),
            amount_gross=payload.get("amountGross"),
            amount_net=payload.get("amountNet"),
            qty=payload.get("qty"), user_id=current_user.id)
    except (ValueError, TypeError) as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    await db.commit()
    return result


@router.post("/charges/{charge_id}/correction")
async def post_charge_correction(
    charge_id: uuid.UUID,
    company_id: str = Query(...),
    payload: dict = Body(default_factory=dict),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Провести расхождение закрытого месяца корректировкой в открытом.

    Закрытый период не переписываем: исходная строка остаётся со своей цифрой,
    а разница уходит отдельной строкой со ссылкой на источник.
    """
    cid = await assert_company_member(company_id, current_user, db)
    try:
        result = await ops_closing.make_correction(
            db, cid, charge_id, target_period=payload.get("period"),
            reason=payload.get("reason"), user_id=current_user.id)
    except ValueError as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    await db.commit()
    return result

# ---------------------------------------------------------------------------
# Кассовый факт: сколько по объектам сети реально заплатили
# ---------------------------------------------------------------------------

@router.post("/payments/upload", status_code=201)
async def upload_ops_payments(
    company_id: str = Query(...),
    sheet: str | None = Query(None, description="лист выгрузки; по умолчанию первый"),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Принять сводную выгрузку списаний казначейства.

    «Хозяйство» знало только ожидания по договорам; сколько ушло со счёта — не знало
    вовсе, и вопрос «платим больше или меньше должного» оставался без ответа.

    Повторная загрузка того же файла не задваивает суммы: ключ строки — период,
    статья и контрагент. Незнакомые формулировки статей не теряются, а попадают в
    «прочие расходы» и возвращаются списком — по нему уточняется карта соответствий.
    """
    cid = await assert_company_member(company_id, current_user, db)
    content = await file.read()
    if not content:
        raise HTTPException(400, "Пустой файл")
    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        parsed = ops_payments.read_rows(ws.iter_rows(values_only=True),
                                        source_label=(file.filename or "")[:200])
        result = await ops_payments.store(db, cid, parsed)
    except ops_payments.PaymentsImportError as e:
        await db.rollback()
        raise HTTPException(400, str(e)) from e
    except KeyError as e:
        await db.rollback()
        raise HTTPException(400, f"Лист выгрузки не найден: {e}") from e
    await db.commit()
    return result


@router.get("/payments")
async def get_ops_payments(
    company_id: str = Query(...),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Факт против ожидания по периодам и статьям.

    Расхождение считается только по месяцам: годовые строки старых лет сравнивать
    не с чем — начислений за те годы в пространстве нет, и показать «минус всё»
    значило бы соврать.
    """
    cid = await assert_company_member(company_id, current_user, db)
    return await ops_payments.summary(db, cid, date_from=date_from, date_to=date_to)


@router.get("/payments/counterparties")
async def get_ops_payments_counterparties(
    company_id: str = Query(...),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    limit: int = Query(50, ge=1, le=500),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Кому платим больше всего. Пока объекты не связаны — это главный разрез."""
    cid = await assert_company_member(company_id, current_user, db)
    return await ops_payments.by_counterparty(
        db, cid, date_from=date_from, date_to=date_to, limit=limit)


@router.get("/payments/coverage")
async def get_ops_payments_coverage(
    company_id: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Сколько бухгалтерских номеров из выгрузки связано с объектами сети.

    Нужна честная цифра: пока номер заказчика в реестре соответствий не ведётся,
    разложить расход по площадкам нечем, и это должно быть видно, а не подразумеваться.
    """
    cid = await assert_company_member(company_id, current_user, db)
    return await ops_payments.objects_coverage(db, cid)
