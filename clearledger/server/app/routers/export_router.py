"""
Экспорт записей в JSON, Excel, CSV.
"""

import csv
import io
import uuid
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth import get_current_user
from app.database import get_db
from app.models import AuditEvent, DataEntry, User

router = APIRouter(prefix="/export", tags=["Экспорт"])


# ---------------------------------------------------------------------------
# Утилиты
# ---------------------------------------------------------------------------

async def _fetch_entries(
    db: AsyncSession,
    company_id: str | None = None,
    status_filter: str | None = None,
) -> list[DataEntry]:
    """Получает записи для экспорта."""
    query = select(DataEntry)
    if company_id:
        try:
            cid = uuid.UUID(company_id)
            query = query.where(DataEntry.company_id == cid)
        except ValueError:
            pass
    if status_filter and status_filter != "all":
        query = query.where(DataEntry.status == status_filter)

    query = query.order_by(DataEntry.created_at.desc())
    result = await db.execute(query)
    return list(result.scalars().all())


def _entry_to_dict(entry: DataEntry) -> dict:
    """Конвертирует запись в плоский словарь для экспорта."""
    return {
        "id": str(entry.id),
        "title": entry.title,
        "category_id": entry.category_id,
        "subcategory_id": entry.subcategory_id,
        "doc_type_id": entry.doc_type_id or "",
        "company_id": str(entry.company_id),
        "status": entry.status,
        "source": entry.source,
        "source_label": entry.source_label,
        "file_url": entry.file_url or "",
        "file_type": entry.file_type or "",
        "file_size": str(entry.file_size) if entry.file_size else "",
        "created_at": entry.created_at.isoformat() if entry.created_at else "",
        "updated_at": entry.updated_at.isoformat() if entry.updated_at else "",
    }


EXPORT_COLUMNS = [
    "id", "title", "category_id", "subcategory_id", "doc_type_id",
    "company_id", "status", "source", "source_label",
    "file_url", "file_type", "file_size", "created_at", "updated_at",
]

COLUMN_LABELS = {
    "id": "ID",
    "title": "Название",
    "category_id": "Категория",
    "subcategory_id": "Подкатегория",
    "doc_type_id": "Тип документа",
    "company_id": "Компания",
    "status": "Статус",
    "source": "Источник",
    "source_label": "Метка источника",
    "file_url": "URL файла",
    "file_type": "Тип файла",
    "file_size": "Размер",
    "created_at": "Создано",
    "updated_at": "Обновлено",
}


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------

@router.get("/json")
async def export_json(
    company_id: str | None = Query(None),
    status: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Экспорт в JSON (массив объектов)."""
    entries = await _fetch_entries(db, company_id, status)
    data = [_entry_to_dict(e) for e in entries]

    # Аудит экспорта
    if company_id:
        try:
            cid = uuid.UUID(company_id)
            event = AuditEvent(
                company_id=cid,
                user_id=str(current_user.id),
                user_name=current_user.name,
                action="exported",
                details=f"Экспорт JSON: {len(data)} записей",
            )
            db.add(event)
        except ValueError:
            pass

    return data


# ---------------------------------------------------------------------------
# Excel (.xlsx)
# ---------------------------------------------------------------------------

@router.get("/excel")
async def export_excel(
    company_id: str | None = Query(None),
    status: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Экспорт в Excel (.xlsx)."""
    entries = await _fetch_entries(db, company_id, status)

    wb = Workbook()
    ws = wb.active
    ws.title = "ClearLedger"

    # Заголовки
    for col_idx, col_key in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=COLUMN_LABELS.get(col_key, col_key))
        cell.font = cell.font.copy(bold=True)

    # Данные
    for row_idx, entry in enumerate(entries, 2):
        d = _entry_to_dict(entry)
        for col_idx, col_key in enumerate(EXPORT_COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=d.get(col_key, ""))

    # Автоширина колонок
    for col_idx, col_key in enumerate(EXPORT_COLUMNS, 1):
        max_len = len(COLUMN_LABELS.get(col_key, col_key))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
            min(max_len + 2, 50)
        )

    # Запись в буфер
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Аудит
    if company_id:
        try:
            cid = uuid.UUID(company_id)
            event = AuditEvent(
                company_id=cid,
                user_id=str(current_user.id),
                user_name=current_user.name,
                action="exported",
                details=f"Экспорт Excel: {len(entries)} записей",
            )
            db.add(event)
        except ValueError:
            pass

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"clearledger_{timestamp}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

@router.get("/csv")
async def export_csv(
    company_id: str | None = Query(None),
    status: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Экспорт в CSV (UTF-8 с BOM для Excel)."""
    entries = await _fetch_entries(db, company_id, status)

    buffer = io.StringIO()
    # BOM для корректного открытия кириллицы в Excel
    buffer.write("\ufeff")

    writer = csv.DictWriter(
        buffer,
        fieldnames=EXPORT_COLUMNS,
        extrasaction="ignore",
    )

    # Заголовки с русскими названиями
    writer.writerow(COLUMN_LABELS)

    for entry in entries:
        writer.writerow(_entry_to_dict(entry))

    # Аудит
    if company_id:
        try:
            cid = uuid.UUID(company_id)
            event = AuditEvent(
                company_id=cid,
                user_id=str(current_user.id),
                user_name=current_user.name,
                action="exported",
                details=f"Экспорт CSV: {len(entries)} записей",
            )
            db.add(event)
        except ValueError:
            pass

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"clearledger_{timestamp}.csv"

    content = buffer.getvalue().encode("utf-8")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
