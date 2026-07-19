"""/api/corporate/* — корпоративное направление ЭЗС (ЮЛ): реестр клиентов,
KPI-обзор, рентабельность (розница vs договор), биллинг. Питает пункт «Корпоратив».
"""
from __future__ import annotations

from datetime import date, datetime
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Response, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth import assert_company_member, get_current_user
from app.database import get_db
from app.models import ChargeSession, User
from app.services.corporate_service import CorporateService
from app.services.export_audit import log_export
from app.services.export_files import xlsx_response

router = APIRouter(prefix="/corporate", tags=["Корпоратив"])


def _d(s: str, field: str) -> date:
    try:
        return date.fromisoformat(s[:10])
    except (ValueError, TypeError) as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail=f"Invalid {field}: {s}") from exc


@router.get("/overview")
async def corporate_overview(
    company_id: str,
    date_from: str,
    date_to: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    """KPI корп-направления + топ-клиенты + алерты."""
    cid = await assert_company_member(company_id, current_user, db)
    return await CorporateService(db).overview(cid, _d(date_from, "date_from"), _d(date_to, "date_to"))


@router.get("/clients")
async def corporate_clients(
    company_id: str,
    date_from: str,
    date_to: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    """Реестр клиентов + метрики/гэп/тариф за период (Клиенты/Тарифы/Рентабельность/Биллинг)."""
    cid = await assert_company_member(company_id, current_user, db)
    return await CorporateService(db).clients(cid, _d(date_from, "date_from"), _d(date_to, "date_to"))


@router.get("/client-card")
async def corporate_client_card(
    company_id: str,
    client: str,
    date_from: str,
    date_to: str,
    history_months: int = 0,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    """Карточка одного ЮЛ: помесячная реализация за период, разрезы потребления
    (станции/регионы/коннекторы/карты), режим эксплуатации и срок отношений.

    `history_months` > 0 удлиняет ТОЛЬКО ряд месяцев (динамика на коротком
    контуре не читается); итоги и разрезы остаются по периоду."""
    cid = await assert_company_member(company_id, current_user, db)
    return await CorporateService(db).client_card(
        cid, client, _d(date_from, "date_from"), _d(date_to, "date_to"),
        max(0, min(history_months, 36)))


@router.get("/billing")
async def corporate_billing(
    company_id: str,
    date_from: str,
    date_to: str,
    client: str | None = None,
    vat_rate: float = 20.0,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict[str, Any]:
    """Данные к выставлению под УПД: сводка на клиента + номенклатура + разбивка НДС."""
    cid = await assert_company_member(company_id, current_user, db)
    return await CorporateService(db).billing(
        cid, _d(date_from, "date_from"), _d(date_to, "date_to"), client, vat_rate)


@router.get("/billing-export")
async def corporate_billing_export(
    company_id: str,
    date_from: str,
    date_to: str,
    client: str | None = None,
    vat_rate: float = 20.0,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Response:
    """Реестр к выставлению под УПД (xlsx, 2 листа): «Реестр» (сводка на клиента,
    НДС 20% выделен) + «Детализация» (номенклатура по договорному тарифу). Фильтры:
    период (обяз.), опц. конкретный клиент (один клиент = один УПД)."""
    cid = await assert_company_member(company_id, current_user, db)
    data = await CorporateService(db).billing(
        cid, _d(date_from, "date_from"), _d(date_to, "date_to"), client, vat_rate)

    import openpyxl
    wb = openpyxl.Workbook()
    vr = int(vat_rate)
    period = f"{date_from[:10]} — {date_to[:10]}"

    ws = wb.active
    ws.title = "Реестр"
    ws.append(["Клиент", "Период", "Сессий", "Энергия, кВтч",
               "Сумма без НДС ₽", f"НДС {vr}% ₽", "Сумма с НДС ₽", "Ср.тариф ₽/кВтч"])
    for s in data["summary"]:
        ws.append([s["client"], period, s["sessions"], s["energy_kwh"],
                   s["net"], s["vat"], s["gross"], s["avg_tariff"]])
    if data["summary"]:
        tot_net = round(sum(s["net"] for s in data["summary"]), 2)
        tot_vat = round(sum(s["vat"] for s in data["summary"]), 2)
        tot_gross = round(sum(s["gross"] for s in data["summary"]), 2)
        ws.append(["ИТОГО", period, sum(s["sessions"] for s in data["summary"]),
                   round(sum(s["energy_kwh"] for s in data["summary"]), 1),
                   tot_net, tot_vat, tot_gross, ""])

    wd = wb.create_sheet("Детализация")
    wd.append(["Клиент", "Услуга (номенклатура)", "Количество, кВтч", "Цена ₽/кВтч (с НДС)",
               "Сумма без НДС ₽", f"НДС {vr}% ₽", "Сумма с НДС ₽"])
    for d in data["detail"]:
        usluga = f"Зарядка ЭЗС, тариф {d['tariff']:g} ₽/кВтч (НДС в т.ч.)"
        wd.append([d["client"], usluga, d["energy_kwh"], d["tariff"],
                   d["net"], d["vat"], d["gross"]])

    # Лист «Транзакции» — приложение к УПД: список сессий, подтверждающих сумму.
    S = ChargeSession
    lo = datetime.combine(_d(date_from, "date_from"), datetime.min.time())
    hi = datetime.combine(_d(date_to, "date_to"), datetime.max.time())
    sc = [S.company_id == cid, S.client_name.is_not(None), S.client_amount.is_not(None),
          S.started_at.is_not(None), S.started_at >= lo, S.started_at <= hi]
    if client:
        sc.append(S.client_name == client)
    sessions = (await db.execute(
        select(S).where(*sc).order_by(S.client_name, S.started_at).limit(100000))).scalars().all()
    vf = vat_rate / (100.0 + vat_rate)
    wtx = wb.create_sheet("Транзакции")
    wtx.append(["№", "ID сессии", "Дата", "Клиент", "Станция", "Регион", "Коннектор",
                "Энергия кВтч", "Тариф ₽/кВтч (с НДС)", "Сумма без НДС ₽", f"НДС {vr}% ₽", "Сумма с НДС ₽"])
    for i, s in enumerate(sessions, 1):
        gross = float(s.client_amount or 0)
        vat = round(gross * vf, 2)
        wtx.append([i, s.session_ext_id,
                    s.started_at.strftime("%d.%m.%Y %H:%M") if s.started_at else "",
                    s.client_name, s.station_code, s.region, s.connector_type,
                    round(float(s.energy_kwh or 0), 3), float(s.client_tariff or 0),
                    round(gross - vat, 2), vat, round(gross, 2)])

    # След выгрузки: основание для выставления УПД — сумма отсюда уходит клиенту.
    total = round(sum(s["gross"] for s in data["summary"]), 2) if data["summary"] else 0.0
    log_export(db, cid, current_user,
               f"Реестр под УПД (xlsx): {len(data['summary'])} клиентов, "
               f"{len(sessions)} сессий, {total} ₽ с НДС {vr}%, "
               f"период {date_from[:10]}…{date_to[:10]}"
               + (f", клиент «{client}»" if client else ""))

    fname = f"Реестр под УПД {date_from[:10]} — {date_to[:10]}"
    if client:
        fname += f" · {client}"
    return xlsx_response(wb, f"{fname}.xlsx")
