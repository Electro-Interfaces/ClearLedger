"""Приём первичных документов в пространство (docs/INTAKE.md).

Загрузка — не «положить файл в папку». Между «файл прочитан» и «документ в учёте»
лежит работа, которую нельзя пропускать:

  1. РАЗОБРАТЬ — вытащить из таблицы номер, дату, контрагента, сумму, строки;
  2. СОПОСТАВИТЬ — найти контрагента и договор среди уже заведённых;
  3. ПРОВЕРИТЬ — сверить с тем, что в пространстве уже накоплено;
  4. ПРИНЯТЬ — и только тогда документ появляется в учёте.

Третий шаг — главный. У пространства уже есть закрытые периоды, справочник
контрагентов, договоры и ранее загруженные документы; новый документ обязан с
ними согласовываться. Расхождение — не повод молча загрузить: это либо ошибка в
файле, либо новое обстоятельство (новый контрагент, новый договор), и человек
должен увидеть его до приёма, а не искать потом в отчётности.

Проверки намеренно разделены на «нельзя» (`error`), «странно» (`warning`) и
«к сведению» (`info`): приём блокирует только первое. Система не знает бизнеса
лучше человека — она знает, что уже загружено.
"""
from __future__ import annotations

import csv
import hashlib
import io
import re
from datetime import date, datetime
from typing import Any

from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import AccountingDoc, Contract, Counterparty, IntakeItem, Period

# Колонки распознаются по заголовку: файлы приходят из чужих систем, и назвать
# колонку «Контрагент» никто не обязан. Список синонимов — единственное место,
# где живёт знание о том, как эти поля называют в жизни.
_COLUMNS: dict[str, tuple[str, ...]] = {
    "number": ("номер", "№", "номер документа", "no", "number", "док. номер"),
    "date": ("дата", "дата документа", "date"),
    "counterparty": ("контрагент", "покупатель", "поставщик", "заказчик", "клиент",
                     "наименование контрагента", "получатель", "плательщик"),
    "inn": ("инн", "инн контрагента", "inn"),
    "amount": ("сумма", "сумма с ндс", "всего", "итого", "сумма документа", "amount"),
    "vat": ("ндс", "сумма ндс", "в т.ч. ндс", "vat"),
    "contract": ("договор", "основание", "контракт"),
    "doc_type": ("вид", "вид документа", "тип", "тип документа", "операция"),
    "item": ("номенклатура", "товар", "услуга", "позиция", "наименование", "товары"),
    "qty": ("количество", "кол-во", "qty"),
    "price": ("цена", "цена за единицу", "price"),
    "line_amount": ("сумма строки", "сумма позиции", "стоимость"),
}

# Вид документа из текста файла. Тот же словарь, что в реестре документов, —
# иначе загруженное назовётся не так, как всё остальное в пространстве.
_DOC_TYPES: list[tuple[str, tuple[str, ...]]] = [
    ("sale", ("реализация", "накладная", "упд", "отгрузка", "акт")),
    ("purchase", ("поступление", "приход", "закупка")),
    ("invoice_out", ("счёт покупателю", "счет покупателю", "счёт на оплату", "счет на оплату")),
    ("invoice_in", ("счёт от поставщика", "счет от поставщика", "входящий счёт")),
    ("vat_invoice_out", ("счёт-фактура выданный", "счет-фактура выданный")),
    ("vat_invoice_in", ("счёт-фактура полученный", "счет-фактура полученный")),
    ("bank_in", ("поступление на расчётный счёт", "поступление на расчетный счет", "приход банк")),
    ("bank_out", ("списание с расчётного счёта", "списание с расчетного счета", "расход банк")),
]


def _norm_header(v: Any) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip().lower())


def _num(v: Any) -> float:
    """Число из ячейки. Из чужих файлов оно приходит как угодно: с пробелами,
    запятой, знаком рубля и неразрывными пробелами Excel."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("\xa0", "").replace(" ", "").replace("₽", "").replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s)
    except ValueError:
        return 0.0


def _date(v: Any) -> str | None:
    """Дата в ISO. Excel отдаёт datetime, CSV — строку в любом из четырёх форматов."""
    if v in (None, ""):
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%y"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _doc_type_of(value: str | None, declared: str | None) -> str:
    """Вид документа: из строки файла, иначе из того, что сказал человек."""
    text_ = (value or "").lower()
    for code, words in _DOC_TYPES:
        if any(w in text_ for w in words):
            return code
    return declared or "sale"


def parse_table(content: bytes, file_name: str) -> tuple[list[dict[str, Any]], list[str]]:
    """Прочитать xlsx/csv в строки словарей. Возвращает (строки, найденные колонки)."""
    rows: list[list[Any]] = []
    if file_name.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    else:
        # Кодировка чужого CSV — почти всегда либо UTF-8, либо windows-1251;
        # разделитель определяем по первой строке, иначе «;» ломает разбор.
        raw = content.decode("utf-8-sig", errors="ignore")
        if raw.count(";") > raw.count(",") and raw.count(";") > 0:
            dialect_delim = ";"
        else:
            dialect_delim = ","
        rows = [r for r in csv.reader(io.StringIO(raw), delimiter=dialect_delim)]

    if not rows:
        return [], []

    # Шапка — первая строка, в которой распозналось хотя бы две наших колонки:
    # в выгрузках 1С сверху бывает название отчёта и период.
    header_idx, mapping = 0, {}
    for i, row in enumerate(rows[:15]):
        found: dict[int, str] = {}
        for j, cell in enumerate(row):
            h = _norm_header(cell)
            for key, names in _COLUMNS.items():
                if h in names or any(h.startswith(n) for n in names):
                    found.setdefault(j, key)
        if len(set(found.values())) >= 2:
            header_idx, mapping = i, found
            break
    if not mapping:
        return [], []

    out: list[dict[str, Any]] = []
    for row in rows[header_idx + 1:]:
        if not any(c not in (None, "") for c in row):
            continue
        rec: dict[str, Any] = {}
        for j, key in mapping.items():
            if j < len(row):
                rec[key] = row[j]
        rec["_raw"] = {str(rows[header_idx][j] or j): row[j]
                       for j in range(len(row)) if j < len(rows[header_idx])}
        out.append(rec)
    return out, sorted(set(mapping.values()))


def fingerprint(doc_type: str, number: str | None, day: str | None,
                inn: str | None, amount: float) -> str:
    """Отпечаток документа. Номер сам по себе не уникален (у разных контрагентов
    он повторяется), поэтому в ключ входят вид, дата, ИНН и сумма."""
    base = f"{doc_type}|{(number or '').strip()}|{day or ''}|{(inn or '').strip()}|{amount:.2f}"
    return hashlib.sha256(base.encode()).hexdigest()[:32]


async def build_items(db: AsyncSession, cid, batch_id, rows: list[dict[str, Any]],
                      declared_type: str | None) -> list[IntakeItem]:
    """Собрать кандидатов: одна строка файла — один документ (строки-позиции
    склеиваются к документу по номеру и дате, если файл построчный)."""
    by_doc: dict[tuple, dict[str, Any]] = {}
    for i, r in enumerate(rows):
        number = str(r.get("number") or "").strip() or None
        day = _date(r.get("date"))
        name = str(r.get("counterparty") or "").strip() or None
        inn = re.sub(r"\D", "", str(r.get("inn") or "")) or None
        key = (number, day, inn or name)
        doc = by_doc.setdefault(key, {
            "row_no": i + 1,
            "doc_type": _doc_type_of(str(r.get("doc_type") or ""), declared_type),
            "number": number, "date": day,
            "counterparty_name": name, "counterparty_inn": inn,
            "contract_name": str(r.get("contract") or "").strip() or None,
            "amount": 0.0, "vat": 0.0, "lines": [], "raw": r.get("_raw"),
        })
        line_amount = _num(r.get("line_amount")) or _num(r.get("amount"))
        if r.get("item"):
            doc["lines"].append({
                "name": str(r.get("item")).strip(),
                "qty": _num(r.get("qty")) or 1,
                "price": _num(r.get("price")),
                "amount": line_amount,
                "vat": _num(r.get("vat")),
            })
        # Сумма документа: если файл построчный, шапка складывается из строк;
        # если документ одной строкой — берём её сумму как есть.
        doc["amount"] = (sum(l["amount"] for l in doc["lines"])
                         if doc["lines"] else _num(r.get("amount")))
        doc["vat"] += _num(r.get("vat")) if doc["lines"] else 0.0
        if not doc["lines"]:
            doc["vat"] = _num(r.get("vat"))

    items = []
    for d in by_doc.values():
        items.append(IntakeItem(
            company_id=cid, batch_id=batch_id, row_no=d["row_no"],
            doc_type=d["doc_type"], number=d["number"], date=d["date"],
            counterparty_name=d["counterparty_name"], counterparty_inn=d["counterparty_inn"],
            contract_name=d["contract_name"],
            amount=round(d["amount"], 2), vat_amount=round(d["vat"], 2),
            lines=d["lines"], raw=d["raw"],
            fingerprint=fingerprint(d["doc_type"], d["number"], d["date"],
                                    d["counterparty_inn"], d["amount"]),
        ))
    return items


async def match_and_verify(db: AsyncSession, cid, items: list[IntakeItem]) -> None:
    """Сопоставить со справочниками и проверить против накопленного.

    Проверки идут от того, что в пространстве УЖЕ ЕСТЬ: закрытые периоды,
    карточки контрагентов, договоры, ранее принятые документы. Это и есть смысл
    приёмки — новый документ сверяется с историей, а не просто складывается.
    """
    # Справочники читаем один раз на пакет: файл на тысячу строк иначе даёт
    # тысячу запросов за тем же самым.
    cps = (await db.execute(select(Counterparty).where(
        Counterparty.company_id == cid))).scalars().all()
    by_inn = {(c.inn or "").strip(): c for c in cps if (c.inn or "").strip()}
    by_name = {_normname(c.name): c for c in cps}

    contracts = (await db.execute(select(Contract).where(
        Contract.company_id == cid))).scalars().all()

    closed = {(p.year, p.month) for p in (await db.execute(select(Period).where(
        Period.company_id == cid, Period.status == "closed"))).scalars().all()}

    for it in items:
        checks: list[dict[str, str]] = []

        # 1. Контрагент: ИНН → нормализованное имя. Новый контрагент — не ошибка,
        # но его надо завести в 1С, иначе документ повиснет без карточки.
        cp = by_inn.get((it.counterparty_inn or "").strip()) if it.counterparty_inn else None
        if cp is None and it.counterparty_name:
            cp = by_name.get(_normname(it.counterparty_name))
        if cp is not None:
            it.counterparty_id = cp.id
            if it.counterparty_inn and cp.inn and it.counterparty_inn != cp.inn:
                checks.append({"level": "warning", "code": "inn_mismatch",
                               "text": f"ИНН в файле ({it.counterparty_inn}) не совпадает "
                                       f"с карточкой ({cp.inn})"})
        else:
            checks.append({"level": "warning", "code": "new_counterparty",
                           "text": "Контрагента нет в справочнике — завести в 1С"})

        # 2. Договор — только внутри своего контрагента.
        if it.contract_name and it.counterparty_id:
            found = next((c for c in contracts
                          if str(c.counterparty_id) == str(it.counterparty_id)
                          and (c.type or "").strip().lower() == it.contract_name.strip().lower()),
                         None)
            if found is not None:
                it.contract_id = found.id
            else:
                checks.append({"level": "warning", "code": "new_contract",
                               "text": f"Договор «{it.contract_name}» у этого контрагента "
                                       f"не заведён"})

        # 3. Обязательные реквизиты: без даты и суммы документ не документ.
        if not it.date:
            checks.append({"level": "error", "code": "no_date", "text": "Не распознана дата"})
        if not it.number:
            checks.append({"level": "warning", "code": "no_number", "text": "Нет номера"})
        if not it.amount:
            checks.append({"level": "error", "code": "no_amount", "text": "Сумма нулевая"})

        # 4. Закрытый период — приём блокируется: закрытый месяц неизменяем, и
        # документ в нём меняет отчётность, которую уже сдали.
        if it.date:
            y, m = int(it.date[:4]), int(it.date[5:7])
            if (y, m) in closed:
                checks.append({"level": "error", "code": "period_closed",
                               "text": f"Период {m:02d}.{y} закрыт — документ туда не принимается"})

        # 5. Дубль: такой документ уже принят. Контрагента сверяем ССЫЛКОЙ, если она
        # нашлась, и лишь потом текстом: в файле ИНН часто не заполняют вовсе, и
        # ключ «номер + дата + ИНН» пропускал повторную загрузку той же накладной.
        dup = (await db.execute(text("""
            SELECT number, date FROM accounting_docs
             WHERE company_id = :cid AND doc_type = :t AND number = :n AND date = :d
               AND (
                 (CAST(:kid AS uuid) IS NOT NULL AND counterparty_id = CAST(:kid AS uuid))
                 OR (CAST(:kid AS uuid) IS NULL AND coalesce(:inn, '') <> ''
                     AND counterparty_inn = :inn)
                 OR (CAST(:kid AS uuid) IS NULL AND coalesce(:inn, '') = ''
                     AND lower(counterparty_name) = lower(coalesce(:name, '')))
               )
             LIMIT 1
        """), {"cid": str(cid), "t": it.doc_type, "n": it.number, "d": it.date,
               "kid": str(it.counterparty_id) if it.counterparty_id else None,
               "inn": it.counterparty_inn, "name": it.counterparty_name})).first()
        if dup:
            checks.append({"level": "error", "code": "duplicate",
                           "text": "Такой документ уже загружен"})
            it.status = "duplicate"

        # 6. Сумма строк против шапки — расхождение значит, что файл прочитан
        # неверно или в нём потеряна строка.
        if it.lines:
            lines_sum = round(sum(_num(l.get("amount")) for l in it.lines), 2)
            if abs(lines_sum - (it.amount or 0)) > 1:
                checks.append({"level": "warning", "code": "lines_mismatch",
                               "text": f"Сумма строк {lines_sum} ≠ сумме документа {it.amount}"})

        # 7. НДС: 20 % от суммы без НДС ± рубль. Ставка бывает и другой, поэтому
        # это предупреждение, а не запрет.
        if it.vat_amount and it.amount:
            implied = it.amount - it.amount / 1.2
            if abs(implied - it.vat_amount) > max(1.0, it.amount * 0.005):
                checks.append({"level": "info", "code": "vat_odd",
                               "text": f"НДС {it.vat_amount} не похож на 20 % "
                                       f"(ожидалось ≈ {implied:.2f})"})

        # 8. Сумма против истории этого контрагента: документ на порядок больше
        # обычного — повод посмотреть глазами, а не тихо принять.
        if it.counterparty_id and it.amount:
            avg = (await db.execute(text("""
                SELECT avg(amount) FROM accounting_docs
                 WHERE company_id = :cid AND counterparty_id = :kid AND doc_type = :t
                   AND amount > 0
            """), {"cid": str(cid), "kid": str(it.counterparty_id), "t": it.doc_type})).scalar()
            if avg and it.amount > float(avg) * 10:
                checks.append({"level": "info", "code": "amount_outlier",
                               "text": f"Сумма в {it.amount / float(avg):.0f} раз выше "
                                       f"обычной для этого контрагента"})

        it.checks = checks
        if it.status != "duplicate":
            it.status = ("blocked" if any(c["level"] == "error" for c in checks)
                         else "warning" if any(c["level"] == "warning" for c in checks)
                         else "ready")


def _normname(v: str | None) -> str:
    """Имя без кавычек и лишних пробелов — тот же ключ, что при сведении документов."""
    return re.sub(r"\s+", " ", re.sub(r"[«»\"'`]", "", (v or "")).strip()).lower()


async def accept_items(db: AsyncSession, cid, items: list[IntakeItem],
                       user_name: str | None = None) -> dict[str, Any]:
    """Принять кандидатов в учёт: создать документы и связать их с кандидатами."""
    created = skipped = 0
    for it in items:
        if it.status in ("blocked", "duplicate", "accepted", "rejected"):
            skipped += 1
            continue
        doc = AccountingDoc(
            company_id=cid,
            external_id=f"intake:{it.id}",
            doc_type=it.doc_type,
            number=it.number or "",
            date=it.date or "",
            counterparty_name=it.counterparty_name or "",
            counterparty_inn=it.counterparty_inn,
            counterparty_id=it.counterparty_id,
            contract_id=it.contract_id,
            amount=it.amount,
            vat_amount=it.vat_amount,
            status_1c="Загружен",
            lines=it.lines or [],
            doc_meta={"источник": "загрузка", "автор": user_name or "", "пакет": str(it.batch_id)},
        )
        db.add(doc)
        await db.flush()
        it.accounting_doc_id = doc.id
        it.status = "accepted"
        created += 1
    await db.commit()
    return {"created": created, "skipped": skipped}
