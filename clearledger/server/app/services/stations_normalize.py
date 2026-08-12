"""Нормализация справочника станций ЭЗС (energy): L1 RAW (Excel) → L2 CLEAN.

Оркестратор канала (source_type='stations_excel'):
  parse_stations_xlsx(content) → сырой паспорт станций (L1); формат определяется
  автоматически:
    • полный справочник HubEx — лист «Станции», 36 колонок (ID станции/Название/…);
    • компактная выгрузка CPO — лист «Stations», 9 колонок (Номер/Название/
      OCPP ID/Местоположение/Протокол/Производитель/Коннекторы/Статус/Владелец) —
      несёт ОПЕРАТИВНЫЕ статусы и OCPP ID, обновляет только присутствующие поля.
  ingest_stations(db, company, rows, channel_id, mode) → нормализация + UPSERT в ServiceLocation

Объект-станция = ServiceLocation (type='ev_charging'). Паспорт разложен в
ТИПИЗИРОВАННЫЕ колонки L2 (широта/долгота/мощность/коннекторы/бренд/OCPP/HubEx…),
неразложенный «хвост» — в extra_metadata. Регион канонизируется и резолвится в
regions (get-or-create → region_id). Идемпотентно (UPSERT по стабильному id):
  • append  — только новые станции (существующие пропускаются);
  • replace — обновить существующие + добавить новые (справочник = источник истины).
Удаление станций, пропавших из файла, НЕ делаем: на service_locations висят
каскадные FK (contract_locations, station_contract_settlements) — снос опасен.
"""
from __future__ import annotations

import hashlib
import io
import logging
import re as _re
import uuid as _uuid
from typing import Any

from sqlalchemy import func, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import ChannelSyncLog, ChargeSession, Region, ServiceLocation
from app.services.mapping import canon_brand, canon_city, canon_region, geo_in_russia

logger = logging.getLogger("clearledger.stations")


def _s(v, maxlen: int | None = None) -> str | None:
    if v is None:
        return None
    out = str(v).strip()
    if not out:
        return None
    return out[:maxlen] if maxlen else out


def _num(v) -> float | None:
    if v is None:
        return None
    try:
        s = str(v).replace(",", ".").strip()
        return float(s) if s else None
    except (ValueError, TypeError):
        return None


def _int(v) -> int | None:
    f = _num(v)
    return int(f) if f is not None else None


def _truthy(v) -> bool | None:
    if v is None:
        return None
    s = str(v).strip().lower()
    if not s:
        return None
    return s in ("да", "true", "1", "yes", "y", "+", "опубликована", "published")


def _status_from_stage(stage) -> str:
    """Стадия жизненного цикла → status точки (active|closed|planned)."""
    s = (stage or "").strip().lower()
    if "закры" in s or "closed" in s or "демонт" in s:
        return "closed"
    if "план" in s or "plan" in s or "проект" in s:
        return "planned"
    return "active"


def _oper_status(status_dev) -> str:
    """Операционный статус станции.

    Четыре статуса CPO переносятся в объекты 1:1 — они показывают ТЕКУЩЕЕ состояние
    станции и не должны схлопываться:
      «Активная» → working | «Нет связи» → no_link | «Отключена» → disabled |
      «Выведена из эксплуатации» → decommissioned.
    unknown остаётся только для объектов БЕЗ данных CPO (нет в выгрузке).
    not_working/on_repair/maintenance — из других контуров (ручная смена, HubEx,
    складской демонтаж), CPO их не выставляет."""
    s = (status_dev or "").strip().lower()
    if not s:
        return "unknown"
    if "выведен" in s or "демонтир" in s or "демонтаж" in s:
        return "decommissioned"  # выведена из эксплуатации — не в сети
    if "отключ" in s:            # CPO «Отключена» — станция намеренно выключена
        return "disabled"
    if "нет связи" in s or "офлайн" in s or "оффлайн" in s or "offline" in s:
        return "no_link"        # нет телеметрии — НЕ поломка станции
    if "ремонт" in s:
        return "on_repair"
    if "обслуж" in s or "maintenance" in s:
        return "maintenance"
    if "не работает" in s or "неисправ" in s or "сломан" in s or "авар" in s:
        return "not_working"
    if "работает" in s or "онлайн" in s or "online" in s or "доступ" in s or "активн" in s:
        return "working"
    return "unknown"


# Классы номера ЭЗС (канон РусГидро, 2026-07-17):
#   «1»…«714»  — станции РусГидро;
#   «65.135»   — станции СНК (партнёр нумерует сам: регион.порядковый);
#   «TS1033»   — станция СНК (единичный формат);
#   «Тест»     — тестовые станции, в учёт НЕ берутся;
#   «72-1»/«73-1» — два исторических исключения (выведены, в отчётах не участвуют).
_RE_SNK_DOT = _re.compile(r"^\d+\.\d+$")
_RE_RH_NUM = _re.compile(r"^\d+$")


def _num_class(number) -> str:
    """Формат номера → класс станции: rushydro | snk | test | legacy."""
    n = (number or "").strip()
    if not n:
        return "legacy"
    if n.lower().startswith("тест") or n.lower().startswith("test"):
        return "test"
    if _RE_SNK_DOT.match(n) or n.upper().startswith("TS"):
        return "snk"
    if _RE_RH_NUM.match(n):
        return "rushydro"
    return "legacy"   # «72-1», «73-1» — старая нумерация


def _is_real_name(v: str | None) -> bool:
    """Подпись ли это вообще. В выгрузке CPO колонка «Название» иногда содержит
    номер станции («230», «268») — принимать такое за имя нельзя: осмысленная
    подпись («ЭЗС StarCharge») информативнее и терять её нельзя."""
    t = str(v or "").strip()
    return bool(t) and not _re.fullmatch(r"[\d\W_]+", t)


def _owner_by_class(num_class: str | None) -> str | None:
    """Класс номера → владелец сети (фолбэк, когда колонка «Владелец» пуста)."""
    return {"snk": "СНК", "rushydro": "РусГидро", "legacy": "РусГидро"}.get(num_class or "")


def _loc_id(company_id, ext_id: str) -> str:
    """Стабильный ServiceLocation.id для станции (детерминирован по company+ext)."""
    return "ezs-" + hashlib.md5(f"{company_id}|{ext_id}".encode()).hexdigest()[:20]


def _parse_compact(ws) -> list[dict[str, Any]]:
    """Компактная выгрузка CPO (лист «Stations», 9 колонок) → строки с меткой compact.

    Ключи: «Номер» = station_number (конформная размерность), OCPP ID уникален.
    Регион вытаскиваем из первого сегмента «Местоположения» (для новых станций)."""
    def g(r, i):
        return r[i] if len(r) > i else None

    rows: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        num, name, ocpp = _s(g(r, 0), 60), _s(g(r, 1), 255), _s(g(r, 2), 120)
        if not (num or ocpp):
            continue
        addr = _s(g(r, 3), 500)
        cls = _num_class(num)
        rows.append({
            "compact": True,
            "number": num, "name": name, "ocpp_id": ocpp,
            "address": addr,
            "region": (addr.split(",")[0].strip() if addr else None),
            "ocpp_protocol": _s(g(r, 4), 40),
            "brand": _s(g(r, 5), 120),
            "connector_types": _s(g(r, 6), 200),
            "status_dev": _s(g(r, 7), 60),
            "owner": _s(g(r, 8), 200),
            # Тест определяется ФОРМАТОМ НОМЕРА («Тест»), а не подстрокой в имени:
            # боевые станции с «тест» в названии (АЗС Импульс и пр.) — не тестовые.
            "num_class": cls,
            "is_test": cls == "test",
        })
    return rows


def parse_stations_xlsx(content: bytes) -> list[dict[str, Any]]:
    """Excel справочника станций → список сырого паспорта (L1). Формат — автоматически:
    компактная выгрузка CPO (заголовок «Номер|Название|OCPP ID…») либо полный
    справочник HubEx (лист «Станции», 36 колонок)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    for cand in wb.worksheets:
        head = next(cand.iter_rows(values_only=True, max_row=1), ()) or ()
        low = [str(h).strip().lower() if h is not None else "" for h in head]
        if len(low) >= 3 and low[0] == "номер" and "ocpp" in low[2]:
            return _parse_compact(cand)
    ws = wb["Станции"] if "Станции" in wb.sheetnames else wb[wb.sheetnames[-1]]

    def g(r, i):
        return r[i] if len(r) > i else None

    rows: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or g(r, 0) is None:
            continue
        tv = g(r, 31)
        rows.append({
            "ext_id": _s(g(r, 0), 64),
            "name": _s(g(r, 1), 255),
            "number": _s(g(r, 2), 60),
            "serial_number": _s(g(r, 3), 120),
            "owner_id": _int(g(r, 4)),
            "owner": _s(g(r, 5), 200),
            "region": _s(g(r, 6), 200),
            "city": _s(g(r, 7), 120),
            "street": _s(g(r, 8), 200),
            "house": _s(g(r, 9), 40),
            "address": _s(g(r, 10), 500),
            "latitude": _num(g(r, 11)),
            "longitude": _num(g(r, 12)),
            "model": _s(g(r, 13), 120),
            "brand": _s(g(r, 14), 120),
            "connectors_count": _int(g(r, 15)),
            "connector_types": _s(g(r, 16), 200),
            "power_kwt": _num(g(r, 17)),
            "ocpp_protocol": _s(g(r, 18), 40),
            "firmware": _s(g(r, 19), 80),
            "status_dev": _s(g(r, 20), 60),
            "stage": _s(g(r, 21), 40),
            "is_published": _truthy(g(r, 22)),
            "is_hidden": _truthy(g(r, 23)),
            "access_type": _s(g(r, 24), 60),
            "location_type": _s(g(r, 25), 60),
            "subsidy": _s(g(r, 26), 60),
            "always_on": _s(g(r, 27), 20),
            "rating": _num(g(r, 28)),
            "success_pct": _num(g(r, 29)),
            "is_test": tv is not None and "тест" in str(tv).lower(),
            "hubex_asset_id": _s(g(r, 32), 80),
            "hubex_name": _s(g(r, 33), 200),
            "hubex_link_status": _s(g(r, 34), 40),
            "hubex_note": _s(g(r, 35), 300),
        })
    return rows


async def _bump(db: AsyncSession, log_id, done: int, total: int, created: int, updated: int) -> None:
    """Промежуточный прогресс в лог (для %-индикатора UI) + commit для поллинга."""
    if log_id is None:
        return
    log = await db.get(ChannelSyncLog, log_id)
    if log is not None:
        log.loaded = created + updated
        log.events = [{"level": "info", "event": "run",
                       "message": f"станции: {done:,}/{total:,}".replace(",", " "),
                       "stations_done": done, "stations_total": total, "loaded": created + updated}]
    await db.commit()


async def _make_region_resolver(db: AsyncSession, company_id):
    """Кэш регионов компании: canon-name → region_id (get-or-create)."""
    region_cache: dict[str, _uuid.UUID] = {}
    for nm, rid in (await db.execute(
        select(Region.name, Region.id).where(Region.company_id == company_id))).all():
        region_cache[nm] = rid

    async def _region_id(raw):
        canon = canon_region(raw)
        if not canon:
            return None
        rid = region_cache.get(canon)
        if rid is None:
            reg = Region(company_id=company_id, name=canon, federal_subject=_s(raw, 200))
            db.add(reg)
            await db.flush()
            rid = reg.id
            region_cache[canon] = rid
        return rid

    return _region_id


async def _ingest_compact(
    db: AsyncSession, company_id, rows: list[dict[str, Any]],
    mode: str = "append", log_id=None,
) -> dict[str, Any]:
    """Компактная CPO-выгрузка → обновление ТОЛЬКО присутствующих в ней полей
    (операционный статус, OCPP ID/протокол, производитель) у существующих объектов
    + создание отсутствующих станций (в т.ч. партнёрских, владелец «СНК»).
    Паспортные поля полного справочника (координаты/мощность/владелец-ID…) не трогаем.

      append  — существующие объекты не трогаем, добавляем только новые станции;
      replace — обновить существующие + добавить новые (выгрузка = источник истины
                по оперативному статусу сети).

    Идентичность железки = OCPP ID (в выгрузке уникален), номер = ключ ПЛОЩАДКИ и
    уникальным не является: пары AC/DC стоят на одном номере (96, 110, 174, 175, 295).
    Поэтому резолвим OCPP-first, номер — лишь фолбэк для объектов без OCPP-ключа;
    иначе вторая станция пары садится на первую и молча теряет статус.

    Тестовые станции (номер «Тест») в учёт не берутся — строки пропускаются целиком.
    """
    _region_id = await _make_region_resolver(db, company_id)

    existing = (await db.execute(select(ServiceLocation).where(
        ServiceLocation.company_id == company_id, ServiceLocation.type == "ev_charging"))).scalars().all()
    by_num: dict[str, list[ServiceLocation]] = {}   # номер → ВСЕ носители (не уникален)
    by_ocpp: dict[str, ServiceLocation] = {}        # только OCPP ID карточки
    by_ident: dict[str, ServiceLocation] = {}       # идентификаторы источников и код карточки
    for loc in existing:
        # Тест-объекты исключены из индексов целиком: тест-строки мы не грузим, а
        # боевая строка не должна сесть на тестовую карточку по совпадению номера
        # (номер «123» носили и боевая «АЗС Импульс» в Ангарске, и тест-объект).
        if loc.is_test:
            continue
        md = loc.extra_metadata or {}
        if md.get("ocppId") and str(md["ocppId"]).strip():
            by_ocpp.setdefault(str(md["ocppId"]).strip(), loc)
        # Разные типы ключей — разные словари (12.08.2026): раньше OCPP ID, id
        # витрины, серийник и код карточки лежали в одном, и совпадение любого с
        # любым считалось попаданием. Так карточка получала чужой OCPP, а следующий
        # прогон разносил её идентичность по трём пространствам имён (№493: код
        # nsp_0335, OCPP в снимке nsp_0306, у боевой соседки nsp_0338).
        for k in (md.get("stationId"), md.get("ext_id"), md.get("asuimStationId"),
                  loc.serial_number, loc.code):
            if k is not None and str(k).strip():
                by_ident.setdefault(str(k).strip(), loc)
        num = loc.station_number or md.get("number")
        if num is not None and str(num).strip():
            by_num.setdefault(str(num).strip(), []).append(loc)

    created = updated = skipped = errors = 0
    tests = 0                   # тест-строки: в учёт не берём
    pair_split = 0              # вторые железки пар AC/DC → отдельные карточки
    ambiguous: list[str] = []   # номер носят несколько карточек — не угадываем
    touched: set[str] = set()   # карточки, уже занятые строкой этого файла
    total = len(rows)
    await _bump(db, log_id, 0, total, 0, 0)

    for idx, row in enumerate(rows):
        # Строка грузится в SAVEPOINT: с уникальными индексами (v2.16) конфликт
        # одной строки раньше отравлял сессию, и весь прогон канала падал на
        # финальном flush — вместо одной непринятой строки терялся весь файл.
        try:
          async with db.begin_nested():
            num, ocpp = row.get("number"), row.get("ocpp_id")
            if row.get("is_test"):
                tests += 1
                continue
            # OCPP-first: номер не уникален (пары AC/DC на одной площадке), OCPP — уникален.
            # Дальше — идентификаторы источников (у части карточек OCPP ID лежит в code),
            # и только потом номер, да и то если он принадлежит ровно одной карточке.
            loc = ((by_ocpp.get(ocpp) or by_ident.get(ocpp)) if ocpp else None)
            if loc is None and num:
                cands = by_num.get(str(num).strip()) or []
                if len(cands) > 1:
                    ambiguous.append(f"№{num} (кандидатов {len(cands)})")
                elif cands:
                    loc = cands[0]
            # Резолв по НОМЕРУ привёл на карточку, уже занятую другой строкой файла →
            # это вторая железка пары AC/DC (её OCPP-ключа в базе ещё нет). Раньше
            # строка молча пропускалась и DC-половина навсегда теряла статус —
            # теперь заводим ей собственную карточку (id детерминирован по OCPP,
            # повторный прогон найдёт её по ключу и просто обновит).
            if loc is not None and loc.id in touched:
                pair_split += 1
                loc = None
            # Карточка могла быть заведена ПРОШЛЫМ прогоном по тому же
            # детерминированному ключу и при этом выпасть из индексов: у соседних
            # карточек бывают загрязнённые code/ocppId (при замене железки OCPP новой
            # станции записывали старой), и by_key уводит ключ на чужой объект.
            # Без этой проверки db.add ронял flush по service_locations_pkey — и одна
            # такая строка обрушивала ВЕСЬ накат, а не только себя.
            if loc is None and (ocpp or num):
                lid = _loc_id(company_id, f"ocpp:{ocpp or num}")
                present = await db.get(ServiceLocation, lid)
                if present is not None:
                    if present.id in touched:
                        skipped += 1
                        logger.warning(
                            "stations_compact: строка %s (№%s / OCPP %s) — дубль ключа "
                            "внутри файла, карточка %s уже занята", idx + 1, num, ocpp, lid,
                        )
                        continue
                    loc = present
            oper = _oper_status(row.get("status_dev"))
            md_extra = {k: v for k, v in {
                "ocppId": ocpp,
                "statusDev": row.get("status_dev"),
                "ownerCpo": row.get("owner"),
                "numClass": row.get("num_class"),
            }.items() if v}
            if loc is not None:
                if mode == "append":
                    skipped += 1
                    continue
                loc.operational_status = oper
                if oper == "decommissioned":
                    loc.status = "closed"
                for fld in ("ocpp_protocol", "brand", "connector_types"):
                    if row.get(fld):
                        setattr(loc, fld, canon_brand(row[fld]) if fld == "brand" else row[fld])
                # поля полного паспорта не затираем — дозаполняем только пустые
                if not loc.owner and row.get("owner"):
                    loc.owner = row["owner"]
                if not loc.address and row.get("address"):
                    loc.address = row["address"]
                # Подпись из справочника CPO — главнее имени из сессий: реестр
                # ведёт сам оператор, а station_name в выгрузке сессий отстаёт и
                # бывает мусорным (на карточке ID 3114 так держалось «Нартис С-60»,
                # хотя по реестру это «Бистро Ням-Ням 1»). Прежнее имя не теряем.
                reg_name = _s(row.get("name"), 255)
                if reg_name and not _is_real_name(reg_name) and _is_real_name(loc.name):
                    reg_name = None      # «230» вместо названия — оставляем свою подпись
                if reg_name and str(loc.name or "").strip() != reg_name:
                    hist = [*((loc.extra_metadata or {}).get("nameHistory") or [])]
                    if loc.name and (not hist or hist[-1] != loc.name):
                        hist.append(loc.name)
                    md_extra["nameHistory"] = hist
                    loc.name = reg_name
                if reg_name:
                    md_extra["nameSource"] = "cpo_registry"
                if not loc.station_number and num:
                    loc.station_number = num
                if loc.region_id is None:
                    loc.region_id = await _region_id(row.get("region"))
                md = {**(loc.extra_metadata or {}), **md_extra}
                # Чужой OCPP ID на карточку не штампуем. Строка, дошедшая по НОМЕРУ,
                # раньше безусловно переписывала `ocppId`, и у карточки оказывались
                # три расходящихся идентификатора; на следующем прогоне все три вели
                # в один индекс и растаскивали идентичность станции.
                own_ocpp = str((loc.extra_metadata or {}).get("ocppId") or "").strip()
                if ocpp and own_ocpp and own_ocpp != str(ocpp).strip():
                    md["ocppId"] = own_ocpp
                    ambiguous.append(f"OCPP {ocpp} ↔ карточка {loc.code} ({own_ocpp})")
                if num:
                    md.setdefault("number", num)
                fs = canon_region(row.get("region")) or row.get("region")
                if fs:
                    md.setdefault("federalSubject", fs)
                loc.extra_metadata = md
                touched.add(loc.id)
                updated += 1
            else:
                canon = canon_region(row.get("region"))
                loc = ServiceLocation(
                    id=_loc_id(company_id, f"ocpp:{ocpp or num}"),
                    company_id=company_id, type="ev_charging",
                    # Код карточки = OCPP ID (уникален), а не номер: номер принадлежит
                    # площадке и на паре AC/DC у двух карточек совпадает.
                    code=ocpp or num,
                    name=row.get("name") or (f"ЭЗС №{num}" if num else ocpp),
                    station_number=num,
                    address=row.get("address"),
                    region_id=await _region_id(row.get("region")),
                    ocpp_protocol=row.get("ocpp_protocol"),
                    brand=canon_brand(row.get("brand")),
                    connector_types=row.get("connector_types"),
                    # Владелец: колонка CPO, а при пустой — по формату номера
                    # («65.135»/«TS1033» = СНК, число = РусГидро).
                    owner=row.get("owner") or _owner_by_class(row.get("num_class")),
                    status="closed" if oper == "decommissioned" else "active",
                    operational_status=oper,
                    is_test=False,   # тест-строки сюда не доходят
                    source_bindings=[],
                    extra_metadata={"number": num, "source": "stations_compact",
                                    "regionRaw": row.get("region"),
                                    "federalSubject": canon or row.get("region"),
                                    **({"nameSource": "cpo_registry"} if row.get("name") else {}),
                                    **md_extra},
                )
                db.add(loc)
                # Номер после этого носят две карточки — следующая строка с тем же
                # номером и без OCPP-ключа станет неоднозначной, и это правильно.
                if num:
                    by_num.setdefault(str(num).strip(), []).append(loc)
                if ocpp:
                    by_ocpp.setdefault(str(ocpp).strip(), loc)
                    by_ident.setdefault(str(ocpp).strip(), loc)
                touched.add(loc.id)
                created += 1
            if (idx + 1) % 200 == 0:
                await db.flush()
                await _bump(db, log_id, idx + 1, total, created, updated)
        except Exception as exc:  # noqa: BLE001
            # Молчаливый счётчик ошибок не даёт понять, ЧТО не загрузилось:
            # строка выпадает из справочника, а накат рапортует «success».
            errors += 1
            logger.warning(
                "stations_compact: строка %s (№%s / OCPP %s) не загружена: %s: %s",
                idx + 1, row.get("number"), row.get("ocpp_id"), type(exc).__name__, exc,
            )

    await db.flush()
    await _bump(db, log_id, total, total, created, updated)
    msg = f"CPO-выгрузка: обновлено {updated}, добавлено {created}, пропущено {skipped}"
    if tests:
        msg += f", тестовых исключено {tests}"
    if pair_split:
        msg += f", разведено пар AC/DC {pair_split}"
    if ambiguous:
        msg += f", неоднозначных ключей {len(ambiguous)}"
        logger.warning("stations_compact: неоднозначные ключи: %s", ", ".join(ambiguous[:20]))
    return {"status": "success", "mode": mode, "created": created, "updated": updated,
            "skipped": skipped, "errors": errors, "tests": tests, "pair_split": pair_split,
            "ambiguous": len(ambiguous), "ambiguous_keys": ambiguous[:50], "message": msg}


async def ingest_stations(
    db: AsyncSession, company_id, rows: list[dict[str, Any]], channel_id=None,
    mode: str = "append", log_id=None,
) -> dict[str, Any]:
    """L1 паспорт → нормализация (регион canon + region_id, статусы) → UPSERT в ServiceLocation."""
    if rows and rows[0].get("compact"):
        return await _ingest_compact(db, company_id, rows, mode=mode, log_id=log_id)
    _region_id = await _make_region_resolver(db, company_id)

    # Индексы СУЩЕСТВУЮЩИХ объектов ЭЗС — резолвим НА них (конформная размерность,
    # не плодим дубли). Ключи разных импортёров: stationId (HubEx) → серийный (code)
    # → № (number). Так канал станций ОБОГАЩАЕТ уже загруженные объекты паспортом.
    existing = (await db.execute(select(ServiceLocation).where(
        ServiceLocation.company_id == company_id, ServiceLocation.type == "ev_charging"))).scalars().all()
    # Ключи РАЗНЫХ типов держим в разных словарях (12.08.2026). Раньше код карточки
    # и её серийник лежали в одном индексе, поэтому серийник входящей строки мог
    # совпасть с КОДОМ чужой карточки — так карточка-артефакт переноса перехватила
    # боевой серийник владивостокской станции и живёт с чужим именем и номером.
    by_sid: dict[str, ServiceLocation] = {}                 # идентификаторы витрины/HubEx и код карточки
    by_serial: dict[str, ServiceLocation] = {}              # только серийные номера
    by_num: dict[str, list[ServiceLocation]] = {}           # номер → ВСЕ носители (номер не уникален)
    for loc in existing:
        md = loc.extra_metadata or {}
        for sid in (md.get("stationId"), md.get("ext_id"), md.get("asuimStationId"), loc.code):
            # code участвует как идентификатор витрины: у 385 станций её id лежит
            # именно там. Но как СЕРИЙНИК код больше не читается.
            if sid is not None and str(sid).strip():
                by_sid.setdefault(str(sid).strip(), loc)
        # Тест-объекты — только по ext_id: НЕ индексируем по serial/№, чтобы боевые
        # строки не сливались в них (и наоборот). Иначе мусор «001/Симулятор» с чужим
        # № затирает боевую станцию.
        if loc.is_test:
            continue
        if loc.serial_number and str(loc.serial_number).strip():
            by_serial.setdefault(str(loc.serial_number).strip(), loc)
        num = loc.station_number or md.get("number")
        if num is not None and str(num).strip():
            by_num.setdefault(str(num).strip(), []).append(loc)

    def _is_non_network(r) -> bool:
        # В витринной выгрузке (partial) «выведена из эксплуатации» — это ТЕКУЩЕЕ
        # состояние той же карточки, а не отдельная мусорная запись: витрина отдаёт
        # весь парк одним списком. Резолвить такие строки только по ext_id нельзя —
        # у карточек прежних импортов этого ключа нет, и первый же прогон завёл
        # дубль на каждую из 59 выведенных станций.
        if r.get("partial"):
            return bool(r.get("is_test"))
        return bool(r.get("is_test")) or "выведен" in str(r.get("status_dev") or "").lower()

    def _own_ids(loc: ServiceLocation) -> set[str]:
        """Идентификаторы витрины/HubEx, которые карточка уже носит."""
        md = loc.extra_metadata or {}
        return {str(v).strip() for v in
                (md.get("asuimStationId"), md.get("stationId"), md.get("ext_id"), loc.code)
                if v is not None and str(v).strip()}

    def _foreign(loc: ServiceLocation, ext_) -> bool:
        """У карточки СВОЙ идентификатор источника, и он не равен идентификатору строки.

        Значит это другая станция, просто с тем же номером: витрина кладёт AC- и
        DC-половину одной точки под общим номером (96, 174, 175) и под одним номером
        может держать две разные станции (295 «Фастовская»). Раньше вторая строка
        резолвилась по номеру на первую карточку и молча затирала её — так до нас не
        доезжали 4 боевые станции парка."""
        if not ext_:
            return False
        ids = _own_ids(loc)
        return bool(ids) and str(ext_).strip() not in ids

    ambiguous: list[str] = []

    def _resolve(r) -> ServiceLocation | None:
        ext_, serial, num = r.get("ext_id"), r.get("serial_number"), r.get("number")
        # Тест/выведенные резолвим ТОЛЬКО по ext_id (уникален) — отдельные объекты,
        # не сливаем по serial/№ с боевыми станциями.
        if _is_non_network(r):
            return by_sid.get(str(ext_).strip()) if ext_ else None
        loc = by_sid.get(str(ext_).strip()) if ext_ else None
        if loc is not None:
            return loc
        # Серийный номер — доказательство сам по себе: именно по нему строка витрины
        # садится на карточку, заведённую из HubEx (у неё нет id витрины вовсе).
        # Проверку «чужой идентификатор» к нему не применяем — иначе обогащение
        # паспорта из витрины перестало бы работать для всего парка HubEx.
        loc = by_serial.get(str(serial).strip()) if serial else None
        if loc is not None:
            return loc
        cands = by_num.get(str(num).strip()) if num else None
        if not cands:
            return None
        if len(cands) > 1:
            # Номер носят несколько карточек — угадывать нельзя: молча выбранный
            # «первый из БД» зависит от порядка строк в базе и меняется на перезаливке.
            ambiguous.append(f"№{num} (кандидатов {len(cands)})")
            return None
        return None if _foreign(cands[0], ext_) else cands[0]

    created = updated = skipped = errors = 0
    total = len(rows)
    await _bump(db, log_id, 0, total, 0, 0)

    for idx, row in enumerate(rows):
        # Строка — в SAVEPOINT (см. компактную ветку выше).
        try:
          async with db.begin_nested():
            ext = row.get("ext_id")
            if not ext:
                errors += 1
                continue
            loc = _resolve(row)
            if loc is not None and mode == "append" and loc.serial_number is not None:
                skipped += 1   # уже обогащён типизированным паспортом — не трогаем
                continue

            rid = await _region_id(row.get("region"))
            # Координаты вне России — мусор источника (симуляторы CPO стоят в
            # Гренландии и в Арктике). Сырьё остаётся в снимке, в паспорт не идёт:
            # иначе карта сети растягивается на полмира из-за трёх строк.
            if row.get("latitude") is not None and not geo_in_russia(
                    row.get("latitude"), row.get("longitude")):
                row = {**row, "latitude": None, "longitude": None}
            passport_extra = {  # неразложенный «хвост» паспорта (мержим, не затираем)
                "ext_id": ext,
                "isHidden": row.get("is_hidden"),
                "locationType": row.get("location_type"),
                "subsidy": row.get("subsidy"),
                "alwaysOn": row.get("always_on"),
                "hubexName": row.get("hubex_name"),
                "hubexNote": row.get("hubex_note"),
                # Канонический регион (fleet группирует по federalSubject → так дедуп
                # «Республика Татарстан (Татарстан)»/«Татарстан» в один регион).
                "federalSubject": canon_region(row.get("region")) or row.get("region"),
                "regionRaw": row.get("region"),
                "statusDev": row.get("status_dev"),
            }
            typed = dict(  # типизированные колонки паспорта (L2) — общие для update/create
                status=_status_from_stage(row.get("stage")),
                operational_status=_oper_status(row.get("status_dev")),
                address=row.get("address"),
                region_id=rid,
                serial_number=row.get("serial_number"),
                station_number=row.get("number"),
                # Город без приставки типа и бренд в каноническом написании: источники
                # пишут «г.Красноярск»/«Красноярск» и «Фора»/«FORA»/«Fora» вперемешку,
                # и справочник разъезжается на пустом месте (см. mapping.canon_*).
                city=canon_city(row.get("city")), street=row.get("street"), house=row.get("house"),
                latitude=row.get("latitude"), longitude=row.get("longitude"),
                power_kwt=row.get("power_kwt"),
                connectors_count=row.get("connectors_count"),
                connector_types=row.get("connector_types"),
                owner=row.get("owner"), owner_id=row.get("owner_id"),
                brand=canon_brand(row.get("brand")), model=row.get("model"),
                ocpp_protocol=row.get("ocpp_protocol"), firmware=row.get("firmware"),
                stage=row.get("stage"), access_type=row.get("access_type"),
                is_published=row.get("is_published"),
                is_test=bool(row.get("is_test")),
                hubex_asset_id=row.get("hubex_asset_id"),
                hubex_link_status=row.get("hubex_link_status"),
                rating=row.get("rating"), success_pct=row.get("success_pct"),
            )
            # Витринная выгрузка (asuim) отдаёт паспорт ЧАСТИЧНО: мощность, даты,
            # рейтинг, процент успеха, id бренда и группы пусты во всех строках.
            # Такой файл дозаполняет, а не переписывает: иначе один прогон стёр бы
            # мощность у 351 станции и обнулил бы их в разрезах по скорости.
            if row.get("partial"):
                # Два значения не становятся None от пустоты и потому проскакивали
                # фильтр ниже: `_oper_status(None)` возвращает «unknown», а `is_test`
                # приводится к False. На частичной выгрузке это гасило рабочий статус
                # станции и снимало метку тестовой — выкидываем их, если витрина
                # ничего про них не сказала.
                if not row.get("status_dev"):
                    typed.pop("operational_status", None)
                if not row.get("is_test"):
                    typed.pop("is_test", None)
                typed = {k: v for k, v in typed.items() if v is not None}
                passport_extra = {k: v for k, v in
                                  {**passport_extra, **(row.get("extra") or {})}.items()
                                  if v is not None}
                # Стадию жизненного цикла витрина не описывает («этап» = Active даже
                # у выведенных из эксплуатации). Берём её из статуса устройства:
                # выведена → closed, остальным своё значение не меняем.
                typed["status"] = ("closed" if typed.get("operational_status") == "decommissioned"
                                   else (loc.status if loc is not None else "active"))
                # Владелец в витрине — ОПЕРАТОР сети (РусГидро/СНК), в Учёте —
                # собственник по договору («ООО Доступная энергия»). Своё не трогаем.
                if loc is not None and loc.owner:
                    typed.pop("owner", None)
                # Тестовую станцию, которой у нас нет, не заводим: это прогоны
                # симуляторов CPO, в сети их не существует.
                if loc is None and row.get("is_test"):
                    skipped += 1
                    continue
            if loc is not None:
                # ОБОГАЩАЕМ существующий объект: заполняем типизированные колонки +
                # мержим паспорт; id / code / source_bindings НЕ трогаем.
                for k, v in typed.items():
                    setattr(loc, k, v)
                loc.extra_metadata = {**(loc.extra_metadata or {}), **passport_extra}
                # Подпись из справочника CPO главнее имени из сессий (см. _ingest_compact).
                reg_name = _s(row.get("name"), 255)
                if reg_name and not _is_real_name(reg_name) and _is_real_name(loc.name):
                    reg_name = None      # «230» вместо названия — оставляем свою подпись
                if reg_name and str(loc.name or "").strip() != reg_name:
                    hist = [*((loc.extra_metadata or {}).get("nameHistory") or [])]
                    if loc.name and (not hist or hist[-1] != loc.name):
                        hist.append(loc.name)
                    loc.extra_metadata = {**loc.extra_metadata, "nameHistory": hist,
                                          "nameSource": "cpo_registry"}
                    loc.name = reg_name
                elif reg_name:
                    loc.extra_metadata = {**loc.extra_metadata, "nameSource": "cpo_registry"}
                updated += 1
            else:
                lid = _loc_id(company_id, ext)
                num_key = str(row.get("number")).strip() if row.get("number") else None
                # Код карточки обязан быть уникальным в компании. Номер занят другой
                # станцией (пара AC/DC на одной точке) → кодом становится id источника.
                code = (row.get("number") if num_key and num_key not in by_num
                        else None) or row.get("serial_number") or ext
                fresh = ServiceLocation(
                    id=lid, company_id=company_id, type="ev_charging",
                    code=code, name=row.get("name") or row.get("number") or ext,
                    source_bindings=[], extra_metadata=passport_extra, **typed)
                db.add(fresh)
                # в индексы — чтобы вторая строка того же объекта в этом же файле
                # нашла свежесозданную карточку, а не завела ещё одну
                by_sid.setdefault(str(ext).strip(), fresh)
                by_sid.setdefault(str(code).strip(), fresh)
                if row.get("serial_number"):
                    by_serial.setdefault(str(row["serial_number"]).strip(), fresh)
                if num_key:
                    by_num.setdefault(num_key, []).append(fresh)
                created += 1

            if (idx + 1) % 200 == 0:
                await db.flush()
                await _bump(db, log_id, idx + 1, total, created, updated)
        except Exception as exc:  # noqa: BLE001
            errors += 1
            logger.warning(
                "stations_passport: строка %s (№%s / ext %s) не загружена: %s: %s",
                idx + 1, row.get("number"), row.get("ext_id"), type(exc).__name__, exc,
            )

    await db.flush()
    # Снятие метки «тестовая» по имени «ЭЗС%» убрано 12.08.2026: автоимя новой
    # карточки — ровно `ЭЗС №<номер>` (см. ниже по файлу), поэтому любая тест-строка,
    # получившая автоимя, на следующем прогоне становилась боевой и входила во все
    # индексы резолва. Признак теста даёт формат номера (_num_class) и явная
    # пометка источника; разбирать спорные карточки — отдельной операцией с отчётом,
    # а не побочным эффектом каждой загрузки.
    await _bump(db, log_id, total, total, created, updated)
    if ambiguous:
        logger.warning("stations_passport: не резолвлено по неоднозначному номеру: %s",
                       ", ".join(ambiguous[:20]))
    message = (f"переписано: обновлено {updated}, добавлено {created}" if mode == "replace"
               else f"добавлено {created}, пропущено {skipped}")
    if ambiguous:
        message += f", неоднозначных номеров {len(ambiguous)}"
    return {"status": "success", "mode": mode, "created": created, "updated": updated,
            "skipped": skipped, "errors": errors, "ambiguous": len(ambiguous),
            "ambiguous_keys": ambiguous[:50], "message": message}


async def build_station_index(db: AsyncSession, company_id) -> dict[str, str]:
    """Индекс № станции → ServiceLocation.id (конформная размерность) для резолва
    сессий/оплат на объект. Ключ — station_number (типизир.) или extra_metadata.number.

    Учитываются и ПРЕЖНИЕ номера карточки (`numberHistory`). Номер станции не
    вечен: при замене оборудования CPO меняет его, а выгрузка сессий ещё какое-то
    время присылает старый («580-1» против «580» в справочнике). Без алиасов такая
    сессия не находит свой объект, и конвейер заводит станцию-дубль заново —
    ровно то, что мы только что развели вручную.

    Один номер могут носить несколько карточек (тест-строки CPO, зеркала). Раньше
    выигрывала первая попавшаяся, и 1 336 сессий ангарской «АЗС Импульс» (№123)
    уехали на тест-станцию с тем же номером, а боевая карточка осталась с нулём.
    Поэтому кандидаты ранжируются: боевая раньше тестовой, подтверждённая
    витриной раньше неподтверждённой, действующая раньше закрытой.
    """
    # номер → (ранг карточки, id): побеждает минимальный ранг.
    idx_r: dict[str, tuple[tuple, str]] = {}
    alias_r: dict[str, tuple[tuple, str]] = {}

    def put(store: dict[str, tuple[tuple, str]], key: str, loc: ServiceLocation) -> None:
        md = loc.extra_metadata or {}
        r = (1 if loc.is_test else 0,
             0 if md.get("asuimStationId") else 1,
             1 if loc.status == "closed" else 0)
        cur = store.get(key)
        if cur is None or r < cur[0]:
            store[key] = (r, loc.id)

    for loc in (await db.execute(select(ServiceLocation).where(
        ServiceLocation.company_id == company_id, ServiceLocation.type == "ev_charging"))).scalars():
        md = loc.extra_metadata or {}
        num = loc.station_number or md.get("number")
        if num is not None and str(num).strip():
            put(idx_r, str(num).strip(), loc)
        for h in (md.get("numberHistory") or []):
            old = str((h or {}).get("было") or "").strip()
            if old:
                put(alias_r, old, loc)
    idx = {k: v[1] for k, v in idx_r.items()}
    # Действующие номера приоритетнее прежних: алиас не должен перехватить номер,
    # который сегодня принадлежит другой станции.
    for old, (_r, lid) in alias_r.items():
        idx.setdefault(old, lid)
    return idx


async def ensure_stations_from_sessions(db: AsyncSession, company_id) -> dict[str, Any]:
    """Durable self-heal каталога L2. Для КАЖДОГО номера станции, который встречается
    в сессиях, но отсутствует в справочнике объектов, создаём объект-ЗАГЛУШКУ ЭЗС из
    данных самих сессий (№/наименование/адрес/регион/коннекторы). Так «сироты»
    (location_id NULL) больше не копятся: неизвестная станция сразу заводится в L2.

    Заглушка помечена extra_metadata.source='auto_from_sessions' и station_number=№,
    поэтому позже официальный справочник (ingest_stations резолвит по № → by_num)
    попадёт на ЭТУ же строку и дообогатит её типизированным паспортом — без дублей."""
    existing = set((await build_station_index(db, company_id)).keys())

    grp = (await db.execute(select(
        ChargeSession.station_code.label("code"),
        func.max(ChargeSession.station_name).label("name"),
        func.max(ChargeSession.address).label("address"),
        func.max(ChargeSession.region).label("region"),
        func.array_agg(func.distinct(ChargeSession.connector_type)).label("conns"),
    ).where(ChargeSession.company_id == company_id, ChargeSession.station_code.is_not(None))
     .group_by(ChargeSession.station_code))).all()

    _region_id = await _make_region_resolver(db, company_id)

    created = 0
    codes: list[str] = []
    for r in grp:
        code = str(r.code).strip()
        # «580-1» — тот же номер с суффиксом коннектора: станция уже есть, заглушку
        # заводить нельзя (иначе она перехватит индекс и заберёт себе заезды).
        base = code.rsplit("-", 1)[0] if "-" in code else None
        if not code or code in existing or (base and base in existing):
            continue
        if _num_class(code) == "test":
            continue   # тестовые станции в каталог не заводим — и их сессии не грузим
        conns = ",".join(sorted({str(c) for c in (r.conns or []) if c})) or None
        rid = await _region_id(r.region)
        canon = canon_region(r.region)
        db.add(ServiceLocation(
            id=_loc_id(company_id, f"sess:{code}"),   # namespace 'sess:' — не коллидит с ext_id каталога
            company_id=company_id, type="ev_charging",
            code=code, station_number=code,
            name=_s(r.name, 255) or f"ЭЗС №{code}",
            address=_s(r.address, 500),
            region_id=rid,
            connector_types=_s(conns, 200),
            status="active", operational_status="unknown", is_test=False,
            source_bindings=[],
            extra_metadata={"number": code, "source": "auto_from_sessions",
                            "regionRaw": r.region, "federalSubject": canon or r.region},
        ))
        created += 1
        codes.append(code)
        existing.add(code)
    await db.flush()
    return {"status": "success", "created": created, "codes": codes}


async def backfill_session_locations(db: AsyncSession, company_id, auto_create: bool = False) -> dict[str, Any]:
    """Материализация связи сессии → объект-станция: проставить ChargeSession.location_id
    по № (резолв на объект). Bulk-UPDATE по каждому station_code. Идемпотентно.

    auto_create=True → сначала self-heal каталога (ensure_stations_from_sessions):
    неизвестные станции заводятся из сессий, поэтому покрытие уходит к ~100%.

    Код вида «580-1» (номер с суффиксом коннектора) резолвится по БАЗОВОМУ номеру,
    если сам код индексу неизвестен. CPO присылает такие коды массово — 11 602
    сессии парка; у четырнадцати станций их спасал рукописный алиас в
    `numberHistory`, а «580-1» алиаса не имел, и 349 заездов владивостокской
    «Авроровской» остались на карточке, чей номер давно другой (тестовая станция в
    Татарстане). Привязка не пересматривалась потому, что неизвестный код просто
    пропускался — теперь он находит свою станцию, а переезды считаются отдельно."""
    created_stations = 0
    if auto_create:
        created_stations = (await ensure_stations_from_sessions(db, company_id))["created"]
    idx = await build_station_index(db, company_id)
    codes = [str(x).strip() for x in (await db.execute(
        select(ChargeSession.station_code).where(ChargeSession.company_id == company_id).distinct()
    )).scalars() if x]
    linked_rows = matched = moved = 0
    for code in codes:
        lid = idx.get(code) or (idx.get(code.rsplit("-", 1)[0]) if "-" in code else None)
        if not lid:
            continue
        matched += 1
        # Сессии, которые сейчас висят на ДРУГОЙ карточке: это переезд, а не первая
        # привязка. Считаем отдельно — молчаливый переезд выручки между станциями
        # должен быть виден в отчёте загрузки.
        moved += int((await db.execute(select(func.count()).select_from(ChargeSession).where(
            ChargeSession.company_id == company_id, ChargeSession.station_code == code,
            ChargeSession.location_id.is_not(None), ChargeSession.location_id != lid,
        ))).scalar() or 0)
        r = await db.execute(update(ChargeSession).where(
            ChargeSession.company_id == company_id, ChargeSession.station_code == code
        ).values(location_id=lid))
        linked_rows += int(r.rowcount or 0)
    await db.flush()
    total = int((await db.execute(select(func.count()).select_from(ChargeSession).where(
        ChargeSession.company_id == company_id))).scalar() or 0)
    return {"status": "success", "sessions_linked": linked_rows, "sessions_total": total,
            "stations_matched": matched, "stations_total": len(codes),
            "stations_created": created_stations, "sessions_moved": moved,
            "message": (f"связано {linked_rows} из {total} сессий "
                        f"({matched}/{len(codes)} станций"
                        + (f", переехало {moved}" if moved else "")
                        + (f", создано {created_stations} из сессий" if created_stations else "") + ")")}


async def refresh_location_names(db: AsyncSession, company_id) -> int:
    """Имя объекта-станции = свежайшее имя из сессий — но ТОЛЬКО если подпись не
    подтверждена справочником CPO.

    Правило приоритета: реестр оператора > выгрузка сессий. Реестр ведёт сам CPO,
    а `station_name` в сессиях отстаёт и бывает мусорным — так карточка ID 3114
    держала «Нартис С-60», хотя по реестру это «Бистро Ням-Ням 1», и правка
    подписи откатывалась при каждой загрузке сессий. Карточки с
    `extra_metadata.nameSource='cpo_registry'` этот проход не трогает.

    Для станций, которых в реестре нет (заведены из сессий), поведение прежнее:
    имя берётся из последней сессии, старое уходит в extra_metadata.nameHistory
    (поиск по нему остаётся). Идемпотентно; возвращает число переименованных."""
    from sqlalchemy import text
    res = await db.execute(text(
        "WITH fresh AS ("
        "  SELECT DISTINCT ON (location_id) location_id, btrim(station_name) AS nm"
        "  FROM charge_sessions"
        "  WHERE company_id = :cid AND location_id IS NOT NULL"
        "    AND station_name IS NOT NULL AND btrim(station_name) <> ''"
        "  ORDER BY location_id, started_at DESC NULLS LAST) "
        "UPDATE service_locations sl SET "
        "  extra_metadata = jsonb_set(coalesce(sl.extra_metadata, '{}'::jsonb), '{nameHistory}',"
        "      coalesce(sl.extra_metadata->'nameHistory', '[]'::jsonb) || to_jsonb(sl.name)),"
        "  name = f.nm "
        "FROM fresh f "
        "WHERE f.location_id = sl.id AND sl.company_id = :cid "
        "  AND btrim(sl.name) IS DISTINCT FROM f.nm"
        "  AND coalesce(sl.extra_metadata->>'nameSource', '') <> 'cpo_registry'"
    ), {"cid": str(company_id)})
    return int(res.rowcount or 0)
