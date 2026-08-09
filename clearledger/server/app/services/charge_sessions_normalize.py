"""Нормализация зарядных сессий ЭЗС (energy): L1 RAW (Excel) → L2 CLEAN.

Вызывается оркестратором канала (source_type='charge_sessions_excel'):
  parse_sessions_xlsx(content) → сырые строки (L1)
  ingest_charge_sessions(db, company, rows, channel_id) → нормализация + сохранение

Нормализация коннектора/типа пользователя — через справочники mapping (kind
'connector'/'user_type'), с channel-override поверх company-default и канон-сида.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import re

from sqlalchemy import delete, func, or_, select, text, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import async_session_factory
from app.models import (
    ChannelSyncLog, ChargeSession, Contract, CorporateClient, Counterparty, Organization,
)
from app.services.analytics_cache import bump_version
from app.services.charge_mart import rebuild_mart
from app.services.mapping import apply, canon_region, load_kind_map
# Те же правила сведения имён, что у входящих договоров из реестров РусГидро — иначе
# одно юрлицо разойдётся на две карточки в общем реестре контрагентов пространства.
from app.services.reestr_normalize import _clean_cp_name, _cp_type, _normname

# Форматы дат в выгрузках сессий. Точки = DD.MM (RU), слэши = MM/DD (US, выгрузка
# ChargeTransactions «Сессии (69)» — «12/31/2025 08:20:44»). Разделитель различает,
# поэтому DD.MM и MM/DD не конфликтуют.
_DT_FORMATS = ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M",
               "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M",
               "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S")


def _num(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", ".").strip())
    except (ValueError, TypeError):
        return 0.0


def _dt(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.replace(tzinfo=None)
    s = str(v).strip()
    for fmt in _DT_FORMATS:
        try:
            return datetime.strptime(s[:26], fmt)
        except ValueError:
            continue
    return None


def _s(v, maxlen: int | None = None) -> str | None:
    if v is None:
        return None
    out = str(v).strip()
    if not out:
        return None
    return out[:maxlen] if maxlen else out


def _cut_key(user_type, charge_type, paid_at) -> str:
    """Разрез учёта «Расчёты по сессиям ЭЗС» — взаимоисключающая корзина по каналу
    расчёта (основа сверки выручки сессий с эквайрингом/ведомостями/1С):
      • ezs_admin  — служебные (ADMIN), вне выручки;
      • ezs_corp   — юрлица (user_type=ЮЛ) → сверка с ведомостями/договорами
                     per-контрагент (client_name из обогащения; корп-оплата
                     отложенная — amount мгновенного списания у них 0);
      • ezs_unpaid — картовая ФЛ-сессия без отметки оплаты → контроль утечки;
      • ezs_app    — ФЛ моб. приложение (оплачено) → сверка с эквайрингом/банком.
    Приоритет: служебные → ЮЛ → без оплаты → приложение. Корп-сигнал = user_type=ЮЛ
    (надёжнее прежней RFID-эвристики: совпадает 1:1 с наличием client_name). RFID —
    лишь способ запуска (charge_type), корзину больше не задаёт. Разрез сверки —
    reconcile_catalog.ezs_settlement. Набор/приоритет корзин меняется здесь."""
    ct = str(charge_type).strip().upper() if charge_type else ""
    ut = str(user_type).strip().upper() if user_type else ""
    if ct == "ADMIN":
        return "ezs_admin"
    if ut == "ЮЛ":
        return "ezs_corp"
    if paid_at is None:
        return "ezs_unpaid"
    return "ezs_app"


def parse_sessions_xlsx(content: bytes) -> list[dict[str, Any]]:
    """Excel ChargeTransactions → список сырых сессий (L1 RAW), по индексам колонок."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        started = _dt(r[5]) if len(r) > 5 else None
        finished = _dt(r[6]) if len(r) > 6 else None
        dur = round((finished - started).total_seconds() / 60, 2) if started and finished and finished >= started else 0.0
        rows.append({
            "session_ext_id": _s(r[0], 64),
            "station_code": _s(r[1] if len(r) > 1 else None, 40),
            "address": _s(r[2] if len(r) > 2 else None, 300),
            "connector_no": _s(r[3] if len(r) > 3 else None, 20),
            "connector_type": _s(r[4] if len(r) > 4 else None, 40),
            "started_at": started, "finished_at": finished, "duration_min": dur,
            "result": _s(r[7] if len(r) > 7 else None, 40),
            "charge_type": _s(r[9] if len(r) > 9 else None, 40),
            "rfid": _s((r[10] if len(r) > 10 else None) or (r[23] if len(r) > 23 else None), 120),
            "user_id": _s(r[11] if len(r) > 11 else None, 160),
            "energy_kwh": _num(r[12]) if len(r) > 12 else 0.0,
            "amount": _num(r[13]) if len(r) > 13 else 0.0,
            "tariff": _num(r[14]) if len(r) > 14 else 0.0,
            "paid_at": _dt(r[16]) if len(r) > 16 else None,
            "station_name": _s(r[17] if len(r) > 17 else None, 160),
            "region": _s(r[18] if len(r) > 18 else None, 120),
            "user_type": _s(r[21] if len(r) > 21 else None, 20),
            "payment_id": _s(r[24] if len(r) > 24 else None, 64),
        })
    return rows


async def ingest_charge_sessions(
    db: AsyncSession, company_id, rows: list[dict[str, Any]], channel_id=None,
    mode: str = "append", log_id=None,
) -> dict[str, Any]:
    """L1 RAW → нормализация (connector/user_type) → дедуп → сохранение (L2).

    mode:
      • 'append'  — подгрузить: добавить только те сессии, которых ещё нет
        (дедуп по «ID сессии»); существующие пропускаются.
      • 'replace' — переписать: удалить сессии ТЕКУЩЕГО скоупа загрузки и
        загрузить заново из таблицы. Скоуп = (company_id, channel_id): при
        загрузке через канал стираются только сессии этого канала; при прямом
        импорте (channel_id=NULL) — только сессии прямого импорта. Так «переписать»
        на одном канале НЕ уничтожает данные другого канала той же компании.
        Дедуп при вставке остаётся company-scoped (UNIQUE(company, session_ext_id)).
    """
    connector_map = await load_kind_map(db, company_id, "connector", channel_id)
    user_type_map = await load_kind_map(db, company_id, "user_type", channel_id)
    # Индекс объектов-станций (№ → location_id) — материализуем связь сессия→объект (FK).
    from app.services.stations_normalize import _num_class, build_station_index
    station_idx = await build_station_index(db, company_id)

    deleted = 0
    if mode == "replace":
        # Удаляем строго в скоупе текущей загрузки, а не по всей компании.
        del_conds = [ChargeSession.company_id == company_id]
        if channel_id is not None:
            del_conds.append(ChargeSession.channel_id == channel_id)
        else:
            del_conds.append(ChargeSession.channel_id.is_(None))
        res = await db.execute(delete(ChargeSession).where(*del_conds))
        deleted = int(res.rowcount or 0)
        await db.flush()

    # existing считаем ПОСЛЕ удаления: в режиме replace множество пустое → грузим всё
    existing: set[str] = set((await db.execute(
        select(ChargeSession.session_ext_id).where(ChargeSession.company_id == company_id)
    )).scalars().all())

    created = skipped = errors = 0
    tests = 0                  # сессии тестовых станций — в учёт не берём
    seen: set[str] = set()
    batch: list[ChargeSession] = []

    total = len(rows)

    async def _bump(done: int) -> None:
        """Прогресс в лог (для %-индикатора UI). Пишем в ОТДЕЛЬНОЙ транзакции, а НЕ
        коммитим основную: раньше `db.commit()` здесь коммитил незавершённый ingest
        (в replace — уже удалённые сессии, частично вставленные батчи) ДО финального
        bump_version. Воркер, обслуживающий дашборд во время загрузки, видел эти
        частичные данные под СТАРОЙ версией кэша и кэшировал их — источник «мигания»
        и рассинхрона воркеров. Теперь данные+визиты+bump_version = один финальный
        commit (атомарно), а прогресс поллинга идёт мимо основной транзакции."""
        if log_id is None:
            return
        async with async_session_factory() as log_db:
            log = await log_db.get(ChannelSyncLog, log_id)
            if log is not None:
                log.loaded = created
                log.events = [{"level": "info", "event": "run",
                               "message": f"загрузка сессий: {done:,}/{total:,}".replace(",", " "),
                               "stations_done": done, "stations_total": total, "loaded": created}]
                await log_db.commit()

    await _bump(0)
    for idx, row in enumerate(rows):
        try:
            sid = row.get("session_ext_id")
            if not sid:
                errors += 1
                continue
            # Сессии тестовых станций (код «Тест») в учёт не берём — это прогоны
            # симуляторов, они искажают выручку и надёжность сети.
            if _num_class(row.get("station_code")) == "test":
                tests += 1
                continue
            if sid in existing or sid in seen:
                skipped += 1
                continue
            seen.add(sid)

            raw_conn = (row.get("connector_type") or "").upper() or None
            connector = apply("connector", raw_conn, connector_map) if raw_conn else None
            user_type = apply("user_type", row.get("user_type"), user_type_map) if row.get("user_type") else None

            batch.append(ChargeSession(
                company_id=company_id,
                channel_id=channel_id,
                session_ext_id=sid,
                station_code=row.get("station_code"),
                location_id=station_idx.get((row.get("station_code") or "").strip()) or None,
                station_name=row.get("station_name"),
                region=canon_region(row.get("region")),
                address=row.get("address"),
                connector_no=row.get("connector_no"),
                connector_type=connector,
                started_at=row.get("started_at"), finished_at=row.get("finished_at"),
                duration_min=row.get("duration_min") or 0.0,
                result=row.get("result"), charge_type=row.get("charge_type"),
                user_type=user_type, user_id=row.get("user_id"),
                # UID карты — в верхнем регистре, как в справочнике витрины
                # (`ezs_rfid_cards.uid`). Выгрузка сессий отдаёт его строчными, и
                # прямой join «сессия ↔ карта» давал 6 % вместо 99,7 %: владелец
                # карты в отчётах терялся, хотя данные есть.
                rfid=(row.get("rfid") or None) and str(row["rfid"]).upper(),
                energy_kwh=row.get("energy_kwh") or 0.0, amount=row.get("amount") or 0.0,
                tariff=row.get("tariff") or 0.0,
                paid_at=row.get("paid_at"), payment_id=row.get("payment_id"),
                # Разрез учёта: помечаем корзину расчёта → ChargeSession в конвейере
                # достоверности (сверка выручки), а не «без разреза» как раньше.
                # Корп-сигнал = канонизированный user_type (ЮЛ), не сырой.
                cut_key=_cut_key(user_type, row.get("charge_type"), row.get("paid_at")),
            ))
            created += 1
            if len(batch) >= 1000:
                db.add_all(batch)
                await db.flush()
                batch = []
                await _bump(idx + 1)   # прогресс каждую 1000 строк
        except Exception:  # noqa: BLE001
            errors += 1

    if batch:
        db.add_all(batch)
        await db.flush()

    # Объекты — источник правды. Сессия резолвится на СУЩЕСТВУЮЩИЙ объект, но НЕ
    # заводит его: станция не из справочника «Объекты» больше НЕ превращается в
    # объект-призрак (auto_create=False — раньше self-heal плодил заглушки, и парк
    # объектов рос из выгрузок сессий вместо справочника). Такие сессии грузятся с
    # location_id=NULL и поднимаются «красной тряпкой» ниже. Повторный прогон после
    # заведения объекта свяжет их — backfill идемпотентен.
    from app.services.stations_normalize import backfill_session_locations, refresh_location_names
    heal = await backfill_session_locations(db, company_id, auto_create=False)

    # «Красная тряпка»: сессии, чья станция отсутствует в справочнике объектов
    # (location_id не проставился). Сессия без объекта недопустима — объекты
    # приоритетны, всё крутится вокруг них. Считаем по всей компании: инвариант —
    # «в базе нет ни одной сессии без объекта», а не только в текущем прогоне.
    unmatched_rows = (await db.execute(
        select(ChargeSession.station_code, func.count().label("n"))
        .where(ChargeSession.company_id == company_id,
               ChargeSession.location_id.is_(None),
               ChargeSession.station_code.is_not(None),
               func.btrim(ChargeSession.station_code) != "")
        .group_by(ChargeSession.station_code)
        .order_by(func.count().desc())
    )).all()
    unmatched_stations = [{"code": r.station_code, "sessions": int(r.n)} for r in unmatched_rows]
    unmatched_sessions = sum(u["sessions"] for u in unmatched_stations)
    # Переименования CPO: имя объекта = свежайшее имя из сессий (история в
    # extra_metadata.nameHistory) — каталог не отстаёт от загрузок.
    renamed = await refresh_location_names(db, company_id)

    # Визиты: склейка смежных попыток одного клиента на одной станции. Считаем
    # по всей компании, а не только по загруженным строкам — новая сессия может
    # примкнуть к визиту, который уже лежит в базе, и изменить его исход.
    from app.services.charge_visits import recompute_visits
    visits = await recompute_visits(db, company_id)

    if mode == "replace":
        message = f"переписано: удалено {deleted}, загружено {created}"
    else:
        message = f"загружено {created}, пропущено {skipped}"
    if tests:
        message += f"; сессий тестовых станций исключено: {tests}"
    if heal.get("stations_created"):
        message += f"; станций заведено из сессий: {heal['stations_created']}"
    if renamed:
        message += f"; объектов переименовано по свежим сессиям: {renamed}"
    if visits.get("visits"):
        message += (f"; визитов {visits['visits']:,}".replace(",", " ")
                    + f", успешных {visits['success_pct']}%"
                    + f", с повторными попытками {visits['retried']:,}".replace(",", " "))
    # Несопоставленные станции — вперёд сообщения: это ошибка загрузки, а не хвост.
    if unmatched_stations:
        preview = ", ".join(u["code"] for u in unmatched_stations[:15])
        more = f" +ещё {len(unmatched_stations) - 15}" if len(unmatched_stations) > 15 else ""
        message = (f"⚠ ОШИБКА: {unmatched_sessions} сессий по {len(unmatched_stations)} станциям "
                   f"ВНЕ справочника «Объекты» ({preview}{more}) — заведите станции в разделе "
                   f"«Объекты» или проверьте выгрузку. Загружено с location_id=NULL, "
                   f"повторный прогон свяжет. · ") + message
    await rebuild_mart(db, company_id)  # витрина L3 — атомарно с данными (коммит в bump)
    await bump_version(db, company_id)  # инвалидировать кеш дашбордов «Продаж»
    return {"status": "success", "mode": mode, "created": created, "skipped": skipped,
            "errors": errors, "deleted": deleted, "tests": tests,
            "stations_created": heal.get("stations_created", 0),
            "unmatched_stations": unmatched_stations,
            "unmatched_sessions": unmatched_sessions,
            "unmatched_station_count": len(unmatched_stations),
            "visits": visits,
            "message": message}


# ---------------------------------------------------------------------------
# Обогащение сессий справочником «Организации» (наименование корпоративного
# клиента → в строки ЮЛ-сессий по телефону). Пример enrichment-шага канала:
# отдельная справочная таблица джойнится в основную по ключу (телефон=user_id).
# ---------------------------------------------------------------------------
def _norm_phone(v) -> str:
    """Любой формат телефона → канон '+7XXXXXXXXXX' (как user_id сессий). '' если не телефон."""
    d = re.sub(r"\D", "", str(v or ""))
    if len(d) == 11 and d[0] in "78":
        d = "7" + d[1:]
    return "+" + d if len(d) == 11 else ""


def _num_or_none(v):
    """Число из ячейки (запятая→точка) или None для пустой."""
    if v is None:
        return None
    try:
        s = str(v).replace(",", ".").strip()
        return float(s) if s else None
    except (ValueError, TypeError):
        return None


# Матрица тарифа использует короткие столбцы CCS2/TYPE2/TYPE1; сессии хранят КАНОН-имена
# коннекторов (normalize_default: «CCS Combo 2»/«CHAdeMO»/«GB/T DC»/«Type 2»). Плюс старые
# прямые импорты — short-формы. Сопоставляем столбец матрицы → все варианты имени коннектора.
# CCS2 = DC-группа (DC-быстрые тарифицируются по цене CCS2 региона).
_MATRIX_CONN: dict[str, tuple[str, ...]] = {
    "CCS2": ("CCS2", "CCS Combo 2", "GBT_DC", "GB/T DC", "CHADEMO", "CHAdeMO"),
    "TYPE2": ("TYPE2", "Type 2"),
    "TYPE1": ("TYPE1",),
}


def parse_orgs_xlsx(content: bytes) -> dict[str, Any]:
    """Улучшенный справочник «Организации» → {orgs, matrix}.

    Лист «Organizations»: col1 «Название», col5 «Телефон», col6 «Режим тарифа»
    (matrix|flat|retail), col7 «Ставка ₽/кВтч» (для flat). Телефон — из col5 или
    первой ячейки-телефона в строке (устойчивость к сдвигу).
    Лист «Tariffs» (опц., для режима matrix): Организация|Регион|CCS2|TYPE2|TYPE1 —
    договорная матрица (резолв по коннектору; DC→CCS2)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb["Organizations"] if "Organizations" in wb.sheetnames else wb[wb.sheetnames[0]]
    orgs: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        name = _s(r[1] if len(r) > 1 else None, 300)
        phone = _norm_phone(r[5] if len(r) > 5 else None)
        if not phone:
            for c in r:
                p = _norm_phone(c)
                if p:
                    phone = p
                    break
        if not (name and phone):
            continue
        mode = (_s(r[6] if len(r) > 6 else None) or "").lower() or None
        rate = _num_or_none(r[7] if len(r) > 7 else None)
        u = _num_or_none(r[3] if len(r) > 3 else None)
        orgs.append({
            "phone": phone, "name": name, "mode": mode, "rate": rate,
            "ext_id": _s(r[0], 40),
            "contract_start": _s(r[2] if len(r) > 2 else None, 20),
            "users": int(u) if u is not None else None,
            "status": _s(r[4] if len(r) > 4 else None, 40),
            "discount": _num_or_none(r[8] if len(r) > 8 else None),  # «Скидка %» (опц. колонка)
        })

    matrix: dict[str, dict[str, dict[str, float]]] = {}
    if "Tariffs" in wb.sheetnames:
        for r in wb["Tariffs"].iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None:
                continue
            oname = _s(r[0], 300)
            region = _s(r[1] if len(r) > 1 else None, 120)
            if not (oname and region):
                continue
            prices = {
                "CCS2": _num_or_none(r[2] if len(r) > 2 else None),
                "TYPE2": _num_or_none(r[3] if len(r) > 3 else None),
                "TYPE1": _num_or_none(r[4] if len(r) > 4 else None),
            }
            matrix.setdefault(oname, {})[region] = {k: v for k, v in prices.items() if v is not None}
    return {"orgs": orgs, "matrix": matrix}


async def enrich_sessions_with_orgs(
    db: AsyncSession, company_id, parsed: dict[str, Any], channel_id=None,
) -> dict[str, Any]:
    """Обогатить ЮЛ-сессии по справочнику: наименование клиента + договорной тариф
    и выручка (energy_kwh × тариф). Идемпотентно (UPDATE по телефону).

    Режимы тарифа (у ЮЛ session.amount=0 — постоплата, выручку считаем тут):
      • flat   — плоская ставка на все сессии клиента;
      • matrix — розница как база, затем переопределение по матрице регион×коннектор
        (DC-коннекторы → цена CCS2; ячейки без цены остаются розницей);
      • retail — розничный тариф самой сессии.
    """
    S = ChargeSession
    orgs = parsed.get("orgs", []) if isinstance(parsed, dict) else (parsed or [])
    matrix = parsed.get("matrix", {}) if isinstance(parsed, dict) else {}

    # Реестр корп-клиентов: полная замена из справочника (справочник = актуальная
    # картина). Включаем ВСЕХ (даже без сессий) — для клиент-центричных табов.
    # Перед заменой запоминаем id организации из витрины АСУиМ: он приходит ДРУГИМ
    # каналом (organizations_ru), в этом справочнике его нет, и пересоздание реестра
    # стирало точную связку ЮЛ, откатывая её к сопоставлению по телефону и имени.
    prev = (await db.execute(select(CorporateClient).where(
        CorporateClient.company_id == company_id,
        CorporateClient.ext_id.is_not(None)))).scalars().all()
    ext_by_phone = {c.phone: c.ext_id for c in prev if c.phone}
    ext_by_name = {(c.name or "").strip().lower(): c.ext_id for c in prev if c.name}

    await db.execute(delete(CorporateClient).where(CorporateClient.company_id == company_id))
    for o in orgs:
        if not (o.get("phone") and o.get("name")):
            continue
        db.add(CorporateClient(
            company_id=company_id, name=o["name"], phone=o["phone"],
            ext_id=(o.get("ext_id") or ext_by_phone.get(o["phone"])
                    or ext_by_name.get(o["name"].strip().lower())),
            mode=(o.get("mode") or "retail"), rate=o.get("rate"),
            matrix=matrix.get(o["name"]) or None, discount_pct=float(o.get("discount") or 0),
            contract_start=o.get("contract_start"), status=o.get("status"), users=o.get("users"),
        ))
    await db.flush()
    await _sync_corporate_contracts(db, company_id, orgs)

    named, priced, matched = await _apply_org_enrichment(db, company_id, orgs, matrix, channel_id)
    await rebuild_mart(db, company_id)  # витрина L3 — атомарно с данными (коммит в bump)
    await bump_version(db, company_id)  # обогащение меняет выручку → сброс кеша
    return await _enrichment_result(db, company_id, len(orgs), named, priced, matched)


# Предмет исходящего договора с корпоративным клиентом ЭЗС (он же ключ идемпотентности
# вместе с контрагентом: номера в справочнике нет). Формулировка — та, что уже заведена
# в реестре руками (КЗ-2025/061 с «Ситидрайв»): второй тип про то же самое развёл бы
# один предмет на две терминологии.
CORP_CONTRACT_TYPE = "Корпоративная зарядка (постоплата)"


async def _sync_corporate_contracts(db: AsyncSession, company_id, orgs: list[dict]) -> None:
    """Завести корпоративных клиентов в общий реестр контрагентов и договоров.

    База договоров одна на компанию, и в ней должны быть обе стороны. Входящие (аренда,
    энергоснабжение, обслуживание) приходят из реестров РусГидро, исходящие — отсюда:
    корпоративный клиент ЭЗС ездит по договору постоплаты, и до сих пор этот договор
    существовал только как `contract_start` в его карточке.

    Номера в выгрузке нет — договор заводится «б/н» с датой начала; уточнённые номер,
    срок и охват проставляются в «Контрагентах» и повторной загрузкой не затираются.
    """
    if not orgs:
        return
    cps = (await db.execute(select(Counterparty).where(
        Counterparty.company_id == company_id))).scalars().all()
    by_norm = {n: c for c in cps if (n := _normname(c.name))}
    existing = (await db.execute(select(Contract).where(
        Contract.company_id == company_id,
        Contract.type == CORP_CONTRACT_TYPE))).scalars().all()
    by_cp = {c.counterparty_id: c for c in existing}
    org_id = (await db.execute(select(Organization.id).where(
        Organization.company_id == company_id).limit(1))).scalar()

    cp_by_phone: dict[str, Counterparty] = {}
    for o in orgs:
        name = str(o.get("name") or "").strip()
        nn = _normname(name)
        if not nn:
            continue
        cp = by_norm.get(nn)
        if cp is None:
            cp = Counterparty(
                company_id=company_id, inn=str(o.get("inn") or "")[:20],
                name=_clean_cp_name(name)[:500], type=_cp_type(name),
                aliases=["corporate"], kind="external",
            )
            db.add(cp)
            await db.flush()
            by_norm[nn] = cp
        else:
            # Одно юрлицо может быть и покупателем, и поставщиком — роль в реестре
            # прикладная, поэтому она копится в aliases, а не заводит вторую карточку.
            if "corporate" not in (cp.aliases or []):
                cp.aliases = sorted(set((cp.aliases or []) + ["corporate"]))
            # Карточку могли завести без ИНН (например сопоставлением ролей — там его
            # взять неоткуда). Справочник ИНН знает, и это единственное место, где он
            # доезжает: иначе «без ИНН» осталось бы навсегда.
            if not cp.inn and o.get("inn"):
                cp.inn = str(o["inn"])[:20]
        cp_by_phone[str(o.get("phone") or "")] = cp
        if any(by_cp.get(k) for k in (str(cp.id), cp.external_ref) if k):
            continue
        contract = Contract(
            company_id=company_id, number="б/н", date=str(o.get("contract_start") or ""),
            counterparty_id=str(cp.id), organization_id=str(org_id or cp.id),
            type=CORP_CONTRACT_TYPE, kind="СПокупателем", basis="договор",
            # Зарядка идёт по всей сети, а не по набору станций.
            scope_type="company",
        )
        db.add(contract)
        by_cp[str(cp.id)] = contract

    # Ссылка клиента на карточку контрагента: реестр клиентов пересоздаётся при каждой
    # загрузке справочника, поэтому связь ставится здесь, а не разовым сопоставлением —
    # иначе следующая загрузка снова оставила бы сведение по имени.
    for phone, cp in cp_by_phone.items():
        if phone:
            await db.execute(update(CorporateClient).where(
                CorporateClient.company_id == company_id, CorporateClient.phone == phone,
            ).values(counterparty_id=cp.id))
    await db.flush()


async def _apply_org_enrichment(db, company_id, orgs, matrix, channel_id) -> tuple[int, int, int]:
    """Применить обогащение к сессиям: client_name + договорной тариф/выручка по
    режиму. Идемпотентно (UPDATE по телефону). Реестр НЕ трогает."""
    S = ChargeSession

    def where(phone):
        # Обогащаем только постоплатные сессии (amount=0) либо явно ЮЛ-помеченные:
        # у телефона организации бывают розничные сессии с реальным списанием —
        # подменять живую оплату расчётной договорной суммой (и уводить сессию
        # из розницы через client_name) нельзя.
        conds = [S.company_id == company_id, S.user_id == phone,
                 or_(S.user_type == "ЮЛ", func.coalesce(S.amount, 0) == 0)]
        if channel_id is not None:
            conds.append(S.channel_id == channel_id)
        return conds

    named = priced = matched = 0
    for o in orgs:
        phone, name = o.get("phone"), o.get("name")
        if not (phone and name):
            continue
        mode, rate = (o.get("mode") or ""), o.get("rate")
        disc = float(o.get("discount") or 0)  # скидка к рознице, % (каршеринг = 25)

        n = int((await db.execute(update(S).where(*where(phone)).values(client_name=name))).rowcount or 0)
        named += n
        if n:
            matched += 1

        if mode == "flat" and rate is not None:
            r = await db.execute(update(S).where(*where(phone)).values(
                client_tariff=rate, client_amount=S.energy_kwh * rate))
            priced += int(r.rowcount or 0)
        elif mode == "retail":
            # розница сессии × (1 − скидка%): биллинг ПК для каршеринга — ровно так
            k = 1 - disc / 100
            r = await db.execute(update(S).where(*where(phone)).values(
                client_tariff=S.tariff * k, client_amount=S.energy_kwh * S.tariff * k))
            priced += int(r.rowcount or 0)
        elif mode == "matrix":
            r = await db.execute(update(S).where(*where(phone)).values(
                client_tariff=S.tariff, client_amount=S.energy_kwh * S.tariff))
            priced += int(r.rowcount or 0)
            for region, prices in matrix.get(name, {}).items():
                for col, conns in _MATRIX_CONN.items():
                    price = prices.get(col)
                    if price is None:
                        continue
                    # Матрица = РОЗНИЦА региона; договорная цена = розница × (1 − скидка%).
                    eff = round(price * (1 - disc / 100), 4)
                    await db.execute(update(S).where(
                        *where(phone), S.region == region, S.connector_type.in_(conns),
                    ).values(client_tariff=eff, client_amount=S.energy_kwh * eff))
    return named, priced, matched


async def _enrichment_result(db, company_id, orgs_count, named, priced, matched) -> dict[str, Any]:
    revenue = float((await db.execute(select(func.coalesce(func.sum(ChargeSession.client_amount), 0)).where(
        ChargeSession.company_id == company_id, ChargeSession.client_amount.is_not(None)))).scalar_one() or 0)
    return {"status": "success", "orgs": orgs_count, "orgs_matched": matched,
            "updated": named, "priced": priced, "revenue": round(revenue, 2),
            "message": f"обогащено {named} сессий ({matched} орг), корп-выручка {round(revenue):,} ₽".replace(",", " ")}


async def enrich_from_registry(db: AsyncSession, company_id, channel_id=None) -> dict[str, Any]:
    """Переприменить обогащение из СОХРАНЁННОГО реестра corporate_clients (без файла).
    Восстанавливает client_name/тариф/выручку после перезагрузки таблицы сессий."""
    clients = (await db.execute(
        select(CorporateClient).where(CorporateClient.company_id == company_id))).scalars().all()
    if not clients:
        return {"status": "skipped", "orgs": 0, "orgs_matched": 0, "updated": 0, "priced": 0,
                "revenue": 0, "message": "реестр корп-клиентов пуст — сначала загрузите справочник «Организации»"}
    orgs = [{"phone": c.phone, "name": c.name, "mode": c.mode,
             "rate": float(c.rate) if c.rate is not None else None,
             "discount": float(c.discount_pct or 0)}
            for c in clients]
    matrix = {c.name: c.matrix for c in clients if c.matrix}
    named, priced, matched = await _apply_org_enrichment(db, company_id, orgs, matrix, channel_id)
    await rebuild_mart(db, company_id)  # витрина L3 — атомарно с данными (коммит в bump)
    await bump_version(db, company_id)  # обогащение меняет выручку → сброс кеша
    return await _enrichment_result(db, company_id, len(orgs), named, priced, matched)


async def enrich_session_cards(db: AsyncSession, company_id) -> dict[str, Any]:
    """Проставить сессиям данные их RFID-карты (номер, статус, владелец).

    В выгрузке сессия несёт от карты только UID вида «26AA5969». По нему нельзя
    ни отобрать заправки по заблокированной карте, ни назвать владельца, ни
    сгруппировать реестр — поэтому справочник витрины разворачивается в саму
    строку сессии, как это уже сделано для корпоративных клиентов.

    Отдельно помечаются заправки НЕ владельцем карты: телефон карты не совпал с
    телефоном сессии. Карта, которой заряжается кто-то другой, — законный случай
    (служебная машина, каршеринг), но его надо видеть списком.

    Идемпотентно: повторный запуск переписывает те же значения. Вызывается после
    загрузки сессий и после загрузки карт — свежими могут оказаться и те, и те.
    """
    res = await db.execute(text("""
        update charge_sessions s
           set card_number       = c.number,
               card_status       = c.status,
               card_owner_ext_id = c.customer_ext_id,
               card_foreign      = (c.phone is not null and s.user_id is not null
                                    and c.phone <> s.user_id)
          from ezs_rfid_cards c
         where c.company_id = s.company_id
           and c.uid = s.rfid
           and s.company_id = :cid
           and s.rfid is not null and s.rfid <> ''
           and (s.card_number is distinct from c.number
                or s.card_status is distinct from c.status
                or s.card_owner_ext_id is distinct from c.customer_ext_id
                or s.card_foreign is distinct from (c.phone is not null and s.user_id is not null
                                                    and c.phone <> s.user_id))
    """), {"cid": str(company_id)})
    updated = res.rowcount or 0

    stat = (await db.execute(text("""
        select count(*) filter (where rfid is not null and rfid <> '')      as by_card,
               count(*) filter (where card_number is not null)              as enriched,
               count(*) filter (where card_foreign)                          as foreign_use,
               count(*) filter (where card_status is not null
                                  and card_status not in ('Активная'))       as not_active
          from charge_sessions where company_id = :cid
    """), {"cid": str(company_id)})).one()

    msg = (f"карты в сессиях: обогащено {int(stat.enriched)} из {int(stat.by_card)}"
           f"; заправок не владельцем {int(stat.foreign_use)}")
    if stat.not_active:
        msg += f"; по неактивным картам {int(stat.not_active)}"
    logger.info("enrich_session_cards: обновлено %s строк; %s", updated, msg)
    return {"status": "success", "updated": updated,
            "by_card": int(stat.by_card), "enriched": int(stat.enriched),
            "foreign": int(stat.foreign_use), "not_active": int(stat.not_active),
            "message": msg}
