"""Акт сверки: результат сопоставления 1С ↔ TradeLedger одним файлом.

Сверка считалась и показывалась только на экране, а контрагенту и бухгалтерии
нужен документ. По PLAN.md акты сверки делают «раз в квартал вручную», и
ручная сверка оценена в 10 000–30 000 ₽/месяц — этот файл закрывает ровно то
место, где человек перекладывает расхождения в Excel руками.

Четыре листа отвечают на четыре разных вопроса, поэтому разделены:
  Сводка          — сколько и на какую сумму (для руководителя);
  Расхождения     — что именно не сходится и насколько (рабочий лист);
  Нет в TradeLedger — 1С знает документ, у нас его нет (недогруз первички);
  Нет в 1С        — у нас есть, в 1С не проведено (недоработка бухгалтерии).

Последние два — разные проблемы с разными адресатами, сваливать их в один
список нельзя: их разбирают разные люди.
"""
from __future__ import annotations

import uuid
from datetime import datetime

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import AccountingDoc, DataEntry
from app.services.reconciliation_service import entry_meta_amount

# Человеческие названия статусов. В БД лежат англоязычные ключи, но файл уходит
# бухгалтеру и контрагенту — там «minor» ничего не значит.
SEVERITY_LABEL = {
    "none": "Совпадает",
    "rounding": "Округление",
    "minor": "Незначительное",
    "material": "Существенное",
    "critical": "Критическое",
    "unmatched": "Нет пары",
    "pending": "Не сверялось",
}

# Порядок разбора: сначала то, что требует действий. Сортировка по этому
# порядку, а не по дате — акт читают сверху вниз и бросают, когда кончается
# важное.
SEVERITY_ORDER = {
    "critical": 0, "material": 1, "minor": 2, "rounding": 3,
    "unmatched": 4, "pending": 5, "none": 6,
}

# Расхождения, которые попадают в рабочий лист. «none» и «pending» — не
# расхождения: первое сошлось, второе ещё не сверялось.
DISCREPANCY_STATUSES = ("rounding", "minor", "material", "critical")

_HEAD_FILL = "FFEFEFEF"
# Заливка строк по важности. Приглушённая — файл печатают, а ядовитый фон
# съедает тонер и мешает читать текст.
_ROW_FILL = {
    "critical": "FFFFD6D6",
    "material": "FFFFE8CC",
    "minor": "FFFFF6CC",
}


def _head(ws, columns: list[str], widths: list[int]) -> None:
    """Шапка таблицы: жирная, залитая, с заморозкой и автофильтром.

    Координаты задаём строками («A5»), а не через ws.cell(): обращение к ячейке
    в openpyxl её СОЗДАЁТ, и лист получал лишнюю пустую строку под шапкой.
    """
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws.append(columns)
    head_row = ws.max_row
    for cell in ws[head_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=_HEAD_FILL)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{head_row + 1}"
    ws.auto_filter.ref = f"A{head_row}:{get_column_letter(len(columns))}{head_row}"


async def build_reconciliation_act(
    db: AsyncSession,
    company_id: uuid.UUID,
    company_name: str,
    date_from: str | None = None,
    date_to: str | None = None,
    limit: int = 20000,
):
    """Собрать акт сверки. Возвращает (Workbook, статистика).

    date_from/date_to — ISO «YYYY-MM-DD». AccountingDoc.date хранится строкой в
    том же формате, поэтому сравнение строковое и границы включительны сами по
    себе (грабля с DateTime и отсечением последнего дня здесь не возникает).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    period = [AccountingDoc.company_id == company_id]
    if date_from:
        period.append(AccountingDoc.date >= date_from[:10])
    if date_to:
        period.append(AccountingDoc.date <= date_to[:10])

    docs = (await db.execute(
        select(AccountingDoc).where(*period).order_by(AccountingDoc.date)
    )).scalars().all()

    matched_ids = {d.matched_entry_id for d in docs if d.matched_entry_id}

    # Записи TradeLedger отбираем по ДАТЕ ДОКУМЕНТА (meta.docDate), а не по дате
    # загрузки: ТТН от 2019 года могла попасть в систему вчера, и по created_at
    # она уехала бы в акт за текущий месяц. Где docDate нет — падаем на
    # created_at, иначе запись потерялась бы из акта совсем.
    entry_day = func.coalesce(
        DataEntry.meta["docDate"].astext,
        func.to_char(DataEntry.created_at, "YYYY-MM-DD"),
    )
    entry_where = [DataEntry.company_id == company_id]
    if date_from:
        entry_where.append(entry_day >= date_from[:10])
    if date_to:
        entry_where.append(entry_day <= date_to[:10])

    entries = (await db.execute(
        select(DataEntry).where(*entry_where).order_by(DataEntry.created_at.desc())
    )).scalars().all()
    entry_by_id = {e.id: e for e in entries}

    # Пары для листа «Расхождения» ищем без периодного фильтра: документ 1С из
    # периода может быть сопоставлен с записью, чей docDate лёг за границу, и
    # тогда сумма «по TradeLedger» в акте молча стала бы пустой.
    missing = {i for i in matched_ids if i and i not in entry_by_id}
    if missing:
        for e in (await db.execute(
            select(DataEntry).where(DataEntry.id.in_(missing))
        )).scalars().all():
            entry_by_id[e.id] = e

    discrepancies = [d for d in docs if d.discrepancy_status in DISCREPANCY_STATUSES]
    discrepancies.sort(key=lambda d: (SEVERITY_ORDER.get(d.discrepancy_status, 9), d.date))
    unmatched_1c = [d for d in docs if d.discrepancy_status == "unmatched"]
    unmatched_cl = [e for e in entries if e.id not in matched_ids]
    matched_ok = [d for d in docs if d.discrepancy_status == "none"]

    wb = openpyxl.Workbook()

    # ── Лист 1: Сводка ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "АКТ СВЕРКИ ДАННЫХ"
    ws["A1"].font = Font(bold=True, size=15)
    ws["A2"] = f"{company_name} · TradeLedger ↔ 1С:Бухгалтерия"
    ws["A2"].font = Font(color="FF666666")

    span = "весь доступный период"
    if date_from or date_to:
        span = f"{date_from[:10] if date_from else '…'} — {date_to[:10] if date_to else '…'}"
    ws["A3"] = f"Период: {span}"
    ws["A4"] = f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A4"].font = Font(color="FF666666")

    ws.append([])
    ws.append(["Показатель", "Документов", "Сумма по 1С, ₽"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=_HEAD_FILL)

    def _sum(rows) -> float:
        return round(sum(float(d.amount or 0) for d in rows), 2)

    ws.append(["Сошлось без расхождений", len(matched_ok), _sum(matched_ok)])
    ws.append(["Есть расхождения", len(discrepancies), _sum(discrepancies)])
    ws.append(["Нет пары в TradeLedger", len(unmatched_1c), _sum(unmatched_1c)])
    ws.append(["Нет пары в 1С", len(unmatched_cl), None])
    ws.append(["Всего документов 1С в периоде", len(docs), _sum(docs)])

    ws.append([])
    ws.append(["Расхождения по важности", "Документов", "Сумма по 1С, ₽"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=_HEAD_FILL)
    for sev in ("critical", "material", "minor", "rounding"):
        rows = [d for d in discrepancies if d.discrepancy_status == sev]
        if rows:
            ws.append([SEVERITY_LABEL[sev], len(rows), _sum(rows)])
            if sev in _ROW_FILL:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill("solid", fgColor=_ROW_FILL[sev])

    not_checked = [d for d in docs if d.discrepancy_status == "pending"]
    if not_checked:
        ws.append([])
        ws.append([f"⚠ Не сверялось: {len(not_checked)} документов — "
                   f"запустите сверку, иначе акт неполон"])
        ws[ws.max_row][0].font = Font(italic=True, color="FFB8760A")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20

    # ── Лист 2: Расхождения ───────────────────────────────────────────────
    ws = wb.create_sheet("Расхождения")
    _head(ws,
          ["Важность", "Что расходится", "Тип", "Номер 1С", "Дата", "Контрагент",
           "ИНН", "Сумма по 1С, ₽", "Сумма по TradeLedger, ₽", "Разница, ₽",
           "НДС по 1С, ₽", "Документ TradeLedger", "Период"],
          [16, 42, 10, 20, 12, 34, 14, 18, 22, 14, 14, 40, 12])
    for d in discrepancies:
        entry = entry_by_id.get(d.matched_entry_id) if d.matched_entry_id else None
        our = entry_meta_amount(entry) if entry else None
        delta = round(float(d.amount or 0) - float(our), 2) if our is not None else None
        ws.append([
            SEVERITY_LABEL.get(d.discrepancy_status, d.discrepancy_status),
            d.discrepancy_summary or "",
            d.doc_type, d.number, d.date,
            d.counterparty_name, d.counterparty_inn,
            round(float(d.amount or 0), 2), our, delta,
            round(float(d.vat_amount), 2) if d.vat_amount is not None else None,
            entry.title if entry else "",
            "закрыт" if d.period_status == "closed" else "открыт",
        ])
        fill = _ROW_FILL.get(d.discrepancy_status)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor=fill)

    # ── Лист 3: документы 1С без пары ─────────────────────────────────────
    ws = wb.create_sheet("Нет в TradeLedger")
    ws.append(["Документ проведён в 1С, но первичного документа в TradeLedger нет — "
               "проверьте, загружен ли он"])
    ws[1][0].font = Font(italic=True, color="FF666666")
    ws.append([])
    _head(ws, ["Тип", "Номер", "Дата", "Контрагент", "ИНН", "Организация",
               "Сумма, ₽", "НДС, ₽", "Статус в 1С", "Склад"],
          [10, 20, 12, 34, 14, 30, 16, 14, 16, 16])
    shown_1c = unmatched_1c[:limit]
    for d in shown_1c:
        ws.append([d.doc_type, d.number, d.date, d.counterparty_name, d.counterparty_inn,
                   d.organization_name, round(float(d.amount or 0), 2),
                   round(float(d.vat_amount), 2) if d.vat_amount is not None else None,
                   d.status_1c, d.warehouse_code])
    if len(unmatched_1c) > len(shown_1c):
        ws.append([])
        ws.append([f"⚠ Показаны первые {len(shown_1c)} из {len(unmatched_1c)} — "
                   f"сузьте период, чтобы увидеть остальные"])
        ws[ws.max_row][0].font = Font(italic=True, color="FFB8760A")

    # ── Лист 4: записи TradeLedger без пары ───────────────────────────────
    ws = wb.create_sheet("Нет в 1С")
    ws.append(["Документ есть в TradeLedger, но в 1С не найден — "
               "проверьте, проведён ли он бухгалтерией"])
    ws[1][0].font = Font(italic=True, color="FF666666")
    ws.append(["Отбор по дате документа; где она не заполнена — по дате загрузки"])
    ws[2][0].font = Font(italic=True, color="FF999999")
    ws.append([])
    _head(ws, ["Название", "Категория", "Тип документа", "Статус", "Источник",
               "Дата документа", "Сумма, ₽", "Загружен"],
          [46, 22, 20, 16, 18, 16, 16, 18])
    shown_cl = unmatched_cl[:limit]
    for e in shown_cl:
        ws.append([e.title, e.category_id, e.doc_type_id or "", e.status,
                   e.source_label or e.source, (e.meta or {}).get("docDate", ""),
                   entry_meta_amount(e),
                   e.created_at.strftime("%d.%m.%Y %H:%M") if e.created_at else ""])
    if len(unmatched_cl) > len(shown_cl):
        ws.append([])
        ws.append([f"⚠ Показаны первые {len(shown_cl)} из {len(unmatched_cl)}"])
        ws[ws.max_row][0].font = Font(italic=True, color="FFB8760A")

    stats = {
        "matched": len(matched_ok),
        "discrepancies": len(discrepancies),
        "unmatched_1c": len(unmatched_1c),
        "unmatched_cl": len(unmatched_cl),
        "total_docs": len(docs),
        "not_checked": len(not_checked),
    }
    return wb, stats
