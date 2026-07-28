"""Банк ЗУ — площадки под установку ЭЗС (девелоперский пайплайн).

Учёт МЕСТ, где сеть строится. Модель — воронка подбора недвижимости с гейтами:
lead → screening → negotiation → dd → decision → contracting → construction → live,
плюс on_hold (заморожена) и archive (отклонена, с причиной). Разбор и обоснование —
`docs/SITES_LAND_BANK_BLUEPRINT.md`.

Источник — сводный Excel «Банк данных ЗУ» (3 листа, ~55 колонок). Ключевые поля
кладём в колонки (фильтры/агрегаты), полный исходный ряд — в raw JSONB.

**Импорт — UPSERT, не REPLACE-ALL.** У площадки есть идентичность: кадастровый
номер → координаты (радиус ~50 м) → нормализованный адрес + город. Файл дополняет
пустые поля и двигает стадию вперёд, но не затирает уже известное и не откатывает
площадку назад по воронке. Прежний REPLACE-ALL стирал историю: каждая загрузка
обнуляла состояние, накопить движение по стадиям было невозможно.

Экономические поля в файле разнородны («800 т.р.», «9р.», «по запросу») —
парсим best-effort, при неудаче NULL; агрегаты по деньгам подписываем «по
распознанным».
"""
from __future__ import annotations

import hashlib
import math
import re
from datetime import date, datetime, timedelta, timezone
from typing import Any

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import EzsSite, Region

# ── Воронка ────────────────────────────────────────────────────────────────
# Порядок = порядок гейтов. Импорт может двигать площадку только ВПЕРЁД по
# этому списку (кроме архива) — иначе файл откатывал бы ручную работу.
STAGE_ORDER = [
    "lead", "screening", "negotiation", "dd", "decision",
    "contracting", "construction", "commissioning", "live",
]
STAGE_LABELS = {
    "lead": "Лид",
    "screening": "Скрининг",
    "negotiation": "Переговоры",
    "dd": "Проработка",
    "decision": "Решение",
    "contracting": "Оформление земли",
    "construction": "Реализация",
    "commissioning": "Пусконаладка",
    "live": "В эксплуатации",
    "on_hold": "Заморожен",
    "archive": "Архив",
}
STAGE_HINTS = {
    "lead": "адрес и источник",
    "screening": "быстрый отсев без затрат",
    "negotiation": "выход на собственника, условия",
    "dd": "ТУ · право · коммерция",
    "decision": "экономика и вердикт",
    "contracting": "договор / сервитут / разрешение",
    "construction": "присоединение ‖ оборудование ‖ монтаж",
    "commissioning": "пусконаладка и приёмка",
    "live": "объект работает в сети",
    "on_hold": "пауза с датой пересмотра",
    "archive": "отклонён, с причиной",
}
ALL_STAGES = STAGE_ORDER + ["on_hold", "archive"]

# Этапы проекта — крупные блоки, которыми мыслит руководитель. Стадии внутри
# этапа последовательны, а вот работы внутри «Реализации» (присоединение,
# закупка оборудования, монтаж) идут ПАРАЛЛЕЛЬНО: оборудование заказывают, не
# дожидаясь исполнения ТП. Поэтому параллельность живёт не в стадиях, а в
# отдельных треках (см. docs/SITES_PROJECT_LIFECYCLE.md).
PHASES = [
    {"key": "select", "label": "Подбор", "hint": "скрининг, переговоры, экономика, решение",
     "stages": ["lead", "screening", "negotiation", "dd", "decision"]},
    {"key": "land", "label": "Земля", "hint": "аренда / сервитут / разрешение",
     "stages": ["contracting"]},
    {"key": "build", "label": "Реализация", "hint": "присоединение, оборудование, монтаж, ПНР",
     "stages": ["construction", "commissioning"]},
    {"key": "operate", "label": "Эксплуатация", "hint": "объект работает в сети",
     "stages": ["live"]},
    {"key": "closed", "label": "Не в работе", "hint": "заморожен или отклонён",
     "stages": ["on_hold", "archive"]},
]
STAGE_PHASE = {s: p["key"] for p in PHASES for s in p["stages"]}
PHASE_LABELS = {p["key"]: p["label"] for p in PHASES}
_STAGE_POS = {s: i for i, s in enumerate(STAGE_ORDER)}
# Сколько дней без касания считаем «площадка забыта».
STALE_DAYS = 30

# Лист Excel → базовая стадия. Для «ЗУ в работе» стадия уточняется по
# заполненности строки (_stage_for_row): лист смешивает лид и переговоры.
SHEET_STAGE = {
    "зу в работе": "negotiation",
    "зу (архив)": "archive",
    "зу архив": "archive",
    # Договор на подписи у собственника, идёт отбор подрядчика — это конец
    # воронки, а не начало. Раньше лист падал в «В проработке».
    "согласованные": "contracting",
}


def _norm(s: Any) -> str:
    return " ".join(str(s or "").strip().lower().split())


# Заголовок (нормализованный) → поле модели. Порядок ВАЖЕН: специфичные —
# раньше общих («статус согласования» до «статус»).
def _match_field(h: str) -> str | None:
    checks: list[tuple[str, tuple[str, ...]]] = [
        ("received_date", ("дата поступления", "дата прихода")),
        ("ownership", ("статус зу", "собственность/аренда")),
        ("tu_status", ("статус согласования",)),
        ("status_raw", ("статус",)),
        ("place_kind", ("признак",)),
        ("install_place", ("место установки",)),
        ("full_address", ("полный адрес",)),
        ("region", ("регион",)),
        ("city", ("город",)),
        ("route", ("трасса",)),
        ("coords_raw", ("координат",)),
        ("map_url", ("ссылка на карт",)),
        ("owner_contact", ("контакт представителя собственника",)),
        ("source_company", ("предоставившего зу - компания", "предоставившего зу — компания",
                            "предоставивший зу - компания", "предоставивший зу — компания")),
        ("source_person", ("предоставившего зу - фио", "предоставившего зу — фио",
                           "предоставивший зу - фио", "предоставивший зу — фио")),
        ("owner", ("собственник",)),
        ("brand", ("бренд",)),
        ("area_m2", ("площадь",)),
        ("free_power_kwt", ("свободная мощность",)),
        ("rent_cost_month", ("стоимость аренды",)),
        ("input_price_kwth", ("входная стоимость",)),
        ("tp_cost", ("стоимость техприса",)),
        ("smr_cost", ("стоимость смр",)),
        # Порядок важен: «итого затраты на подключение (ТУ)» — отдельная графа
        # этапа 5, а не то же самое, что затраты этапа 3 по техпрису.
        ("tu_total_cost", ("итого затраты на подключение (ту)",)),
        ("connection_cost", ("итого затраты на подключ",)),
        ("distance_to_tp_m", ("расстояние до тп",)),
        ("tp_reconstruction", ("реконструкция тп",)),
        ("long_term_contract", ("долгосрочный договор",)),
        ("has_video", ("видеонаблюдение",)),
        ("access_24x7", ("свободный доступ",)),
        ("has_mobile", ("сотовая связь",)),
        ("planned_power_kwt", ("мощность эзс к установке",)),
        ("planned_ezs_count", ("кол-во эзс",)),
        ("ports_gbt", ("порты эзс - gbt", "порты эзс gbt")),
        ("ports_ccs", ("порты эзс - ccs", "порты эзс ccs")),
        ("ports_chademo", ("порты эзс - chademo", "порты эзс chademo")),
        ("ports_type", ("порты эзс - type", "порты эзс type")),
        ("supplier", ("поставщик",)),
        ("contractor", ("подрядчик",)),
        ("tech_conn_type", ("тип технологического присоед",)),
        ("substation_owner", ("принадлежность подстанции",)),
        ("line_owner", ("принадлежность кл",)),
        ("transformer_kva", ("мощность силового трансформатора",)),
        ("line_type", ("тип и сечение кл",)),
        ("extra_power_possible", ("возможность доп",)),
        ("transformer_swap_possible", ("возможность замены трансформатора",)),
        ("tu_contract_cost", ("стоимость договора с электросетевой",)),
        ("tu_works_cost", ("стоимость мероприятий общества",)),
        ("tp_term_months", ("сроки мероприятий электросетевой",)),
        ("applicant_term_months", ("сроки мероприятий заявителя",)),
        ("dop_service", ("доп.сервис", "доп сервис")),
        ("archive_reason", ("причина согласия", "причина отказа")),
        ("cadastral_no", ("кадастров",)),
        ("comment", ("комментарий",)),
        ("address", ("адрес",)),   # после «полный адрес»
    ]
    for field, keys in checks:
        if any(h.startswith(k) or k in h for k in keys):
            return field
    return None


def _num(v: Any) -> float | None:
    """Число из разнородной строки: «800 т.р.»→800000, «9р.»→9, «по запросу»→None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lower().replace("\xa0", " ")
    if not s or s in ("-", "—", "по запросу", "н/д", "нет данных", "нет"):
        return None
    mult = 1.0
    if "млн" in s:
        mult = 1e6
    elif "т.р" in s or "тыс" in s or "т руб" in s:
        mult = 1e3
    m = re.search(r"[-+]?\d[\d ]*(?:[.,]\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0).replace(" ", "").replace(",", ".")) * mult
    except ValueError:
        return None


def _int(v: Any) -> int | None:
    n = _num(v)
    return int(round(n)) if n is not None else None


def _bool(v: Any) -> bool | None:
    """«да»/«есть»/«+» → True, «нет»/«отсутствует»/«-» → False, прочее → None.

    В графах вида «Свободный доступ (да/нет)» пишут и «для электромобилей», и
    «?». Такое значение не «нет» — это «не выяснено», иначе площадка молча
    получит отказной признак и уедет из выборки.
    """
    t = _norm(v)
    if not t:
        return None
    if t.startswith(("да", "есть", "имеет", "+")):
        return True
    if t.startswith(("нет", "отсут", "-", "не ")):
        return False
    return None


def _coords(v: Any) -> tuple[float | None, float | None]:
    """(lat, lon) из «56.83, 60.59» / «59.57 150.80». DMS/мусор → (None, None)."""
    if v is None:
        return (None, None)
    nums = re.findall(r"[-+]?\d{1,3}[.,]\d{3,}", str(v))
    if len(nums) >= 2:
        try:
            lat = float(nums[0].replace(",", ".")); lon = float(nums[1].replace(",", "."))
        except ValueError:
            return (None, None)
        if 40 <= lat <= 82 and 18 <= lon <= 190:   # правдоподобный охват РФ
            return (lat, lon)
    return (None, None)


def _s(v: Any, limit: int | None = None) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return s[:limit] if limit else s


# ── Нормализация региона ───────────────────────────────────────────────────
# В файле 104 разных значения на 85 субъектов: города вместо регионов («Тюмень»),
# инверсия («Карелия республика»), регистр («Пермский Край»). Аналитика ЭЗС
# считает по справочнику `regions` — площадки должны ложиться на тот же ключ.
_REGION_NOISE = re.compile(
    r"\b(область|обл|край|республика|респ|автономн\w*|округ|ао|г|город)\b\.?", re.I)

# Города, которыми в файле подменяют субъект (кроме городов федерального значения).
_CITY_TO_REGION = {
    "спб": "Санкт-Петербург",
    "санкт петербург": "Санкт-Петербург",
    "питер": "Санкт-Петербург",
    "мск": "Москва",
    "тюмень": "Тюменская область",
    "курган": "Курганская область",
    "в новгород": "Новгородская область",
    "великий новгород": "Новгородская область",
    "н новгород": "Нижегородская область",
    "нижний новгород": "Нижегородская область",
    "екатеринбург": "Свердловская область",
    "владивосток": "Приморский край",
    "хабаровск": "Хабаровский край",
    "чита": "Забайкальский край",
    "пермь": "Пермский край",
    "казань": "Республика Татарстан",
    "уфа": "Республика Башкортостан",
    "самара": "Самарская область",
    "саратов": "Саратовская область",
    "ростов на дону": "Ростовская область",
    "краснодар": "Краснодарский край",
    "воронеж": "Воронежская область",
    "волгоград": "Волгоградская область",
    "новосибирск": "Новосибирская область",
    "омск": "Омская область",
    "красноярск": "Красноярский край",
    "иркутск": "Иркутская область",
    "барнаул": "Алтайский край",
    "кемерово": "Кемеровская область",
    "томск": "Томская область",
    "челябинск": "Челябинская область",
    "магадан": "Магаданская область",
    "южно сахалинск": "Сахалинская область",
    "петрозаводск": "Республика Карелия",
    "мурманск": "Мурманская область",
    "псков": "Псковская область",
    "тверь": "Тверская область",
    "смоленск": "Смоленская область",
    "кострома": "Костромская область",
    "калуга": "Калужская область",
    "рязань": "Рязанская область",
    "тамбов": "Тамбовская область",
    "липецк": "Липецкая область",
    "белгород": "Белгородская область",
    "тула": "Тульская область",
    "владимир": "Владимирская область",
    "ярославль": "Ярославская область",
    "иваново": "Ивановская область",
    "киров": "Кировская область",
    "пенза": "Пензенская область",
    "ульяновск": "Ульяновская область",
    "оренбург": "Оренбургская область",
    "астрахань": "Астраханская область",
    "ставрополь": "Ставропольский край",
    "элиста": "Республика Калмыкия",
    "саранск": "Республика Мордовия",
    "чебоксары": "Чувашская Республика",
    "ижевск": "Удмуртская Республика",
    "благовещенск": "Амурская область",
    "биробиджан": "Еврейская автономная область",
    "улан удэ": "Республика Бурятия",
    "якутск": "Республика Саха (Якутия)",
}


def _region_key(s: str) -> str:
    """Ядро названия субъекта: «Республика Карелия»/«Карелия республика» → «карелия»."""
    t = _norm(s).replace("ё", "е")
    t = re.sub(r"[«»\"'.,()\-]", " ", t)
    t = _REGION_NOISE.sub(" ", t)
    return " ".join(t.split())


def _region_quality(name: str) -> tuple[int, int, int]:
    """Насколько имя субъекта «полное». Справочник сети засорён обрубками
    («Алтайский» рядом с «Алтайский край», «Иркутская обл» рядом с «Иркутская
    область») — на площадке должно оказаться полное название. Канцелярская
    приставка «г Москва» проигрывает чистому «Москва»."""
    t = _norm(name)
    full = 1 if re.search(r"\b(область|край|республика|округ|автономн\w*)\b", t) else 0
    clean = 0 if re.match(r"^(г|гор|город)\b\.?", t) else 1
    return (full, clean, len(t))


class RegionResolver:
    """Сопоставляет значение из файла со справочником `regions` компании."""

    def __init__(self, regions: list[Region]):
        self.by_key: dict[str, Region] = {}
        for r in regions:
            for name in (r.name, r.federal_subject):
                if not name:
                    continue
                k = _region_key(name)
                if not k:
                    continue
                cur = self.by_key.get(k)
                if cur is None or _region_quality(r.name) > _region_quality(cur.name):
                    self.by_key[k] = r
        self.unmatched: dict[str, int] = {}

    def resolve(self, raw: str | None, city: str | None = None) -> tuple[str | None, Any]:
        """(канон-имя, Region|None). Канон возвращаем даже без справочника."""
        if not raw:
            return (None, None)
        key = _region_key(raw)
        if not key:
            return (None, None)
        alias = _CITY_TO_REGION.get(key)
        if alias:
            key = _region_key(alias)
        reg = self.by_key.get(key)
        if reg is not None:
            return (reg.name, reg)
        # Справочник может не содержать субъект (сеть туда ещё не пришла) —
        # канон всё равно фиксируем, чтобы группировка не рассыпалась по регистру.
        canon = alias or str(raw).strip()
        self.unmatched[canon] = self.unmatched.get(canon, 0) + 1
        return (canon[:160], None)


# ── Идентичность площадки ──────────────────────────────────────────────────
_ADDR_NOISE = re.compile(
    r"\b(г|гор|город|ул|улица|пр|просп|проспект|ш|шоссе|д|дом|обл|область|край|"
    r"респ|республика|стр|строение|корп|корпус|мкр|микрорайон|пос|поселок|"
    r"посёлок|дер|деревня|б/н|бн)\b\.?", re.I)
_CAD_RE = re.compile(r"\b\d{2}:\d{2}:\d{6,7}:\d{1,6}\b")


def _addr_key(*parts: str | None) -> str:
    """Адресный ключ. Пустой, если адрес не идентифицирует точку.

    Требуем цифру: «Санкт-Петербург и ЛО» или «ТЦ на объездной» — это не адрес,
    а описание района, и склеивать по нему разные площадки нельзя.
    """
    t = _norm(" ".join(p for p in parts if p))
    t = re.sub(r"[«»\"'.,()№#]", " ", t)
    t = _ADDR_NOISE.sub(" ", t)
    t = " ".join(t.split())
    return t if re.search(r"\d", t) else ""


def _find_cadastral(raw: dict[str, str]) -> str | None:
    """Кадастровый номер, если он затесался в комментарий/адрес (колонки в файле нет)."""
    for v in raw.values():
        m = _CAD_RE.search(str(v))
        if m:
            return m.group(0)
    return None


def _dedup_key(cadastral: str | None, lat: float | None, lon: float | None,
               addr_key: str, fallback: str = "") -> str | None:
    if cadastral:
        return f"cad:{cadastral}"
    if lat is not None and lon is not None:
        return f"geo:{lat:.3f},{lon:.3f}"      # ячейка ~110×60 м
    if addr_key:
        # В файле встречаются ячейки с несколькими адресами сразу — режем длину,
        # иначе ключ не влезает в колонку.
        return f"adr:{addr_key[:180]}"
    if fallback:
        # Ни кадастра, ни координат, ни адреса (24 строки файла). Чтобы такие
        # площадки не создавались заново при каждой загрузке — ключ из содержимого.
        return "row:" + hashlib.sha1(fallback.encode("utf-8")).hexdigest()[:24]
    return None


def _km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    return 111.0 * math.hypot(lat1 - lat2, (lon1 - lon2) * math.cos(math.radians(lat1)))


class SiteIndex:
    """Поиск существующей площадки: кадастр → координаты (50 м) → адрес."""

    NEAR_KM = 0.05

    def __init__(self, sites: list[EzsSite]):
        self.by_cad: dict[str, EzsSite] = {}
        self.by_addr: dict[str, EzsSite] = {}
        self.by_key: dict[str, EzsSite] = {}
        self.geo: list[EzsSite] = []
        for s in sites:
            self.add(s)

    def find(self, cadastral: str | None, lat: float | None, lon: float | None,
             addr_key: str, key: str | None = None) -> tuple[EzsSite | None, str]:
        """(площадка, способ совпадения) — способ нужен, чтобы поймать конфликт
        «координаты рядом, а адреса разные»: это чаще ошибка координат, чем дубль."""
        if cadastral and cadastral in self.by_cad:
            return (self.by_cad[cadastral], "cad")
        # Точный адрес сильнее близости координат: две площадки могут стоять в
        # 50 м друг от друга (разные стороны трассы, соседние ТЦ), а вот один и
        # тот же адрес — это одна площадка, как бы ни были записаны координаты.
        if addr_key and addr_key in self.by_addr:
            return (self.by_addr[addr_key], "addr")
        if lat is not None and lon is not None:
            best, best_km = None, self.NEAR_KM
            for s in self.geo:
                d = _km(lat, lon, s.lat, s.lon)
                if d <= best_km:
                    best, best_km = s, d
            if best is not None:
                other = _addr_key(best.full_address or best.address, best.city)
                if addr_key and other and addr_key != other:
                    return (best, "geo_conflict")
                return (best, "geo")
        # Строки без адреса и координат опознаются только по ключу-хешу.
        return ((self.by_key.get(key), "key") if key else (None, ""))

    def add(self, s: EzsSite) -> None:
        if s.cadastral_no:
            self.by_cad.setdefault(s.cadastral_no, s)
        if s.lat is not None and s.lon is not None:
            self.geo.append(s)
        k = _addr_key(s.full_address or s.address, s.city)
        if k:
            self.by_addr.setdefault(k, s)
        if s.dedup_key:
            self.by_key.setdefault(s.dedup_key, s)


# ── Стадия строки ──────────────────────────────────────────────────────────
def _stage_for_row(sheet_stage: str, vals: dict[str, Any], raw: dict[str, str]) -> str:
    """Уточняет стадию внутри листа «ЗУ в работе» по тому, что уже добыто.

    Лист смешивает всё активное: от «дали адрес» до «запросили ТУ». Ставить всем
    одну стадию — значит показывать 261 площадку как один этап, которого нет.
    """
    if sheet_stage != "negotiation":
        return sheet_stage
    # Пошла техническая проработка: заметки по согласованию/ТУ, подрядчик,
    # поставщик, посчитанные затраты на подключение.
    if any(vals.get(f) for f in ("tu_status", "contractor", "supplier", "connection_cost",
                                 "tech_conn_type", "planned_power_kwt")):
        return "dd"
    # Есть выход на собственника — переговоры идут.
    contact = raw.get("Контакт представителя собственника") or vals.get("owner")
    if contact:
        return "negotiation"
    return "lead"


def _iso_date(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return s[:10]   # openpyxl отдаёт datetime для дат-ячеек


def _today() -> str:
    return date.today().isoformat()


# Поля, которые импорт может дополнять (пустое → значение из файла).
_FILL_FIELDS = (
    "status_raw", "received_date", "region", "city", "address", "full_address",
    "place_kind", "install_place", "route", "coords_raw", "map_url", "owner",
    "brand", "area_m2", "ownership", "free_power_kwt", "connection_cost",
    "rent_cost_month", "planned_power_kwt", "planned_ezs_count", "ports_gbt",
    "ports_ccs", "ports_chademo", "ports_type", "supplier", "contractor",
    "tu_status", "tech_conn_type", "dop_service", "comment", "archive_reason",
    "cadastral_no", "lat", "lon", "region_norm", "region_id",
    # графы чек-листа согласования, добавленные v2.26
    "free_power_num", "input_price_kwth", "tp_cost", "smr_cost", "distance_to_tp_m",
    "tp_term_months", "long_term_contract", "has_video", "access_24x7", "has_mobile",
    "owner_contact", "source_company", "source_person",
)

# Графы, которые описывают ПРИСОЕДИНЕНИЕ, а не участок: они едут в карточку
# `ezs_tech_connections` (поле модели → ключ из `_match_field`).
_TC_FROM_IMPORT = {
    "needs_reconstruction": "tp_reconstruction",
    "substation_owner": "substation_owner",
    "line_owner": "line_owner",
    "transformer_kva": "transformer_kva",
    "line_type": "line_type",
    "extra_power_possible": "extra_power_possible",
    "transformer_swap_possible": "transformer_swap_possible",
    "cost": "tu_contract_cost",
    "works_cost": "tu_works_cost",
    "total_cost": "tu_total_cost",
    "applicant_term_months": "applicant_term_months",
}
_TC_BOOL = {"needs_reconstruction", "extra_power_possible", "transformer_swap_possible"}
_TC_NUM = {"cost", "works_cost", "total_cost", "applicant_term_months"}


def _tc_values(vals: dict[str, Any]) -> dict[str, Any]:
    """Значения карточки присоединения из строки файла (пустые ключи опускаем)."""
    out: dict[str, Any] = {}
    for field, src in _TC_FROM_IMPORT.items():
        v = vals.get(src)
        if v is None or str(v).strip() == "":
            continue
        parsed = _bool(v) if field in _TC_BOOL else (_num(v) if field in _TC_NUM else _s(v, 200))
        if parsed is not None:
            out[field] = parsed
    return out


async def import_sites_xlsx(db: AsyncSession, company_id, content: bytes, dry_run: bool) -> dict[str, Any]:
    """Импорт сводного «Банк данных ЗУ» (3 листа → стадии). UPSERT по идентичности.

    Возвращает отчёт качества: что создано/обновлено/сдвинуто по стадиям, дубли
    внутри файла, нераспознанные регионы, конфликты стадий.
    """
    import io

    import openpyxl

    regions = (await db.execute(
        select(Region).where(Region.company_id == company_id))).scalars().all()
    resolver = RegionResolver(list(regions))
    existing = (await db.execute(
        select(EzsSite).where(EzsSite.company_id == company_id))).scalars().all()
    index = SiteIndex(list(existing))

    now = datetime.now(timezone.utc)
    today = _today()
    report: dict[str, Any] = {
        "dryRun": dry_run, "sheets": [], "total": 0, "withCoords": 0, "unknownSheets": [],
        "created": 0, "updated": 0, "unchanged": 0, "stageMoved": 0, "reactivated": 0,
        "archived": 0, "fileDuplicates": [], "nearConflicts": [],
        "regionsUnmatched": [], "skippedNoKey": 0, "withCadastral": 0,
        "techConnections": {"created": 0, "updated": 0},
    }
    # Карточки присоединения пишем после flush: у новых площадок id ещё нет.
    tc_pending: list[tuple[Any, dict[str, Any]]] = []
    seen_in_file: dict[str, str] = {}   # dedup_key → адрес первой строки
    touched: set[int] = set()           # площадки, уже обработанные этой загрузкой
    # События истории пишем после commit площадок: у новых записей id появляется
    # только после flush, а ссылаться на несуществующую строку нельзя.
    events: list[tuple[Any, str, str, str | None, str | None]] = []

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        sheet_stage = SHEET_STAGE.get(_norm(sn))
        if sheet_stage is None:
            report["unknownSheets"].append(sn)
            continue
        header_idx, colmap = None, {}
        for ri, row in enumerate(rows[:5]):
            cm = {}
            for ci, cell in enumerate(row):
                f = _match_field(_norm(cell))
                if f and f not in cm.values():
                    cm[ci] = f
            if len(cm) >= 5:
                header_idx, colmap = ri, cm
                break
        if header_idx is None:
            report["sheets"].append({"sheet": sn, "stage": sheet_stage, "rows": 0,
                                     "note": "заголовок не найден"})
            continue
        headers = {ci: (str(rows[header_idx][ci]).strip() if rows[header_idx][ci] else f"col{ci}")
                   for ci in range(len(rows[header_idx]))}

        cnt = 0
        for ri, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            vals: dict[str, Any] = {}
            raw: dict[str, str] = {}
            for ci, cell in enumerate(row):
                if cell is None or str(cell).strip() == "":
                    continue
                raw[headers.get(ci, f"col{ci}")] = str(cell).strip()
                f = colmap.get(ci)
                if f:
                    vals[f] = cell
            meaningful = {k for k in vals if k not in ("status_raw",)}
            if not meaningful and not raw.get("Регион"):
                continue

            lat, lon = _coords(vals.get("coords_raw"))
            city = _s(vals.get("city"), 160)
            region_raw = _s(vals.get("region"), 160)
            region_norm, region = resolver.resolve(region_raw, city)
            cadastral = _s(vals.get("cadastral_no"), 40) or _find_cadastral(raw)
            addr_key = _addr_key(_s(vals.get("full_address")) or _s(vals.get("address")), city)
            # Адрес входит в fallback обязательно: без него «Псков, ул. Калинина»
            # и «Псков, ул. Воеводы Шуйского» дают один хеш и склеиваются.
            fallback = _norm(" ".join(str(v) for v in (
                region_raw, city, vals.get("full_address") or vals.get("address"),
                vals.get("install_place"), vals.get("owner")) if v))
            key = _dedup_key(cadastral, lat, lon, addr_key, fallback)
            stage = _stage_for_row(sheet_stage, vals, raw)

            fields: dict[str, Any] = {
                "status_raw": _s(vals.get("status_raw"), 80),
                "received_date": _iso_date(vals.get("received_date")),
                "region": region_raw, "region_norm": region_norm,
                "region_id": region.id if region is not None else None,
                "city": city,
                "address": _s(vals.get("address")), "full_address": _s(vals.get("full_address")),
                "place_kind": _place_kind(vals.get("place_kind")),
                "install_place": _s(vals.get("install_place"), 300),
                "route": _s(vals.get("route"), 80), "lat": lat, "lon": lon,
                "coords_raw": _s(vals.get("coords_raw"), 120), "map_url": _s(vals.get("map_url")),
                "owner": _s(vals.get("owner"), 400), "brand": _s(vals.get("brand"), 160),
                "area_m2": _num(vals.get("area_m2")), "ownership": _s(vals.get("ownership"), 60),
                "free_power_kwt": _s(vals.get("free_power_kwt"), 80),
                # Мощность в файле пишут текстом («100кВт», «60-80 квт»): числом
                # берём первое значение, иначе гейт стадии dd не закрыть ничем.
                "free_power_num": _num(vals.get("free_power_kwt")),
                "connection_cost": _num(vals.get("connection_cost")),
                "input_price_kwth": _num(vals.get("input_price_kwth")),
                "tp_cost": _num(vals.get("tp_cost")),
                "smr_cost": _num(vals.get("smr_cost")),
                "distance_to_tp_m": _num(vals.get("distance_to_tp_m")),
                "tp_term_months": _num(vals.get("tp_term_months")),
                "long_term_contract": _bool(vals.get("long_term_contract")),
                "has_video": _bool(vals.get("has_video")),
                "access_24x7": _bool(vals.get("access_24x7")),
                "has_mobile": _bool(vals.get("has_mobile")),
                "owner_contact": _s(vals.get("owner_contact")),
                "source_company": _s(vals.get("source_company"), 300),
                "source_person": _s(vals.get("source_person")),
                "rent_cost_month": _num(vals.get("rent_cost_month")),
                "planned_power_kwt": _num(vals.get("planned_power_kwt")),
                "planned_ezs_count": _int(vals.get("planned_ezs_count")),
                "ports_gbt": _s(vals.get("ports_gbt"), 40), "ports_ccs": _s(vals.get("ports_ccs"), 40),
                "ports_chademo": _s(vals.get("ports_chademo"), 40),
                "ports_type": _s(vals.get("ports_type"), 40),
                "supplier": _s(vals.get("supplier"), 300), "contractor": _s(vals.get("contractor"), 400),
                "tu_status": _s(vals.get("tu_status")),
                "tech_conn_type": _s(vals.get("tech_conn_type"), 300),
                "dop_service": _s(vals.get("dop_service"), 300), "comment": _s(vals.get("comment")),
                "archive_reason": _archive_reason(stage, vals, raw),
                "cadastral_no": cadastral,
            }

            if not key:
                # Ни адреса, ни координат, ни собственника — место не опознать.
                # Такую строку нельзя ни отличить от соседней, ни найти при
                # следующей загрузке: пропускаем и показываем в отчёте.
                report["skippedNoKey"] += 1
                continue
            if cadastral:
                report["withCadastral"] += 1

            found, how = index.find(cadastral, lat, lon, addr_key, key)
            if how == "geo_conflict":
                # Точка в 50 м от известной площадки, но адрес другой. Молча
                # склеивать нельзя — заводим отдельную запись и показываем в отчёте.
                report["nearConflicts"].append({
                    "sheet": sn, "row": ri,
                    "address": (fields["full_address"] or fields["address"] or "")[:120],
                    "near": (found.full_address or found.address or "")[:120],
                })
                found = None
                # Свой ключ, а не общая гео-ячейка с соседом — иначе следующая
                # загрузка снова не узнает эту площадку и создаст дубль.
                key = _dedup_key(cadastral, None, None, addr_key, fallback)
            # Вторая строка файла на ту же площадку — дубль в источнике. Не
            # применяем её поверх первой: иначе raw «мигает» между загрузками и
            # каждый повторный импорт выглядит как изменение.
            if found is not None and id(found) in touched:
                report["fileDuplicates"].append({
                    "sheet": sn, "row": ri, "key": key or "",
                    "address": (fields["full_address"] or fields["address"] or "")[:120],
                    "first": seen_in_file.get(key or "", "")[:120],
                })
                cnt += 1
                continue
            if key:
                seen_in_file.setdefault(
                    key, fields["full_address"] or fields["address"] or f"{sn}:{ri}")

            if found is None:
                site = EzsSite(
                    company_id=company_id, stage=stage, stage_since=today,
                    dedup_key=key, source_sheet=sn, row_no=ri, raw=raw,
                    first_seen_at=now, last_seen_at=now, updated_at=now, **fields)
                if not dry_run:
                    db.add(site)
                    events.append((site, "import", f"Заведена из файла «{sn}»", None, None))
                index.add(site)
                touched.add(id(site))
                report["created"] += 1
            else:
                prev_stage = found.stage
                changed = _apply_update(found, fields, raw, sn, ri, key, now, not dry_run)
                moved = _advance_stage(found, stage, today, not dry_run)
                if moved == "forward":
                    report["stageMoved"] += 1
                elif moved == "reactivated":
                    report["reactivated"] += 1
                elif moved == "archived":
                    report["archived"] += 1
                if moved and not dry_run:
                    events.append((found, "stage",
                                   f"Импорт: {STAGE_LABELS.get(prev_stage, prev_stage)} → "
                                   f"{STAGE_LABELS.get(stage, stage)}", prev_stage, stage))
                if changed or moved:
                    report["updated"] += 1
                else:
                    report["unchanged"] += 1
                if not dry_run:
                    found.last_seen_at = now
                touched.add(id(found))

            tc_vals = _tc_values(vals)
            if tc_vals:
                tc_pending.append((site if found is None else found, tc_vals))

            if lat is not None:
                report["withCoords"] += 1
            cnt += 1
        report["sheets"].append({"sheet": sn, "stage": sheet_stage, "rows": cnt})
    wb.close()

    report["total"] = report["created"] + report["updated"] + report["unchanged"]
    report["regionsUnmatched"] = [{"value": v, "count": n} for v, n in
                                  sorted(resolver.unmatched.items(), key=lambda x: -x[1])[:20]]
    if not dry_run:
        await db.flush()          # id новых площадок — до записи событий
        from app.models import EzsSiteEvent
        for site, kind, text_, frm, to in events:
            db.add(EzsSiteEvent(company_id=company_id, site_id=site.id, kind=kind,
                                text=text_, from_stage=frm, to_stage=to))
        await _sync_tech_connections(db, company_id, tc_pending, now, report)
        await db.commit()
    else:
        # В предпросмотре считаем, сколько карточек присоединения появилось бы.
        report["techConnections"]["created"] = len(tc_pending)
    return report


async def _sync_tech_connections(db: AsyncSession, company_id, pending: list[tuple[Any, dict[str, Any]]],
                                 now: datetime, report: dict[str, Any]) -> None:
    """Заводит карточку присоединения по данным файла и дополняет пустые поля.

    Заполненное НЕ перетираем: карточку ведут руками (переписка с сетевой
    организацией, номера ТУ), а файл знает только то, что снял отдел развития
    на осмотре.
    """
    from app.models import EzsTechConnection

    if not pending:
        return
    site_ids = {s.id for s, _ in pending}
    existing = {tc.site_id: tc for tc in (await db.execute(
        select(EzsTechConnection).where(
            EzsTechConnection.company_id == company_id,
            EzsTechConnection.site_id.in_(site_ids)))).scalars()}
    for site, vals in pending:
        tc = existing.get(site.id)
        if tc is None:
            tc = EzsTechConnection(company_id=company_id, site_id=site.id,
                                   status="draft", created_at=now, **vals)
            db.add(tc)
            existing[site.id] = tc
            report["techConnections"]["created"] += 1
            continue
        changed = False
        for f, v in vals.items():
            if getattr(tc, f, None) is None:
                setattr(tc, f, v)
                changed = True
        if changed:
            tc.updated_at = now
            report["techConnections"]["updated"] += 1


def _place_kind(v: Any) -> str | None:
    """«город»/«Город»/«трасса» → канон; мусор («Магнит», «псковэнерго») → None."""
    t = _norm(v)
    if not t:
        return None
    if t.startswith("город") or t == "г":
        return "город"
    if t.startswith("трасс"):
        return "трасса"
    return None


def _archive_reason(stage: str, vals: dict[str, Any], raw: dict[str, str]) -> str | None:
    """Причина отклонения — только для архива и только из своей колонки.

    Подставлять сюда «Комментарий» нельзя: там лежат служебные пометки прошлых
    чисток, и они выглядели бы как настоящая причина отказа.
    """
    if stage != "archive":
        return None
    txt = _s(vals.get("archive_reason")) or raw.get("Причина согласия / отказа")
    return txt[:200] if txt else None


def _apply_update(site: EzsSite, fields: dict[str, Any], raw: dict[str, str],
                  sheet: str, row_no: int, key: str | None, now: datetime,
                  apply: bool) -> bool:
    """Дополняет пустые поля значениями из файла. Известное НЕ затирает.

    Файл — не единственный источник: часть данных ведут руками в карточке, и
    очередная выгрузка не должна их стирать (в старом REPLACE-ALL стирала всё).
    apply=False — предпросмотр: только считаем, объекты сессии не трогаем.
    """
    changed = False
    manual = set(site.manual_fields or [])   # поля, введённые руками, — не наше дело
    for f in _FILL_FIELDS:
        if f in manual:
            continue
        new = fields.get(f)
        if new is None or new == "":
            continue
        if getattr(site, f, None) in (None, ""):
            if apply:
                setattr(site, f, new)
            changed = True
    # raw обновляем целиком — это снимок последней строки файла.
    merged = dict(site.raw or {})
    merged.update(raw)
    if merged != (site.raw or {}):
        if apply:
            site.raw = merged
        changed = True
    if apply:
        site.source_sheet = sheet
        site.row_no = row_no
        if key and not site.dedup_key:
            site.dedup_key = key
        if changed:
            site.updated_at = now
    return changed


def _advance_stage(site: EzsSite, stage: str, today: str, apply: bool) -> str | None:
    """Двигает стадию по правилам воронки. Назад по воронке файл не откатывает.

    • вперёд по STAGE_ORDER — двигаем;
    • в архив — двигаем всегда (в файле площадка ушла в отказ);
    • из архива в активную — реактивация (собственник вернулся к разговору);
    • назад внутри активных — игнорируем: обычно это отставшая выгрузка.
    """
    def move(kind: str) -> str:
        if apply:
            site.prev_stage, site.stage, site.stage_since = site.stage, stage, today
        return kind

    if stage == site.stage:
        return None
    if stage == "archive":
        return move("archived")
    if site.stage == "archive":
        return move("reactivated")
    cur, new = _STAGE_POS.get(site.stage, -1), _STAGE_POS.get(stage, -1)
    if new > cur >= 0:
        return move("forward")
    return None


def _risk_conditions(risk: str) -> list[Any]:
    """Фильтры под цифры обзора портфеля — чтобы «297 без ответственного» можно
    было раскрыть и увидеть, кто именно за этим числом стоит.

    Считаются ровно так же, как в `portfolio_overview`: иначе список не сойдётся
    с цифрой, и доверия к экрану не будет.
    """
    from sqlalchemy import exists, select as _select

    from app.models import EzsSiteEquipment, EzsTechConnection

    S = EzsSite
    today = date.today().isoformat()
    d30 = (date.today() - timedelta(days=30)).isoformat()
    d90 = (date.today() - timedelta(days=90)).isoformat()
    active = [S.stage.in_(STAGE_ORDER)]

    if risk == "step_overdue":
        return [*active, S.next_action_due.is_not(None), S.next_action_due < today]
    if risk == "no_owner":
        return [*active, S.owner_user_id.is_(None)]
    if risk == "no_next":
        return [*active, S.next_action.is_(None)]
    if risk == "stuck_90":
        return [*active, func.coalesce(S.stage_since, "1970-01-01") < d90]
    if risk == "no_touch_30":
        return [*active, func.coalesce(
            func.to_char(S.last_touch_at, "YYYY-MM-DD"), "1970-01-01") < d30]
    if risk == "tp_overdue":
        return [*active, exists(_select(EzsTechConnection.id).where(
            EzsTechConnection.site_id == S.id,
            EzsTechConnection.due_date.is_not(None),
            EzsTechConnection.done_date.is_(None),
            EzsTechConnection.due_date < today,
            EzsTechConnection.status.notin_(["done", "rejected"])))]
    if risk == "eq_overdue":
        return [*active, exists(_select(EzsSiteEquipment.id).where(
            EzsSiteEquipment.site_id == S.id,
            EzsSiteEquipment.due_date.is_not(None),
            EzsSiteEquipment.supplied_date.is_(None),
            EzsSiteEquipment.due_date < today,
            EzsSiteEquipment.status.in_(["planned", "ordered"])))]
    return []


async def list_sites(
    db: AsyncSession, company_id, *, stage: str | None = None, region: str | None = None,
    search: str | None = None, owner_id=None, overdue: bool = False,
    risk: str | None = None, page: int = 1, page_size: int = 100,
) -> dict[str, Any]:
    from app.models import User

    S = EzsSite
    conds = [S.company_id == company_id]
    if risk:
        conds += _risk_conditions(risk)
    if stage == "active":            # вся живая часть воронки одним фильтром
        conds.append(S.stage.in_(STAGE_ORDER))
    elif stage:
        conds.append(S.stage == stage)
    if region:
        conds.append(func.coalesce(S.region_norm, S.region) == region)
    if owner_id:
        conds.append(S.owner_user_id == owner_id)
    if overdue:
        # Просрочен следующий шаг — то, ради чего у площадки вообще есть срок.
        conds.append(S.next_action_due.is_not(None))
        conds.append(S.next_action_due < _today())
        conds.append(S.stage.in_(STAGE_ORDER))
    if search:
        like = f"%{search.lower()}%"
        conds.append(func.lower(func.coalesce(S.full_address, "") + " " + func.coalesce(S.city, "")
                                 + " " + func.coalesce(S.owner, "") + " " + func.coalesce(S.install_place, "")).like(like))
    total = int((await db.execute(select(func.count()).select_from(S).where(*conds))).scalar_one() or 0)
    rows = (await db.execute(
        select(S, func.coalesce(User.name, User.email))
        .outerjoin(User, User.id == S.owner_user_id).where(*conds)
        .order_by(func.coalesce(S.region_norm, S.region).nulls_last(), S.city.nulls_last(), S.id)
        .offset((page - 1) * page_size).limit(page_size)
    )).all()
    items = []
    for s, owner_name in rows:
        row = _site_out(s)
        row.update({"ownerName": owner_name, "nextAction": s.next_action,
                    "nextActionDue": s.next_action_due,
                    "lastTouchAt": s.last_touch_at.isoformat() if s.last_touch_at else None})
        items.append(row)
    return {"total": total, "page": page, "pageSize": page_size, "items": items}


def _site_out(s: EzsSite) -> dict[str, Any]:
    return {
        "id": str(s.id), "projectNo": s.project_no, "title": s.title,
        "stage": s.stage, "stageLabel": STAGE_LABELS.get(s.stage, s.stage),
        "phase": STAGE_PHASE.get(s.stage), "phaseLabel": PHASE_LABELS.get(STAGE_PHASE.get(s.stage, "")),
        "stageSince": s.stage_since, "prevStage": s.prev_stage,
        "archiveReason": s.archive_reason, "cadastralNo": s.cadastral_no,
        "statusRaw": s.status_raw, "receivedDate": s.received_date,
        "region": s.region_norm or s.region, "regionRaw": s.region,
        "city": s.city, "address": s.address, "fullAddress": s.full_address,
        "placeKind": s.place_kind, "installPlace": s.install_place, "route": s.route,
        "lat": s.lat, "lon": s.lon, "mapUrl": s.map_url,
        "owner": s.owner, "brand": s.brand, "areaM2": s.area_m2, "ownership": s.ownership,
        "freePowerKwt": s.free_power_kwt,
        "connectionCost": float(s.connection_cost) if s.connection_cost is not None else None,
        "rentCostMonth": float(s.rent_cost_month) if s.rent_cost_month is not None else None,
        "plannedPowerKwt": s.planned_power_kwt, "plannedEzsCount": s.planned_ezs_count,
        "portsGbt": s.ports_gbt, "portsCcs": s.ports_ccs, "portsChademo": s.ports_chademo, "portsType": s.ports_type,
        "supplier": s.supplier, "contractor": s.contractor, "tuStatus": s.tu_status,
        "techConnType": s.tech_conn_type, "dopService": s.dop_service, "comment": s.comment,
    }


async def site_detail(db: AsyncSession, company_id, site_id) -> dict[str, Any] | None:
    from app.models import User
    from app.services.ezs_site_work import (
        site_doc_kinds, site_equipment_supplied, site_out_full,
    )

    row = (await db.execute(
        select(EzsSite, func.coalesce(User.name, User.email))
        .outerjoin(User, User.id == EzsSite.owner_user_id)
        .where(EzsSite.company_id == company_id, EzsSite.id == site_id)
    )).first()
    if row is None:
        return None
    s, owner_name = row
    out = site_out_full(s, owner_name, doc_kinds=await site_doc_kinds(db, s.id),
                        equipment_supplied=await site_equipment_supplied(db, s.id))
    out["raw"] = s.raw or {}       # все 55 исходных колонок для карточки
    out["sourceSheet"] = s.source_sheet
    out["firstSeenAt"] = s.first_seen_at.isoformat() if s.first_seen_at else None
    out["lastSeenAt"] = s.last_seen_at.isoformat() if s.last_seen_at else None
    return out


async def sites_overview(db: AsyncSession, company_id) -> dict[str, Any]:
    S = EzsSite
    base = S.company_id == company_id
    total = int((await db.execute(select(func.count()).select_from(S).where(base))).scalar_one() or 0)
    by_stage = {r.stage: int(r.n) for r in (await db.execute(
        select(S.stage, func.count().label("n")).where(base).group_by(S.stage))).all()}
    reg_expr = func.coalesce(S.region_norm, S.region, "— не указан")
    regions = (await db.execute(
        select(reg_expr.label("region"), func.count().label("n"))
        .where(base).group_by(reg_expr).order_by(func.count().desc()).limit(15))).all()
    agg = (await db.execute(select(
        func.count().filter(S.lat.is_not(None)).label("with_coords"),
        func.coalesce(func.sum(S.planned_ezs_count), 0).label("ezs"),
        func.coalesce(func.sum(S.planned_power_kwt), 0).label("power"),
        func.count().filter(S.connection_cost.is_not(None)).label("with_cost"),
        func.coalesce(func.sum(S.connection_cost), 0).label("cost_sum"),
        func.count().filter(S.cadastral_no.is_not(None)).label("with_cad"),
        func.count().filter(S.region_id.is_not(None)).label("with_region"),
    ).where(base))).one()

    active = sum(by_stage.get(s, 0) for s in STAGE_ORDER)
    funnel = [{"stage": st, "label": STAGE_LABELS[st], "hint": STAGE_HINTS[st],
               "count": by_stage.get(st, 0)} for st in STAGE_ORDER]

    # Управляемость активной части: кто ведёт, что просрочено, что забыто.
    stale_before = (datetime.now(timezone.utc) - timedelta(days=STALE_DAYS))
    work = (await db.execute(select(
        func.count().filter(S.owner_user_id.is_(None)).label("no_owner"),
        func.count().filter(S.next_action.is_(None)).label("no_next"),
        func.count().filter(S.next_action_due.is_not(None),
                            S.next_action_due < _today()).label("overdue"),
        func.count().filter(func.coalesce(S.last_touch_at, S.first_seen_at) < stale_before)
        .label("stale"),
    ).where(base, S.stage.in_(STAGE_ORDER)))).one()
    return {
        "total": total,
        "active": active,
        "onHold": by_stage.get("on_hold", 0),
        "archived": by_stage.get("archive", 0),
        "funnel": funnel,
        # Совместимость с прежним контрактом фронта (byStage) — те же числа.
        "byStage": funnel + [{"stage": s, "label": STAGE_LABELS[s], "hint": STAGE_HINTS[s],
                              "count": by_stage.get(s, 0)} for s in ("on_hold", "archive")],
        "byRegion": [{"region": r[0], "count": int(r[1])} for r in regions],
        "withCoords": int(agg.with_coords or 0),
        "plannedEzs": int(agg.ezs or 0),
        "plannedPowerKwt": round(float(agg.power or 0), 1),
        "withKnownCost": int(agg.with_cost or 0),
        "connectionCostSum": round(float(agg.cost_sum or 0), 2),
        "quality": {
            "withCadastral": int(agg.with_cad or 0),
            "regionMatched": int(agg.with_region or 0),
            "withCoords": int(agg.with_coords or 0),
        },
        # Считается только по активной части: у архива нет ни срока, ни смысла в нём.
        "work": {
            "noOwner": int(work.no_owner or 0),
            "noNextAction": int(work.no_next or 0),
            "overdue": int(work.overdue or 0),
            "stale": int(work.stale or 0),
            "staleDays": STALE_DAYS,
        },
    }
