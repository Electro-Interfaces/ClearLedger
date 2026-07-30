"""Проект ЭЗС: документы, техприсоединение, бюджет, портфель.

Волны 4–8 (`docs/SITES_PROJECT_LIFECYCLE.md`). Проект — это площадка
(`ezs_sites`) с номером, этапами и тремя спутниками:

  • документы (`ezs_site_docs`) — доказательная база; часть пунктов гейта
    закрывается именно приложенным файлом, а не галочкой;
  • техприсоединение (`ezs_tech_connections`) — со своим циклом и сроками:
    именно оно определяет срок всего проекта, а не стройка;
  • бюджет (`ezs_site_costs`) — план и факт по статьям.

Отдельно — «Ждёт учёта»: мост в бухгалтерию. Система показывает, что пора
завести в учёт, но записи создаёт человек: бухгалтерский контур не должен
наполняться побочным эффектом смены статуса.
"""
from __future__ import annotations


import uuid as _uuid
from datetime import date, datetime, timedelta, timezone
from typing import Any

from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Contract, EzsSite, EzsSiteCost, EzsSiteDoc, EzsSiteEquipment, EzsTechConnection,
    ServiceLocation, User,
)
from app.services.ezs_changes import make_change
from app.services.ezs_site_work import log_event
from app.services.ezs_checklist import norm_days
from app.services.ezs_sites import (
    PHASE_LABELS, PHASES, STAGE_LABELS, STAGE_ORDER, STAGE_PHASE,
)

# Норматив стадии прямо в SQL: параметр перед `::` (`:norms::jsonb`) SQLAlchemy
# в text() не связывает — падает «A value is required for bind parameter».
# Значения — целые из нашего кода, не пользовательский ввод.
_NORM_CASE = "(case stage " + " ".join(
    f"when '{s}' then {norm_days(s)}" for s in STAGE_ORDER) + " else 90 end)"


def _stage_position_sql(column: str) -> str:
    allowed = {"e.from_stage", "e.to_stage"}
    if column not in allowed:
        raise ValueError(f"Недопустимая колонка стадии: {column}")
    whens = " ".join(f"when '{stage}' then {index}"
                     for index, stage in enumerate(STAGE_ORDER))
    return f"(case {column} {whens} else -1 end)"


# ── Документы ──────────────────────────────────────────────────────────────
DOC_KINDS = [
    {"key": "egrn", "label": "Выписка ЕГРН"},
    {"key": "site_plan", "label": "Схема участка"},
    {"key": "photo", "label": "Фото места"},
    {"key": "offer", "label": "Коммерческое предложение"},
    {"key": "contract", "label": "Договор на землю"},
    {"key": "tu", "label": "Технические условия"},
    {"key": "tp_contract", "label": "Договор техприсоединения"},
    {"key": "project", "label": "Проектная документация"},
    {"key": "act_mount", "label": "Акт монтажа"},
    {"key": "act", "label": "Акт приёмки / ввода"},
    {"key": "equipment_doc", "label": "Документы на оборудование"},
    {"key": "other", "label": "Прочее"},
]
DOC_LABELS = {d["key"]: d["label"] for d in DOC_KINDS}


async def list_docs(db: AsyncSession, company_id, site_id) -> list[dict[str, Any]]:
    from app.models import SourceFile
    rows = (await db.execute(
        select(EzsSiteDoc, SourceFile.file_name, SourceFile.size, User.name, User.email)
        .outerjoin(SourceFile, SourceFile.id == EzsSiteDoc.file_id)
        .outerjoin(User, User.id == EzsSiteDoc.uploaded_by)
        .where(EzsSiteDoc.company_id == company_id, EzsSiteDoc.site_id == site_id)
        .order_by(EzsSiteDoc.created_at.desc()))).all()
    return [{
        "id": str(d.id), "kind": d.kind, "kindLabel": DOC_LABELS.get(d.kind, d.kind),
        "title": d.title, "note": d.note, "stage": d.stage,
        "stageLabel": STAGE_LABELS.get(d.stage or "", d.stage),
        "fileId": str(d.file_id) if d.file_id else None, "fileName": fname,
        "fileSize": size, "uploadedBy": name or email,
        "createdAt": d.created_at.isoformat() if d.created_at else None,
    } for d, fname, size, name, email in rows]


async def add_doc(db: AsyncSession, company_id, site: EzsSite, *, file_id, kind: str,
                  title: str | None, note: str | None, user: User | None) -> dict[str, Any]:
    doc = EzsSiteDoc(
        company_id=company_id, site_id=site.id,
        file_id=_uuid.UUID(str(file_id)) if file_id else None,
        kind=kind if kind in DOC_LABELS else "other",
        title=title, note=note, stage=site.stage,
        uploaded_by=user.id if user is not None else None,
    )
    db.add(doc)
    await db.flush()
    await log_event(db, site, "doc", user=user,
                    text=f"Приложен документ: {DOC_LABELS.get(doc.kind, doc.kind)}"
                         + (f" — {title}" if title else ""),
                    changes=[make_change(
                        f"document:{doc.kind}", None, str(doc.id),
                        label=f"Документ: {DOC_LABELS.get(doc.kind, doc.kind)}",
                        category="decision", new_display=title or "добавлен",
                    )])
    return {"id": str(doc.id)}


async def delete_doc(db: AsyncSession, company_id, site: EzsSite, doc_id, user: User | None) -> bool:
    doc = (await db.execute(select(EzsSiteDoc).where(
        EzsSiteDoc.company_id == company_id, EzsSiteDoc.site_id == site.id,
        EzsSiteDoc.id == doc_id))).scalar_one_or_none()
    if doc is None:
        return False
    await log_event(db, site, "doc", user=user,
                    text=f"Удалён документ: {DOC_LABELS.get(doc.kind, doc.kind)}",
                    changes=[make_change(
                        f"document:{doc.kind}", str(doc.id), None,
                        label=f"Документ: {DOC_LABELS.get(doc.kind, doc.kind)}",
                        category="decision", old_display=doc.title or "был приложен",
                        new_display="удалён",
                    )])
    await db.delete(doc)
    return True


# ── Техприсоединение ───────────────────────────────────────────────────────
TC_STATUSES = [
    {"key": "draft", "label": "Не начато"},
    {"key": "applied", "label": "Заявка подана"},
    {"key": "specs", "label": "ТУ получены"},
    {"key": "contract", "label": "Договор ТП"},
    {"key": "in_progress", "label": "Мероприятия идут"},
    {"key": "done", "label": "Присоединение исполнено"},
    {"key": "rejected", "label": "Отказ сетевой"},
]
TC_LABELS = {s["key"]: s["label"] for s in TC_STATUSES}
TC_FIELDS = {"status", "grid_operator", "application_no", "application_date", "specs_no",
             "specs_date", "contract_no", "contract_date", "power_kwt", "voltage", "cost",
             "due_date", "done_date", "needs_reconstruction", "note",
             # паспорт питающей сети и деньги ТУ (графы AQ–BA банка ЗУ)
             "substation_owner", "line_owner", "transformer_kva", "line_type",
             "extra_power_possible", "transformer_swap_possible",
             "works_cost", "total_cost", "applicant_term_months"}
_TC_NUM = {"power_kwt", "cost", "works_cost", "total_cost", "applicant_term_months"}
_TC_BOOL_FIELDS = {"needs_reconstruction", "extra_power_possible", "transformer_swap_possible"}


def _tc_out(tc: EzsTechConnection) -> dict[str, Any]:
    overdue = bool(tc.due_date and not tc.done_date and tc.due_date < date.today().isoformat()
                   and tc.status not in ("done", "rejected"))
    return {
        "id": str(tc.id), "siteId": str(tc.site_id),
        "status": tc.status, "statusLabel": TC_LABELS.get(tc.status, tc.status),
        "gridOperator": tc.grid_operator,
        "applicationNo": tc.application_no, "applicationDate": tc.application_date,
        "specsNo": tc.specs_no, "specsDate": tc.specs_date,
        "contractNo": tc.contract_no, "contractDate": tc.contract_date,
        "powerKwt": tc.power_kwt, "voltage": tc.voltage,
        "cost": float(tc.cost) if tc.cost is not None else None,
        "dueDate": tc.due_date, "doneDate": tc.done_date,
        "needsReconstruction": tc.needs_reconstruction, "note": tc.note,
        "substationOwner": tc.substation_owner, "lineOwner": tc.line_owner,
        "transformerKva": tc.transformer_kva, "lineType": tc.line_type,
        "extraPowerPossible": tc.extra_power_possible,
        "transformerSwapPossible": tc.transformer_swap_possible,
        "worksCost": float(tc.works_cost) if tc.works_cost is not None else None,
        "totalCost": float(tc.total_cost) if tc.total_cost is not None else None,
        "applicantTermMonths": tc.applicant_term_months,
        "overdue": overdue,
    }


async def get_tech_connection(db: AsyncSession, company_id, site_id) -> dict[str, Any] | None:
    tc = (await db.execute(select(EzsTechConnection).where(
        EzsTechConnection.company_id == company_id,
        EzsTechConnection.site_id == site_id))).scalars().first()
    return _tc_out(tc) if tc else None


async def upsert_tech_connection(db: AsyncSession, company_id, site: EzsSite,
                                 patch: dict[str, Any], user: User | None) -> dict[str, Any]:
    tc = (await db.execute(select(EzsTechConnection).where(
        EzsTechConnection.company_id == company_id,
        EzsTechConnection.site_id == site.id))).scalars().first()
    created = tc is None
    if tc is None:
        tc = EzsTechConnection(company_id=company_id, site_id=site.id)
        db.add(tc)
    changes: list[dict[str, Any]] = []
    for f, v in patch.items():
        if f not in TC_FIELDS:
            continue
        old = getattr(tc, f, None)
        if v in ("", None):
            setattr(tc, f, None)
        elif f in _TC_NUM:
            try:
                setattr(tc, f, float(str(v).replace(",", ".").replace(" ", "")))
            except ValueError:
                pass
        elif f in _TC_BOOL_FIELDS:
            setattr(tc, f, bool(v))
        else:
            setattr(tc, f, str(v))
        new = getattr(tc, f, None)
        if old != new:
            changes.append(make_change(
                f"tc.{f}", old, new,
                old_display=TC_LABELS.get(old, old) if f == "status" and old else None,
                new_display=TC_LABELS.get(new, new) if f == "status" and new else None,
            ))
    tc.updated_at = datetime.now(timezone.utc)
    # Пункт 5.6 гейта («фиксация сроков мероприятий ТУ») смотрит в графу площадки,
    # а срок задают здесь — месяцами заявителя или датой мероприятий. Без зеркала
    # человек заполняет присоединение целиком, а чек-лист продолжает требовать срок.
    if site.tp_term_months is None:
        months = tc.applicant_term_months
        if months is None and tc.due_date and tc.application_date:
            try:
                days = (date.fromisoformat(tc.due_date) - date.fromisoformat(tc.application_date)).days
                months = round(days / 30.44, 1) if days > 0 else None
            except ValueError:
                months = None
        if months:
            changes.append(make_change("tp_term_months", site.tp_term_months, months))
            site.tp_term_months = months
    await db.flush()
    if changes:
        await log_event(
            db, site, "edit", user=user, changes=changes,
            text=("Заведено техприсоединение: " if created else "Техприсоединение: ")
                 + ", ".join(item["label"] for item in changes),
        )
    elif created:
        await log_event(db, site, "note", text="Заведено техприсоединение", user=user)
    return _tc_out(tc)


async def tech_connections_report(db: AsyncSession, company_id) -> dict[str, Any]:
    """Реестр присоединений по проектам: статусы, сроки, просрочки, стоимость."""
    rows = (await db.execute(
        select(EzsTechConnection, EzsSite.project_no, EzsSite.title, EzsSite.region_norm,
               EzsSite.region, EzsSite.city, EzsSite.address, EzsSite.stage)
        .join(EzsSite, EzsSite.id == EzsTechConnection.site_id)
        .where(EzsTechConnection.company_id == company_id)
        .order_by(EzsTechConnection.due_date.nulls_last()))).all()
    items = []
    by_status: dict[str, int] = {}
    overdue = 0
    cost_sum = 0.0
    for tc, pno, title, rnorm, rraw, city, addr, stage in rows:
        out = _tc_out(tc)
        out.update({"projectNo": pno, "title": title, "region": rnorm or rraw,
                    "city": city, "address": addr, "stage": stage,
                    "stageLabel": STAGE_LABELS.get(stage, stage)})
        items.append(out)
        by_status[tc.status] = by_status.get(tc.status, 0) + 1
        overdue += 1 if out["overdue"] else 0
        cost_sum += out["cost"] or 0
    return {
        "total": len(items), "overdue": overdue, "costSum": round(cost_sum, 2),
        "byStatus": [{"key": s["key"], "label": s["label"], "count": by_status.get(s["key"], 0)}
                     for s in TC_STATUSES],
        "items": items,
    }


async def tech_connections_by_operator(db: AsyncSession, company_id) -> dict[str, Any]:
    """Присоединения в разрезе сетевой организации.

    Сетевая — первый разрез этой части работы, а не подпись в карточке: процессы,
    сроки и стоимость у разных сетевых различаются в разы, и «сколько ждать» без
    ответа «у кого» ничего не значит. Считаем то, что определяет план проекта:
    сколько заявок в работе, за сколько дней доходят до ТУ и до исполнения,
    сколько стоит присоединение и как часто приходит отказ.

    Отдельно — владелец подстанции и линии (графы осмотра AQ–AR): у собственных
    сетей «Общество» присоединение идёт иначе, чем через чужую организацию.
    """
    rows = (await db.execute(text("""
        select coalesce(nullif(trim(t.grid_operator), ''), '— не указана') as operator,
               t.grid_operator_counterparty_id as counterparty_id,
               t.status,
               t.application_date, t.specs_date, t.due_date, t.done_date,
               t.cost, t.total_cost, t.power_kwt,
               s.stage
        from ezs_tech_connections t join ezs_sites s on s.id = t.site_id
        where t.company_id = :cid
    """), {"cid": company_id})).mappings().all()

    def days(a: str | None, b: str | None) -> int | None:
        if not a or not b:
            return None
        try:
            return (date.fromisoformat(b[:10]) - date.fromisoformat(a[:10])).days
        except ValueError:
            return None

    def median(xs: list[int]) -> int | None:
        if not xs:
            return None
        xs = sorted(xs)
        mid = len(xs) // 2
        return xs[mid] if len(xs) % 2 else (xs[mid - 1] + xs[mid]) // 2

    today = date.today().isoformat()
    agg: dict[str, dict[str, Any]] = {}
    for r in rows:
        a = agg.setdefault(r["operator"], {
            "operator": r["operator"],
            "counterpartyId": str(r["counterparty_id"]) if r["counterparty_id"] else None,
            "total": 0, "done": 0, "rejected": 0, "overdue": 0, "inProgress": 0,
            "costSum": 0.0, "totalCostSum": 0.0, "powerSum": 0.0,
            "_toSpecs": [], "_toDone": [],
        })
        a["total"] += 1
        if r["status"] == "done":
            a["done"] += 1
        elif r["status"] == "rejected":
            a["rejected"] += 1
        elif r["status"] != "draft":
            a["inProgress"] += 1
        if r["due_date"] and not r["done_date"] and r["due_date"] < today \
                and r["status"] not in ("done", "rejected"):
            a["overdue"] += 1
        a["costSum"] += float(r["cost"] or 0)
        a["totalCostSum"] += float(r["total_cost"] or 0)
        a["powerSum"] += float(r["power_kwt"] or 0)
        d1 = days(r["application_date"], r["specs_date"])
        if d1 is not None:
            a["_toSpecs"].append(d1)
        d2 = days(r["application_date"], r["done_date"])
        if d2 is not None:
            a["_toDone"].append(d2)

    items = []
    for a in agg.values():
        to_specs, to_done = a.pop("_toSpecs"), a.pop("_toDone")
        a["daysToSpecs"] = median(to_specs)
        a["daysToDone"] = median(to_done)
        a["costSum"] = round(a["costSum"], 2)
        a["totalCostSum"] = round(a["totalCostSum"], 2)
        # Стоимость на киловатт — единственная цифра, которую можно сравнивать
        # между сетевыми: сами суммы зависят от мощности площадки.
        a["costPerKwt"] = (round(a["totalCostSum"] / a["powerSum"], 2)
                           if a["powerSum"] and a["totalCostSum"] else None)
        a["rejectPct"] = round(a["rejected"] / a["total"] * 100, 1) if a["total"] else None
        items.append(a)
    items.sort(key=lambda i: -i["total"])

    owners = [{"owner": r["owner"], "count": int(r["n"])} for r in (await db.execute(text("""
        select coalesce(nullif(trim(substation_owner), ''), '— не указан') as owner, count(*) n
        from ezs_tech_connections where company_id = :cid group by 1 order by 2 desc limit 15
    """), {"cid": company_id})).mappings().all()]

    named = sum(i["total"] for i in items if i["operator"] != "— не указана")
    return {
        "items": items,
        "substationOwners": owners,
        "total": sum(i["total"] for i in items),
        "withOperator": named,
        # Честная подпись под пустым разрезом: считать по сетевой можно только то,
        # у чего сетевая названа.
        "hint": ("Сетевая организация не заполнена ни в одной карточке — разрез "
                 "появится, когда её начнут указывать." if named == 0 else None),
    }


# ── Оборудование проекта ───────────────────────────────────────────────────
EQ_STATUSES = [
    {"key": "planned", "label": "Запланировано"},
    {"key": "ordered", "label": "Заказано"},
    {"key": "supplied", "label": "Поставлено"},
    {"key": "installed", "label": "Смонтировано"},
    {"key": "cancelled", "label": "Отменено"},
]
EQ_LABELS = {s["key"]: s["label"] for s in EQ_STATUSES}
EQ_FIELDS = {"status", "title", "manufacturer", "power_kwt", "connectors", "qty", "supplier",
             "price", "order_date", "due_date", "supplied_date", "installed_date", "note"}
_EQ_NUM = {"power_kwt", "price"}


def _eq_out(e: EzsSiteEquipment) -> dict[str, Any]:
    overdue = bool(e.due_date and not e.supplied_date and e.status in ("planned", "ordered")
                   and e.due_date < date.today().isoformat())
    return {
        "id": str(e.id), "siteId": str(e.site_id), "status": e.status,
        "statusLabel": EQ_LABELS.get(e.status, e.status),
        "title": e.title, "manufacturer": e.manufacturer, "powerKwt": e.power_kwt,
        "connectors": e.connectors, "qty": e.qty, "supplier": e.supplier,
        "price": float(e.price) if e.price is not None else None,
        "orderDate": e.order_date, "dueDate": e.due_date,
        "suppliedDate": e.supplied_date, "installedDate": e.installed_date,
        "note": e.note, "overdue": overdue,
    }


async def list_equipment(db: AsyncSession, company_id, site_id) -> dict[str, Any]:
    rows = (await db.execute(select(EzsSiteEquipment).where(
        EzsSiteEquipment.company_id == company_id, EzsSiteEquipment.site_id == site_id)
        .order_by(EzsSiteEquipment.created_at))).scalars().all()
    items = [_eq_out(e) for e in rows]
    active = [i for i in items if i["status"] != "cancelled"]
    return {
        "items": items,
        "priceTotal": round(sum((i["price"] or 0) * (i["qty"] or 1) for i in active), 2),
        # Гейт «Оборудование поставлено» закрывается, когда всё нужное приехало.
        "allSupplied": bool(active) and all(i["status"] in ("supplied", "installed") for i in active),
        "allInstalled": bool(active) and all(i["status"] == "installed" for i in active),
    }


async def upsert_equipment(db: AsyncSession, company_id, site: EzsSite, payload: dict[str, Any],
                           user: User | None) -> dict[str, Any]:
    eid = payload.get("id")
    row = None
    if eid:
        row = (await db.execute(select(EzsSiteEquipment).where(
            EzsSiteEquipment.company_id == company_id, EzsSiteEquipment.site_id == site.id,
            EzsSiteEquipment.id == _uuid.UUID(str(eid))))).scalar_one_or_none()
    created = row is None
    if row is None:
        row = EzsSiteEquipment(company_id=company_id, site_id=site.id)
        db.add(row)
    changes: list[dict[str, Any]] = []
    for f, v in payload.items():
        if f not in EQ_FIELDS:
            continue
        old = getattr(row, f, None)
        if v in ("", None):
            setattr(row, f, None)
        elif f in _EQ_NUM:
            try:
                setattr(row, f, float(str(v).replace(",", ".").replace(" ", "")))
            except ValueError:
                pass
        elif f == "qty":
            try:
                row.qty = max(1, int(float(str(v).replace(",", "."))))
            except ValueError:
                pass
        else:
            setattr(row, f, str(v))
        new = getattr(row, f, None)
        if old != new:
            changes.append(make_change(
                f"equipment.{f}", old, new,
                old_display=EQ_LABELS.get(old, old) if f == "status" and old else None,
                new_display=EQ_LABELS.get(new, new) if f == "status" and new else None,
            ))
    row.updated_at = datetime.now(timezone.utc)
    await db.flush()
    if changes:
        await log_event(
            db, site, "edit", user=user, changes=changes,
            text=(f"Оборудование в план: {row.title or '—'}"
                  if created else f"Оборудование «{row.title or '—'}»: ")
                 + ", ".join(item["label"] for item in changes),
        )
    elif created:
        await log_event(db, site, "note", user=user,
                        text=f"Оборудование в план: {row.title or '—'} × {row.qty}")
    return _eq_out(row)


async def delete_equipment(
    db: AsyncSession, company_id, site_id, eq_id,
    *, site: EzsSite | None = None, user: User | None = None,
) -> bool:
    row = (await db.execute(select(EzsSiteEquipment).where(
        EzsSiteEquipment.company_id == company_id, EzsSiteEquipment.site_id == site_id,
        EzsSiteEquipment.id == eq_id))).scalar_one_or_none()
    if row is None:
        return False
    if site is not None:
        label = row.title or "Позиция оборудования"
        await log_event(
            db, site, "edit", user=user, text=f"Удалено оборудование: {label}",
            changes=[make_change(
                "equipment.status", row.status, None,
                label=f"Оборудование: {label}", category="technical",
                old_display=EQ_LABELS.get(row.status, row.status),
                new_display="удалено",
            )],
        )
    await db.delete(row)
    return True


async def equipment_report(db: AsyncSession, company_id) -> dict[str, Any]:
    """Сводный реестр потребности в оборудовании по всем проектам."""
    rows = (await db.execute(
        select(EzsSiteEquipment, EzsSite.project_no, EzsSite.title, EzsSite.city,
               EzsSite.address, EzsSite.stage)
        .join(EzsSite, EzsSite.id == EzsSiteEquipment.site_id)
        .where(EzsSiteEquipment.company_id == company_id)
        .order_by(EzsSiteEquipment.due_date.nulls_last()))).all()
    items, by_status, overdue, total_price, total_qty = [], {}, 0, 0.0, 0
    for e, pno, title, city, addr, stage in rows:
        out = _eq_out(e)
        out.update({"projectNo": pno, "projectTitle": title, "city": city, "address": addr,
                    "stage": stage, "stageLabel": STAGE_LABELS.get(stage, stage)})
        items.append(out)
        by_status[e.status] = by_status.get(e.status, 0) + 1
        overdue += 1 if out["overdue"] else 0
        if e.status != "cancelled":
            total_price += (out["price"] or 0) * (e.qty or 1)
            total_qty += e.qty or 1
    return {
        "total": len(items), "overdue": overdue, "qty": total_qty,
        "priceTotal": round(total_price, 2),
        "byStatus": [{"key": s["key"], "label": s["label"], "count": by_status.get(s["key"], 0)}
                     for s in EQ_STATUSES],
        "items": items,
    }


# ── Бюджет ─────────────────────────────────────────────────────────────────
# `capital` — попадает ли статья в стоимость будущего объекта (счёт 08 → 01) или
# остаётся расходом периода. Это не оформление, а развилка учёта: по ФСБУ 26/2020
# в капвложения идёт то, что приводит объект в состояние, пригодное к
# использованию, а управленческие расходы и потери от простоев (п. 16) — нет.
# У отменённого проекта капитализированное списывается (Дт 91.02 Кт 08), расходы
# периода списаны уже тогда, когда были понесены.
COST_KINDS = [
    {"key": "tp", "label": "Техприсоединение", "capital": True},
    {"key": "equipment", "label": "Оборудование", "capital": True},
    {"key": "smr", "label": "СМР и монтаж", "capital": True},
    {"key": "design", "label": "Проектирование", "capital": True},
    {"key": "rent", "label": "Аренда (за период стройки)", "capital": True},
    {"key": "survey", "label": "Изыскания и экспертизы", "capital": True},
    {"key": "admin", "label": "Сопровождение и администрирование", "capital": False},
    {"key": "other", "label": "Прочее", "capital": False},
]
COST_LABELS = {c["key"]: c["label"] for c in COST_KINDS}
COST_CAPITAL = {c["key"]: bool(c["capital"]) for c in COST_KINDS}


async def list_costs(db: AsyncSession, company_id, site_id) -> dict[str, Any]:
    rows = (await db.execute(select(EzsSiteCost).where(
        EzsSiteCost.company_id == company_id, EzsSiteCost.site_id == site_id)
        .order_by(EzsSiteCost.created_at))).scalars().all()
    items = [{
        "id": str(c.id), "kind": c.kind, "kindLabel": COST_LABELS.get(c.kind, c.kind),
        "capital": COST_CAPITAL.get(c.kind, False),
        "title": c.title, "docRef": c.doc_ref, "note": c.note,
        "plan": float(c.plan_amount) if c.plan_amount is not None else None,
        "fact": float(c.fact_amount) if c.fact_amount is not None else None,
    } for c in rows]

    def total(field: str, capital: bool | None = None) -> float:
        return round(sum(i[field] or 0 for i in items
                         if capital is None or i["capital"] is capital), 2)

    return {
        "items": items,
        "planTotal": total("plan"), "factTotal": total("fact"),
        # Разделение для моста в учёт: капитализируемое пойдёт в стоимость
        # объекта, расходы периода — нет; при отмене проекта их судьба разная.
        "capitalPlan": total("plan", True), "capitalFact": total("fact", True),
        "expensePlan": total("plan", False), "expenseFact": total("fact", False),
    }


async def costs_report(db: AsyncSession, company_id) -> dict[str, Any]:
    """Бюджет портфеля: по статьям, с отклонением плана и факта.

    Отдельно капвложения и расходы периода — это разные судьбы денег, и в одну
    сумму их складывать нельзя. Отдельно — «в работе» и «отменённые»: у вторых
    капитализированные затраты подлежат списанию, и сумма по ним отвечает на
    вопрос «во что обошлись несостоявшиеся проекты».
    """
    rows = (await db.execute(text("""
        select c.kind,
               case when s.stage = any(:active) then 'active'
                    when s.stage = 'on_hold' then 'on_hold' else 'closed' end as bucket,
               count(distinct c.site_id) sites,
               coalesce(sum(c.plan_amount), 0) plan,
               coalesce(sum(c.fact_amount), 0) fact
        from ezs_site_costs c join ezs_sites s on s.id = c.site_id
        where c.company_id = :cid
        group by 1, 2
    """), {"cid": company_id, "active": STAGE_ORDER})).mappings().all()

    by_kind: dict[str, dict[str, Any]] = {}
    buckets = {"active": {"plan": 0.0, "fact": 0.0}, "on_hold": {"plan": 0.0, "fact": 0.0},
               "closed": {"plan": 0.0, "fact": 0.0}}
    for r in rows:
        k = r["kind"]
        item = by_kind.setdefault(k, {
            "kind": k, "label": COST_LABELS.get(k, k), "capital": COST_CAPITAL.get(k, False),
            "plan": 0.0, "fact": 0.0, "sites": 0,
        })
        item["plan"] += float(r["plan"] or 0)
        item["fact"] += float(r["fact"] or 0)
        item["sites"] += int(r["sites"] or 0)
        b = buckets[r["bucket"]]
        b["plan"] += float(r["plan"] or 0)
        b["fact"] += float(r["fact"] or 0)

    items = []
    for it in by_kind.values():
        it["plan"] = round(it["plan"], 2)
        it["fact"] = round(it["fact"], 2)
        it["variance"] = round(it["fact"] - it["plan"], 2)
        # Процент отклонения без плана бессмыслен: делить не на что.
        it["variancePct"] = round(it["variance"] / it["plan"] * 100, 1) if it["plan"] else None
        items.append(it)
    items.sort(key=lambda i: -(i["plan"] or 0))

    cap = {"plan": round(sum(i["plan"] for i in items if i["capital"]), 2),
           "fact": round(sum(i["fact"] for i in items if i["capital"]), 2)}
    exp = {"plan": round(sum(i["plan"] for i in items if not i["capital"]), 2),
           "fact": round(sum(i["fact"] for i in items if not i["capital"]), 2)}
    return {
        "items": items,
        "capital": cap, "expense": exp,
        "buckets": [
            {"key": "active", "label": "В работе", **{k: round(v, 2) for k, v in buckets["active"].items()}},
            {"key": "on_hold", "label": "Приостановлены (остаются на счёте 08)",
             **{k: round(v, 2) for k, v in buckets["on_hold"].items()}},
            {"key": "closed", "label": "Отменены (подлежат списанию)",
             **{k: round(v, 2) for k, v in buckets["closed"].items()}},
        ],
        "planTotal": round(cap["plan"] + exp["plan"], 2),
        "factTotal": round(cap["fact"] + exp["fact"], 2),
    }


async def upsert_cost(db: AsyncSession, company_id, site: EzsSite, payload: dict[str, Any],
                      user: User | None) -> dict[str, Any]:
    cid = payload.get("id")
    row = None
    if cid:
        row = (await db.execute(select(EzsSiteCost).where(
            EzsSiteCost.company_id == company_id, EzsSiteCost.site_id == site.id,
            EzsSiteCost.id == _uuid.UUID(str(cid))))).scalar_one_or_none()
    created = row is None
    if row is None:
        row = EzsSiteCost(company_id=company_id, site_id=site.id)
        db.add(row)
    old_kind = row.kind
    old_details = (row.title, row.doc_ref, row.note)
    was = (row.plan_amount, row.fact_amount)
    row.kind = str(payload.get("kind") or "other")
    row.title = payload.get("title") or None
    row.doc_ref = payload.get("doc_ref") or None
    row.note = payload.get("note") or None
    for f, key in (("plan_amount", "plan"), ("fact_amount", "fact")):
        v = payload.get(key)
        if v in ("", None):
            setattr(row, f, None)
        else:
            try:
                setattr(row, f, float(str(v).replace(",", ".").replace(" ", "")))
            except ValueError:
                setattr(row, f, None)
    # Бюджет — единственное, что правилось молча: кто и на сколько поменял сумму,
    # восстановить было нельзя. Пишем суммы «было → стало» прямо в текст события.
    label = COST_LABELS.get(row.kind, row.kind)
    changes = []
    if old_kind != row.kind:
        changes.append(make_change(
            "budget.kind", old_kind, row.kind,
            old_display=COST_LABELS.get(old_kind, old_kind) if old_kind else None,
            new_display=label,
        ))
    for field, old, new in (
        ("budget.title", old_details[0], row.title),
        ("budget.doc_ref", old_details[1], row.doc_ref),
        ("budget.note", old_details[2], row.note),
    ):
        if old != new:
            changes.append(make_change(field, old, new))
    if was[0] != row.plan_amount:
        changes.append(make_change("budget.plan", was[0], row.plan_amount))
    if was[1] != row.fact_amount:
        changes.append(make_change("budget.fact", was[1], row.fact_amount))
    if changes:
        await log_event(
            db, site, "edit", user=user, changes=changes,
            text=(f"Бюджет: добавлена статья «{label}» — "
                  if created else f"Бюджет «{label}»: ")
                 + ", ".join(item["label"] for item in changes),
        )
    await db.flush()
    return {"id": str(row.id)}


def _money(v: Any) -> str:
    return "—" if v is None else f"{float(v):,.0f} ₽".replace(",", " ")


async def delete_cost(db: AsyncSession, company_id, site_id, cost_id,
                      site: EzsSite | None = None, user: User | None = None) -> bool:
    row = (await db.execute(select(EzsSiteCost).where(
        EzsSiteCost.company_id == company_id, EzsSiteCost.site_id == site_id,
        EzsSiteCost.id == cost_id))).scalar_one_or_none()
    if row is None:
        return False
    if site is not None:
        label = COST_LABELS.get(row.kind, row.kind)
        changes = [make_change(
            "budget.kind", row.kind, None,
            old_display=label, new_display="статья удалена",
        )]
        if row.plan_amount is not None:
            changes.append(make_change("budget.plan", row.plan_amount, None))
        if row.fact_amount is not None:
            changes.append(make_change("budget.fact", row.fact_amount, None))
        await log_event(db, site, "note", user=user,
                        text=f"Бюджет: удалена статья «{label}» "
                             f"(план {_money(row.plan_amount)}, факт {_money(row.fact_amount)})",
                        changes=changes)
    await db.delete(row)
    return True


# ── Субсидия: требования программы как чек-лист ────────────────────────────
SUBSIDY_MIN_POWER_KWT = 149.0
SUBSIDY_MIN_SPOTS = 2
SUBSIDY_YEARS = 5


def subsidy_check(site: EzsSite) -> dict[str, Any]:
    """Соответствие требованиям господдержки — проверяется до закупки, а не после.

    Требования программы: мощность от 149 кВт, круглосуточный доступ, не менее
    двух машино-мест, на трассе — в составе МФЗ/АЗС с сервисом; после ввода —
    обязательство эксплуатировать пять лет.
    """
    power = site.planned_power_kwt
    items = [
        {"key": "power", "label": f"Мощность от {SUBSIDY_MIN_POWER_KWT:.0f} кВт",
         "done": bool(power and power >= SUBSIDY_MIN_POWER_KWT),
         "value": f"{power:.0f} кВт" if power else None},
        {"key": "spots", "label": f"Машино-мест для зарядки: от {SUBSIDY_MIN_SPOTS}",
         "done": bool(site.parking_spots and site.parking_spots >= SUBSIDY_MIN_SPOTS),
         "value": str(site.parking_spots) if site.parking_spots else None},
        {"key": "access", "label": "Круглосуточный доступ", "done": bool(site.access_24x7),
         "value": None},
        {"key": "ports", "label": "Несколько разъёмов",
         "done": sum(1 for p in (site.ports_gbt, site.ports_ccs, site.ports_chademo,
                                 site.ports_type) if p) >= 2, "value": None},
        {"key": "service", "label": "На трассе — сервис рядом (магазин, кафе)",
         "done": site.place_kind != "трасса" or bool(site.dop_service), "value": site.dop_service},
    ]
    obligation_until = None
    if site.commissioned_on:
        try:
            y, m, d = (int(x) for x in site.commissioned_on.split("-")[:3])
            obligation_until = f"{y + SUBSIDY_YEARS:04d}-{m:02d}-{d:02d}"
        except (ValueError, TypeError):
            obligation_until = None
    return {
        "planned": bool(site.subsidy_planned),
        "amount": float(site.subsidy_amount) if site.subsidy_amount is not None else None,
        "items": items,
        "done": sum(1 for i in items if i["done"]), "total": len(items),
        "eligible": all(i["done"] for i in items),
        "commissionedOn": site.commissioned_on,
        "obligationUntil": obligation_until,
        "obligationYears": SUBSIDY_YEARS,
    }


# ── Портфель и мост в учёт ─────────────────────────────────────────────────
async def portfolio(db: AsyncSession, company_id) -> dict[str, Any]:
    """Обзор портфеля: этапы, сроки, бюджет, риски."""
    S = EzsSite
    by_stage = {r.stage: int(r.n) for r in (await db.execute(
        select(S.stage, func.count().label("n")).where(S.company_id == company_id)
        .group_by(S.stage))).all()}
    phases = []
    for p in PHASES:
        phases.append({"key": p["key"], "label": p["label"], "hint": p["hint"],
                       "count": sum(by_stage.get(s, 0) for s in p["stages"]),
                       "stages": [{"stage": s, "label": STAGE_LABELS[s],
                                   "count": by_stage.get(s, 0)} for s in p["stages"]]})

    # Деньги портфеля — про то, что в работе. Затраты отклонённых и замороженных
    # проектов в общей сумме растворяли реальный бюджет и завышали его.
    money = (await db.execute(text("""
        select coalesce(sum(c.plan_amount), 0) plan, coalesce(sum(c.fact_amount), 0) fact
        from ezs_site_costs c join ezs_sites s on s.id = c.site_id
        where c.company_id = :cid and s.stage <> all(:closed)
    """), {"cid": company_id, "closed": ["archive", "on_hold"]})).mappings().one()

    tc = (await db.execute(text("""
        select count(*) total,
               count(*) filter (where status = 'done') done,
               count(*) filter (where due_date is not null and done_date is null
                                and due_date < to_char(now(), 'YYYY-MM-DD')
                                and status not in ('done','rejected')) overdue
        from ezs_tech_connections where company_id = :cid
    """), {"cid": company_id})).mappings().one()

    # Сколько проектов дошло до работающей станции — главный KPI развития.
    realized = (await db.execute(select(func.count()).select_from(S).where(
        S.company_id == company_id, S.location_id.is_not(None)))).scalar_one()

    docs = (await db.execute(select(func.count()).select_from(EzsSiteDoc)
                             .where(EzsSiteDoc.company_id == company_id))).scalar_one()

    eq = (await db.execute(text("""
        select count(*) total,
               count(*) filter (where status in ('supplied','installed')) supplied,
               count(*) filter (where due_date is not null and supplied_date is null
                                and status in ('planned','ordered')
                                and due_date < to_char(now(), 'YYYY-MM-DD')) overdue
        from ezs_site_equipment where company_id = :cid
    """), {"cid": company_id})).mappings().one()

    return {
        "phases": phases,
        "active": sum(by_stage.get(s, 0) for s in STAGE_ORDER),
        "total": sum(by_stage.values()),
        "realized": int(realized or 0),
        "budget": {"plan": float(money["plan"] or 0), "fact": float(money["fact"] or 0)},
        "techConnections": {"total": int(tc["total"] or 0), "done": int(tc["done"] or 0),
                            "overdue": int(tc["overdue"] or 0)},
        "docs": int(docs or 0),
        "equipment": {"total": int(eq["total"] or 0), "supplied": int(eq["supplied"] or 0),
                      "overdue": int(eq["overdue"] or 0)},
    }


async def portfolio_overview(db: AsyncSession, company_id) -> dict[str, Any]:
    """Обзор портфеля как рабочий экран, а не витрина.

    Отвечает на четыре вопроса руководителя развития:
      1. что горит — просрочки и заблокированные гейты, с числами и ссылками;
      2. где затык — на какой стадии проекты стоят дольше всего и куда не проходят;
      3. когда ждать станции — прогноз ввода по датам ТП и поставок;
      4. что изменилось — движение портфеля за 30 и 90 дней.

    Блоки, для которых данных нет, возвращают `empty` с подсказкой, чем их
    наполнить: экран из нулей хуже честного «нечего показывать».
    """
    today = date.today().isoformat()
    d30 = (date.today() - timedelta(days=30)).isoformat()
    d90 = (date.today() - timedelta(days=90)).isoformat()

    # ── 1. Что горит ───────────────────────────────────────────────────────
    risks = (await db.execute(text(f"""
        with active as (
            select s.* from ezs_sites s
            where s.company_id = :cid and s.stage = any(:active)
        ),
        stage_events as (
            select e.site_id, e.from_stage, e.to_stage
            from ezs_site_events e join active a on a.id = e.site_id
            where e.company_id = :cid and e.kind = 'stage'
        ),
        event_stats as (
            select site_id, count(*) as stage_events,
                   bool_or({_stage_position_sql("e.to_stage")} >= 0
                       and {_stage_position_sql("e.from_stage")} >= 0
                       and {_stage_position_sql("e.to_stage")}
                           < {_stage_position_sql("e.from_stage")}) as has_rework
            from stage_events e group by site_id
        ),
        repeat_stats as (
            select site_id, sum(entries - 1) as repeat_entries
            from (
                select site_id, to_stage, count(*) as entries
                from stage_events where to_stage is not null
                group by site_id, to_stage having count(*) > 1
            ) x group by site_id
        ),
        flags as (
            select a.*,
                   coalesce(es.stage_events, 0) as stage_events,
                   coalesce(es.has_rework, false) as has_rework,
                   coalesce(rs.repeat_entries, 0) as repeat_entries,
                   not exists (
                       select 1 from ezs_site_participants p where p.site_id = a.id
                   ) as no_participants,
                   (
                       (coalesce(a.commissioned_on, '') <> '' and a.stage <> 'live')
                       or (a.stage = 'live' and coalesce(a.commissioned_on, '') = '')
                       or (a.kind = 'decommission' and coalesce(a.commissioned_on, '') <> '')
                   ) as commissioning_mismatch,
                   (a.next_action_due is not null and a.next_action_due < :today) as step_overdue,
                   exists (
                       select 1 from ezs_tech_connections t where t.site_id = a.id
                         and t.due_date is not null and t.done_date is null
                         and t.due_date < :today and t.status not in ('done','rejected')
                   ) as tp_overdue,
                   exists (
                       select 1 from ezs_site_equipment eq where eq.site_id = a.id
                         and eq.due_date is not null and eq.supplied_date is null
                         and eq.status in ('planned','ordered') and eq.due_date < :today
                   ) as eq_overdue
            from active a
            left join event_stats es on es.site_id = a.id
            left join repeat_stats rs on rs.site_id = a.id
        )
        select
          count(*) filter (where step_overdue) as step_overdue,
          count(*) filter (where tp_overdue) as tp_overdue,
          count(*) filter (where eq_overdue) as eq_overdue,
          count(*) filter (where owner_user_id is null) as no_owner,
          count(*) filter (where next_action is null) as no_next,
          count(*) filter (where stage_since is not null
             and (current_date - stage_since::date) > {_NORM_CASE}) as stage_overdue,
          count(*) filter (where coalesce(to_char(last_touch_at, 'YYYY-MM-DD'),
             '1970-01-01') < :d30) as no_touch_30,
          count(*) filter (where no_participants) as no_participants,
          count(*) filter (where has_rework) as rework,
          count(*) filter (where commissioning_mismatch) as commissioning_mismatch,
          count(*) filter (where step_overdue or tp_overdue or eq_overdue) as at_risk,
          count(*) filter (where step_overdue or tp_overdue or eq_overdue
             or owner_user_id is null or next_action is null
             or (stage_since is not null
                 and (current_date - stage_since::date) > {_NORM_CASE})
             or coalesce(to_char(last_touch_at, 'YYYY-MM-DD'), '1970-01-01') < :d30
             or no_participants or has_rework or commissioning_mismatch) as attention_total,
          count(*) as active_total,
          count(*) filter (where stage_events > 0) as with_history,
          count(*) filter (where stage_events = 0) as without_history,
          coalesce(sum(stage_events), 0) as stage_events,
          coalesce(sum(repeat_entries), 0) as repeat_entries,
          count(*) filter (where has_rework) as rework_projects
        from flags
    """), {"cid": company_id, "active": STAGE_ORDER, "today": today,
           "d30": d30})).mappings().one()

    attention = [
        {"key": "step_overdue", "label": "Просрочен следующий шаг", "count": int(risks["step_overdue"]),
         "hint": "срок в карточке прошёл, шаг не сделан", "filter": "overdue", "tone": "critical"},
        {"key": "tp_overdue", "label": "Просрочено техприсоединение", "count": int(risks["tp_overdue"]),
         "hint": "срок мероприятий сетевой прошёл", "filter": "tp", "tone": "critical"},
        {"key": "eq_overdue", "label": "Просрочена поставка оборудования", "count": int(risks["eq_overdue"]),
         "hint": "плановая дата поставки прошла", "filter": "equipment", "tone": "critical"},
        {"key": "commissioning_mismatch", "label": "Дата ввода расходится со стадией",
         "count": int(risks["commissioning_mismatch"]),
         "hint": "проверьте основание перевода капвложений в основные средства",
         "filter": "", "tone": "critical"},
        {"key": "stage_overdue", "label": "Стадия идёт дольше норматива", "count": int(risks["stage_overdue"]),
         # На свежезагруженном портфеле stage_since — это дата поступления из файла,
         # а не дата входа в стадию внутри системы. Формально проект действительно
         # стоит дольше норматива, но приписка «движения нет» звучала бы упрёком за
         # то, что систему только начали вести.
         "hint": "проект стоит на стадии дольше норматива регламента", "filter": "", "tone": "warning"},
        {"key": "no_touch_30", "label": "Без касаний больше 30 дней", "count": int(risks["no_touch_30"]),
         "hint": "нет записи о звонке, встрече или письме за 30 дней", "filter": "", "tone": "warning"},
        {"key": "rework", "label": "Возвращались на более раннюю стадию",
         "count": int(risks["rework"]),
         "hint": "есть обратный переход — проверьте причину переделки", "filter": "", "tone": "warning"},
        {"key": "no_owner", "label": "Без ответственного", "count": int(risks["no_owner"]),
         "hint": "проект не закреплён ни за кем — назначьте в реестре", "filter": "", "tone": "warning"},
        {"key": "no_next", "label": "Без следующего шага", "count": int(risks["no_next"]),
         "hint": "не записано, что делать дальше и к какому сроку", "filter": "", "tone": "warning"},
        {"key": "no_participants", "label": "Не назначен состав проекта",
         "count": int(risks["no_participants"]),
         "hint": "нет ни одного участника с ролью по регламенту", "filter": "", "tone": "warning"},
    ]

    # ── 2. Где затык: воронка со сроком и проходимостью ────────────────────
    # Проходимость считаем по истории: сколько проектов, побывавших на стадии,
    # ушли ВПЕРЁД. Именно вперёд: раньше засчитывалось любое более позднее
    # событие, поэтому возврат с пусконаладки на СМР, пауза и отказ в архив
    # читались как «стадия пройдена» — метрика показывала успех там, где проект
    # развернули назад.
    passed = {r["stage"]: r for r in (await db.execute(text(f"""
        with stage_events as (
            select id, site_id, from_stage, to_stage, created_at
            from ezs_site_events
            where company_id = :cid and kind = 'stage'
        ),
        explicit_entries as (
            select site_id, to_stage as stage, count(*) as entries
            from stage_events where to_stage is not null
            group by site_id, to_stage
        ),
        implicit_entries as (
            select distinct e.site_id, e.from_stage as stage, 1 as entries
            from stage_events e
            where e.from_stage is not null and not exists (
                select 1 from stage_events prior
                where prior.site_id = e.site_id and prior.to_stage = e.from_stage
                  and (prior.created_at, prior.id) < (e.created_at, e.id)
            )
        ),
        visits as (
            select site_id, stage, sum(entries) as entries
            from (
                select * from explicit_entries
                union all
                select * from implicit_entries
            ) entered
            group by site_id, stage
        ),
        outcomes as (
            select site_id, from_stage as stage,
                   bool_or({_stage_position_sql("e.to_stage")}
                       > {_stage_position_sql("e.from_stage")}) as advanced,
                   bool_or({_stage_position_sql("e.to_stage")} >= 0
                       and {_stage_position_sql("e.from_stage")} >= 0
                       and {_stage_position_sql("e.to_stage")}
                           < {_stage_position_sql("e.from_stage")}) as returned,
                   bool_or(to_stage = 'on_hold') as paused,
                   bool_or(to_stage = 'archive') as archived
            from stage_events e where from_stage is not null
            group by site_id, from_stage
        )
        select v.stage, count(*) as visited,
               count(*) filter (where coalesce(o.advanced, false)) as advanced,
               count(*) filter (where coalesce(o.advanced, false)
                   and v.entries = 1) as first_pass,
               coalesce(sum(greatest(v.entries - 1, 0)), 0) as reentries,
               count(*) filter (where coalesce(o.returned, false)) as returned,
               count(*) filter (where coalesce(o.paused, false)) as paused,
               count(*) filter (where coalesce(o.archived, false)) as archived
        from visits v
        left join outcomes o on o.site_id = v.site_id and o.stage = v.stage
        group by v.stage
    """), {"cid": company_id})).mappings().all()}

    durations = (await phase_durations(db, company_id))["stages"]
    dur_by_stage = {d["stage"]: d for d in durations}

    counts = {r.stage: int(r.n) for r in (await db.execute(
        select(EzsSite.stage, func.count().label("n"))
        .where(EzsSite.company_id == company_id).group_by(EzsSite.stage))).all()}

    # Застрявшие — по нормативу своей стадии (см. `STAGE_NORM_DAYS`).
    stuck_by_stage = {r["stage"]: int(r["n"]) for r in (await db.execute(text("""
        select stage, count(*) n from ezs_sites
        where company_id = :cid and stage = any(:active) and stage_since is not null
          and (current_date - stage_since::date) > """ + _NORM_CASE + """
        group by 1
    """), {"cid": company_id, "active": STAGE_ORDER})).mappings().all()}

    funnel = []
    for st in STAGE_ORDER:
        v = passed.get(st, {})
        visited = int(v.get("visited") or 0)
        advanced = int(v.get("advanced") or 0)
        first_pass = int(v.get("first_pass") or 0)
        funnel.append({
            "stage": st, "label": STAGE_LABELS[st],
            "phase": STAGE_PHASE.get(st), "count": counts.get(st, 0),
            "medianDays": dur_by_stage.get(st, {}).get("medianDays", 0),
            "p85Days": dur_by_stage.get(st, {}).get("p85Days", 0),
            "open": dur_by_stage.get(st, {}).get("open", 0),
            "openMedianDays": dur_by_stage.get(st, {}).get("openMedianDays", 0),
            "visited": visited, "advanced": advanced,
            "conversion": round(advanced / visited * 100) if visited else None,
            "firstPass": first_pass,
            "firstPassRate": round(first_pass / advanced * 100) if advanced else None,
            "reentries": int(v.get("reentries") or 0),
            "returned": int(v.get("returned") or 0),
            "paused": int(v.get("paused") or 0),
            "archivedFromStage": int(v.get("archived") or 0),
            "stuck": stuck_by_stage.get(st, 0),
            "normDays": norm_days(st),
        })
    # Узкое место — стадия с наибольшим числом застрявших; при равенстве берём
    # ту, где дольше медиана. Считаем только по активным стадиям с проектами.
    live = [f for f in funnel if f["count"] > 0]
    bottleneck = max(live, key=lambda f: (f["stuck"], f["medianDays"]), default=None)

    # ── 3. Когда ждать станции ─────────────────────────────────────────────
    forecast = (await db.execute(text("""
        with dates as (
            select s.id,
                   greatest(
                     coalesce((select max(t.due_date) from ezs_tech_connections t
                               where t.site_id = s.id and t.done_date is null), ''),
                     coalesce((select max(e.due_date) from ezs_site_equipment e
                               where e.site_id = s.id and e.supplied_date is null), '')
                   ) as ready_date
            from ezs_sites s
            where s.company_id = :cid and s.stage in ('construction','commissioning')
        )
        select case when ready_date = '' then 'без даты'
                    else to_char(to_date(ready_date, 'YYYY-MM-DD'), 'YYYY') || ' Q'
                         || to_char(to_date(ready_date, 'YYYY-MM-DD'), 'Q') end as bucket,
               count(*) as n
        from dates group by 1 order by 1
    """), {"cid": company_id})).mappings().all()

    commissioned = (await db.execute(text("""
        select to_char(to_date(commissioned_on, 'YYYY-MM-DD'), 'YYYY') || ' Q'
               || to_char(to_date(commissioned_on, 'YYYY-MM-DD'), 'Q') as bucket, count(*) n
        from ezs_sites
        where company_id = :cid and commissioned_on is not null and commissioned_on <> ''
        group by 1 order by 1
    """), {"cid": company_id})).mappings().all()

    # ── 4. Что изменилось ──────────────────────────────────────────────────
    movement = (await db.execute(text("""
        select
          (select count(*) from ezs_sites
             where company_id = :cid and first_seen_at >= now() - interval '30 days')  as added_30,
          (select count(*) from ezs_sites
             where company_id = :cid and first_seen_at >= now() - interval '90 days')  as added_90,
          (select count(distinct site_id) from ezs_site_events
             where company_id = :cid and kind = 'stage'
               and created_at >= now() - interval '30 days'
               and to_stage <> 'archive')                                               as moved_30,
          (select count(distinct site_id) from ezs_site_events
             where company_id = :cid and kind = 'stage' and to_stage = 'archive'
               and created_at >= now() - interval '30 days')                            as archived_30,
          (select count(distinct site_id) from ezs_site_events
             where company_id = :cid and kind = 'stage' and to_stage = 'live'
               and created_at >= now() - interval '90 days')                            as live_90,
          (select count(*) from ezs_site_events
             where company_id = :cid and kind in ('touch','note')
               and created_at >= now() - interval '30 days')                            as touches_30
    """), {"cid": company_id})).mappings().one()

    # ── 5. Кто ведёт ───────────────────────────────────────────────────────
    owners = (await db.execute(text("""
        select coalesce(u.name, u.email, '— не назначен') as owner,
               count(*) as projects,
               count(*) filter (where s.next_action_due is not null
                                 and s.next_action_due < :today) as overdue
        from ezs_sites s
        left join users u on u.id = s.owner_user_id
        where s.company_id = :cid and s.stage = any(:active)
        group by 1 order by projects desc limit 10
    """), {"cid": company_id, "active": STAGE_ORDER, "today": today})).mappings().all()

    # ── 6. Деньги ──────────────────────────────────────────────────────────
    # Только проекты в работе: затраты отклонённых и замороженных — это другая
    # история (по ФСБУ 26/2020 их судьба решается отдельно), и в бюджете портфеля
    # они завышали цифру.
    money = (await db.execute(text("""
        select coalesce(sum(c.plan_amount), 0) plan, coalesce(sum(c.fact_amount), 0) fact,
               count(distinct c.site_id) sites
        from ezs_site_costs c join ezs_sites s on s.id = c.site_id
        where c.company_id = :cid and s.stage = any(:active)
    """), {"cid": company_id, "active": STAGE_ORDER})).mappings().one()
    eq_money = (await db.execute(text("""
        select coalesce(sum(e.price * e.qty), 0) total
        from ezs_site_equipment e join ezs_sites s on s.id = e.site_id
        where e.company_id = :cid and e.status <> 'cancelled' and s.stage = any(:active)
    """), {"cid": company_id, "active": STAGE_ORDER})).scalar()

    active_total = sum(counts.get(s, 0) for s in STAGE_ORDER)
    measurement_total = int(risks["active_total"] or 0)
    with_history = int(risks["with_history"] or 0)
    kind_rows = (await db.execute(text("""
        select coalesce(kind, 'new_build') as kind, count(*) as n
        from ezs_sites
        where company_id = :cid and stage = any(:active)
        group by 1 order by 2 desc
    """), {"cid": company_id, "active": STAGE_ORDER})).mappings().all()
    return {
        "active": active_total,
        "total": sum(counts.values()),
        "live": counts.get("live", 0),
        "archived": counts.get("archive", 0),
        "onHold": counts.get("on_hold", 0),
        "attention": [a for a in attention if a["count"] > 0],
        "attentionAll": attention,
        "atRisk": int(risks["at_risk"] or 0),
        "attentionTotal": int(risks["attention_total"] or 0),
        "measurement": {
            "withHistory": with_history,
            "withoutHistory": int(risks["without_history"] or 0),
            "coverage": round(with_history / measurement_total * 100) if measurement_total else 0,
            "stageEvents": int(risks["stage_events"] or 0),
            "repeatEntries": int(risks["repeat_entries"] or 0),
            "reworkProjects": int(risks["rework_projects"] or 0),
            "kinds": [{"kind": r["kind"], "count": int(r["n"])} for r in kind_rows],
        },
        "funnel": funnel,
        "bottleneck": bottleneck,
        "forecast": [{"bucket": r["bucket"], "count": int(r["n"])} for r in forecast],
        "commissioned": [{"bucket": r["bucket"], "count": int(r["n"])} for r in commissioned],
        "movement": {k: int(v or 0) for k, v in dict(movement).items()},
        "owners": [{"owner": r["owner"], "projects": int(r["projects"]),
                    "overdue": int(r["overdue"])} for r in owners],
        "budget": {"plan": float(money["plan"] or 0), "fact": float(money["fact"] or 0),
                   "sites": int(money["sites"] or 0),
                   "equipment": float(eq_money or 0)},
        "phases": [{"key": p["key"], "label": p["label"], "hint": p["hint"],
                    "count": sum(counts.get(s, 0) for s in p["stages"])} for p in PHASES],
    }


async def phase_durations(db: AsyncSession, company_id) -> dict[str, Any]:
    """Сколько проекты стоят на этапах — по датам входа в стадию из истории.

    Берём разницу между соседними сменами стадий: сколько проект реально провёл
    в каждой, а не сколько числится сейчас. Для текущей стадии — время с момента
    входа до сегодня, такие записи помечены `open`.
    """
    rows = (await db.execute(text("""
        with moves as (
            select e.site_id, e.to_stage as stage, e.created_at,
                   lead(e.created_at) over (partition by e.site_id order by e.created_at) as next_at
            from ezs_site_events e
            where e.company_id = :cid and e.kind = 'stage' and e.to_stage is not null
        )
        select stage, count(*) as n,
               count(*) filter (where next_at is not null) as completed,
               percentile_cont(0.5) within group (
                   order by extract(epoch from (next_at - created_at)) / 86400
               ) filter (where next_at is not null) as median_days,
               percentile_cont(0.85) within group (
                   order by extract(epoch from (next_at - created_at)) / 86400
               ) filter (where next_at is not null) as p85_days,
               count(*) filter (where next_at is null) as still_open,
               percentile_cont(0.5) within group (
                   order by extract(epoch from (now() - created_at)) / 86400
               ) filter (where next_at is null) as open_median_days
        from moves group by stage
    """), {"cid": company_id})).mappings().all()
    by_stage = {r["stage"]: {"count": int(r["n"]),
                             "completed": int(r["completed"]),
                             "medianDays": round(float(r["median_days"] or 0), 1),
                             "p85Days": round(float(r["p85_days"] or 0), 1),
                             "open": int(r["still_open"]),
                             "openMedianDays": round(float(r["open_median_days"] or 0), 1)}
                for r in rows}
    return {
        "stages": [{"stage": s, "label": STAGE_LABELS[s],
                    **by_stage.get(s, {"count": 0, "completed": 0, "medianDays": 0,
                                       "p85Days": 0, "open": 0, "openMedianDays": 0})}
                   for s in STAGE_ORDER],
        "note": "медиана и P85 — только по завершённым визитам; открытые стадии показаны отдельно",
    }


async def export_portfolio_xlsx(db: AsyncSession, company_id) -> bytes:
    """Выгрузка портфеля: проекты с этапом, ведением, ТП, бюджетом и субсидией."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    rows = (await db.execute(text("""
        select s.project_no, s.title, coalesce(s.region_norm, s.region) as region, s.city,
               coalesce(s.address, s.full_address, s.install_place) as address,
               s.stage, s.stage_since, s.owner, u.name as owner_user,
               s.next_action, s.next_action_due,
               s.control_form, s.contract_start, s.contract_end,
               s.planned_power_kwt, s.parking_spots, s.subsidy_planned, s.commissioned_on,
               tc.status as tc_status, tc.grid_operator, tc.due_date as tc_due, tc.done_date as tc_done,
               tc.cost as tc_cost,
               (select coalesce(sum(c.plan_amount), 0) from ezs_site_costs c where c.site_id = s.id) as plan_amount,
               (select coalesce(sum(c.fact_amount), 0) from ezs_site_costs c where c.site_id = s.id) as fact_amount,
               (select count(*) from ezs_site_docs d where d.site_id = s.id) as docs
        from ezs_sites s
        left join users u on u.id = s.owner_user_id
        left join ezs_tech_connections tc on tc.site_id = s.id
        where s.company_id = :cid and s.stage = any(:active)
        order by s.project_no
    """), {"cid": company_id, "active": STAGE_ORDER})).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Проекты"
    headers = [
        "Проект", "Название", "Регион", "Город", "Адрес", "Этап", "Стадия", "В стадии с",
        "Собственник", "Ответственный", "Следующий шаг", "Срок",
        "Форма контроля", "Договор с", "Договор по",
        "Мощность, кВт", "Машино-мест", "Субсидия", "Введён",
        "ТП статус", "Сетевая", "ТП срок", "ТП факт", "ТП стоимость",
        "Бюджет план", "Бюджет факт", "Документов",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows:
        ws.append([
            r["project_no"], r["title"], r["region"], r["city"], r["address"],
            PHASE_LABELS.get(STAGE_PHASE.get(r["stage"], ""), ""),
            STAGE_LABELS.get(r["stage"], r["stage"]), r["stage_since"],
            r["owner"], r["owner_user"], r["next_action"], r["next_action_due"],
            r["control_form"], r["contract_start"], r["contract_end"],
            float(r["planned_power_kwt"] or 0) or None, r["parking_spots"],
            "да" if r["subsidy_planned"] else "", r["commissioned_on"],
            TC_LABELS.get(r["tc_status"] or "", ""), r["grid_operator"],
            r["tc_due"], r["tc_done"], float(r["tc_cost"] or 0) or None,
            float(r["plan_amount"] or 0) or None, float(r["fact_amount"] or 0) or None,
            int(r["docs"] or 0),
        ])
    widths = [16, 24, 22, 18, 40, 14, 16, 12, 26, 20, 30, 12, 18, 12, 12, 14, 12, 10, 12,
              18, 22, 12, 12, 14, 14, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    # Лист бюджета: одна строка — статья затрат проекта. Раньше выгрузка отдавала
    # только итог план/факт на проект, и разобраться, ЧТО именно подорожало, было
    # нельзя, а капвложения не отделялись от расходов периода.
    costs = (await db.execute(text("""
        select s.project_no, coalesce(s.address, s.full_address) as address, s.stage,
               c.kind, c.title, c.plan_amount, c.fact_amount, c.doc_ref
        from ezs_site_costs c join ezs_sites s on s.id = c.site_id
        where c.company_id = :cid
        order by s.project_no, c.created_at
    """), {"cid": company_id})).mappings().all()
    wsc = wb.create_sheet("Бюджет")
    cost_headers = ["Проект", "Адрес", "Стадия", "Статья", "Судьба затрат", "Описание",
                    "План", "Факт", "Отклонение", "Основание факта"]
    wsc.append(cost_headers)
    for c in wsc[1]:
        c.font = Font(bold=True)
    for r in costs:
        plan = float(r["plan_amount"] or 0)
        fact = float(r["fact_amount"] or 0)
        wsc.append([
            r["project_no"], r["address"], STAGE_LABELS.get(r["stage"], r["stage"]),
            COST_LABELS.get(r["kind"], r["kind"]),
            "капвложение" if COST_CAPITAL.get(r["kind"]) else "расход периода",
            r["title"], plan or None, fact or None,
            round(fact - plan, 2) if (plan or fact) else None, r["doc_ref"],
        ])
    for i, w in enumerate([16, 40, 16, 24, 18, 30, 14, 14, 14, 24], start=1):
        wsc.column_dimensions[wsc.cell(row=1, column=i).column_letter].width = w
    wsc.freeze_panes = "A2"

    # Лист присоединений: срок проекта определяет именно оно, а в выгрузке
    # портфеля от него помещались четыре колонки.
    tcs = (await db.execute(text("""
        select s.project_no, coalesce(s.address, s.full_address) as address, s.stage,
               t.status, t.grid_operator, t.application_no, t.application_date,
               t.specs_no, t.specs_date, t.contract_no, t.contract_date,
               t.power_kwt, t.voltage, t.cost, t.works_cost, t.total_cost,
               t.due_date, t.done_date, t.needs_reconstruction, t.applicant_term_months,
               t.substation_owner, t.transformer_kva
        from ezs_tech_connections t join ezs_sites s on s.id = t.site_id
        where t.company_id = :cid
        order by s.project_no
    """), {"cid": company_id})).mappings().all()
    wst = wb.create_sheet("Присоединение")
    tc_headers = ["Проект", "Адрес", "Стадия", "Статус ТП", "Сетевая организация",
                  "Заявка №", "Заявка от", "ТУ №", "ТУ от", "Договор ТП №", "Договор от",
                  "Мощность, кВт", "Напряжение", "Договор с сетевой, ₽",
                  "Мероприятия Общества, ₽", "Итого ТУ, ₽", "Срок мероприятий",
                  "Факт исполнения", "Реконструкция ТП", "Срок заявителя, мес.",
                  "Принадлежность ПС", "Трансформатор"]
    wst.append(tc_headers)
    for c in wst[1]:
        c.font = Font(bold=True)
    for r in tcs:
        wst.append([
            r["project_no"], r["address"], STAGE_LABELS.get(r["stage"], r["stage"]),
            TC_LABELS.get(r["status"], r["status"]), r["grid_operator"],
            r["application_no"], r["application_date"], r["specs_no"], r["specs_date"],
            r["contract_no"], r["contract_date"],
            float(r["power_kwt"] or 0) or None, r["voltage"],
            float(r["cost"] or 0) or None, float(r["works_cost"] or 0) or None,
            float(r["total_cost"] or 0) or None,
            r["due_date"], r["done_date"],
            "да" if r["needs_reconstruction"] else ("нет" if r["needs_reconstruction"] is False else ""),
            r["applicant_term_months"], r["substation_owner"], r["transformer_kva"],
        ])
    for i, w in enumerate([16, 36, 16, 20, 26, 14, 12, 14, 12, 14, 12, 12, 12,
                           18, 20, 16, 14, 14, 16, 16, 20, 16], start=1):
        wst.column_dimensions[wst.cell(row=1, column=i).column_letter].width = w
    wst.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def awaiting_accounting(db: AsyncSession, company_id) -> dict[str, Any]:
    """«Ждёт учёта» — где проект ушёл вперёд, а бухгалтерия про это не знает."""
    S = EzsSite

    no_contract = (await db.execute(
        select(S).where(S.company_id == company_id, S.contract_id.is_(None),
                        S.contract_start.is_not(None)))).scalars().all()
    no_location = (await db.execute(
        select(S).where(S.company_id == company_id, S.location_id.is_(None),
                        S.stage.in_(["commissioning", "live"])))).scalars().all()
    no_supply = (await db.execute(text("""
        select s.id, s.project_no, s.city, s.address
        from ezs_sites s
        where s.company_id = :cid and s.stage in ('construction','commissioning','live')
          -- Поставка ищется у ЭТОГО проекта, а не у компании: подзапрос без связи
          -- по site_id отвечал «поставки есть» на весь список, стоило появиться
          -- одному документу в компании, и раздел молча пустел.
          and not exists (
              select 1 from ezs_site_equipment e
              where e.site_id = s.id and e.status in ('supplied', 'installed'))
        limit 200
    """), {"cid": company_id})).mappings().all()

    def brief(s: EzsSite) -> dict[str, Any]:
        return {"id": str(s.id), "projectNo": s.project_no, "title": s.title,
                "region": s.region_norm or s.region, "city": s.city,
                "address": s.address or s.full_address, "stage": s.stage,
                "stageLabel": STAGE_LABELS.get(s.stage, s.stage)}

    return {
        "contractMissing": [brief(s) for s in no_contract],
        "locationMissing": [brief(s) for s in no_location],
        "supplyMissing": [{"id": str(r["id"]), "projectNo": r["project_no"],
                           "city": r["city"], "address": r["address"]} for r in no_supply],
    }


async def link_contract(db: AsyncSession, company_id, site: EzsSite, contract_id,
                        user: User | None) -> dict[str, Any]:
    """Привязать существующий договор учёта к проекту (создавать — руками в учёте)."""
    c = (await db.execute(select(Contract).where(
        Contract.company_id == company_id,
        Contract.id == _uuid.UUID(str(contract_id))))).scalar_one_or_none()
    if c is None:
        return {"ok": False, "message": "Договор не найден"}
    old_contract_id = site.contract_id
    old_contract = None
    if old_contract_id and old_contract_id != c.id:
        old_contract = (await db.execute(select(Contract).where(
            Contract.company_id == company_id,
            Contract.id == old_contract_id))).scalar_one_or_none()
    site.contract_id = c.id
    await log_event(db, site, "edit", user=user,
                    text=f"Привязан договор № {c.number} от {c.date}"
                         + (f" ({c.basis})" if c.basis else ""),
                    changes=[make_change(
                        "contract_id", old_contract_id, c.id,
                        old_display=(f"№ {old_contract.number} от {old_contract.date}"
                                     if old_contract else None),
                        new_display=f"№ {c.number} от {c.date}",
                    )] if old_contract_id != c.id else None)
    from app.services.ezs_lifecycle import sync_from_site
    await sync_from_site(db, company_id, site)
    return {"ok": True, "contract": {"id": str(c.id), "number": c.number, "date": c.date,
                                     "basis": c.basis, "validUntil": c.valid_until}}


async def link_location(db: AsyncSession, company_id, site: EzsSite, location_id,
                        user: User | None) -> dict[str, Any]:
    """Связать проект с объектом сети — цикл замкнут, проект стал станцией."""
    loc = (await db.execute(select(ServiceLocation).where(
        ServiceLocation.company_id == company_id,
        ServiceLocation.id == str(location_id)))).scalar_one_or_none()
    if loc is None:
        return {"ok": False, "message": "Объект не найден"}
    old_location_id = site.location_id
    old_location = None
    if old_location_id and old_location_id != loc.id:
        old_location = (await db.execute(select(ServiceLocation).where(
            ServiceLocation.company_id == company_id,
            ServiceLocation.id == old_location_id))).scalar_one_or_none()
    site.location_id = loc.id
    # Дату ввода здесь НЕ ставим. Привязка объекта — это регистрация станции в
    # реестре сети; её делают и до ввода, и на проекте, который потом отклонят.
    # Дата ввода — основание перевода капвложений со счёта 08 на 01, и ставит её
    # ровно один путь: маршрут, дошедший до стадии ввода, через закрытый чек-лист
    # (projects_process._reflect_commissioning).
    await log_event(
        db, site, "edit", user=user, text=f"Связан объект сети: {loc.name}",
        changes=[make_change(
            "location_id", old_location_id, loc.id,
            old_display=old_location.name if old_location else None,
            new_display=loc.name,
        )] if old_location_id != loc.id else None,
    )
    from app.services.ezs_lifecycle import sync_from_site
    await sync_from_site(db, company_id, site)
    return {"ok": True, "location": {"id": str(loc.id), "name": loc.name, "code": loc.code}}


async def project_roadmap(db: AsyncSession, company_id, site: EzsSite) -> dict[str, Any]:
    """Схема реализации проекта — дорожная карта от участка до эксплуатации.

    Собирает в одну ленту то, что разложено по вкладкам: стадии воронки с
    гейтами, техприсоединение, оборудование, документы, договор и объект сети.
    Каждый шаг знает своё состояние (сделано / в работе / ждёт / просрочено),
    дату и что именно его держит.

    Смысл экрана — видеть путь целиком: где мы, что осталось и что мешает.
    Табличные вкладки отвечают «как заполнить», а схема — «где мы в проекте».
    """
    from app.services.ezs_site_work import GATES, gate_state, site_doc_kinds

    doc_kinds = await site_doc_kinds(db, site.id)
    tc = await get_tech_connection(db, company_id, site.id)
    eq = await list_equipment(db, company_id, site.id)
    docs = await list_docs(db, company_id, site.id)
    today = date.today().isoformat()
    off_path = site.stage in ("archive", "on_hold")
    # У снятого с пути проекта (архив / заморозка) «текущая» стадия — та, на
    # которой его сняли (prev_stage), а не сам архив. Иначе схема красит весь путь
    # крестами и показывает 0%, хотя проект реально прошёл часть пути.
    effective = site.prev_stage if off_path and site.prev_stage in STAGE_ORDER else site.stage
    cur_pos = STAGE_ORDER.index(effective) if effective in STAGE_ORDER else -1
    # prev_stage мы знаем только для проектов, которые вели в системе; у
    # импортированных из архива истории нет — там честно «неизвестно».
    unknown_progress = off_path and site.prev_stage not in STAGE_ORDER

    def stage_state(stage: str) -> str:
        pos = STAGE_ORDER.index(stage)
        if unknown_progress:
            return "unknown"           # импортированный архив: докуда дошёл — не знаем
        if cur_pos < 0:
            return "waiting"
        if pos < cur_pos:
            return "done"
        if pos == cur_pos:
            return "stopped" if off_path else "current"
        return "waiting"

    steps: list[dict[str, Any]] = []
    for stage in STAGE_ORDER:
        st = stage_state(stage)
        g = gate_state(site, stage, doc_kinds=doc_kinds,
                       equipment_supplied=eq["allSupplied"])
        # Что именно держит текущий шаг — самое ценное на схеме.
        blocking = g["blocking"] if st == "current" else []
        steps.append({
            "key": stage, "kind": "stage", "label": STAGE_LABELS[stage],
            "phase": STAGE_PHASE.get(stage), "phaseLabel": PHASE_LABELS.get(STAGE_PHASE.get(stage, ""), ""),
            "state": st,
            "date": site.stage_since if st == "current" else None,
            "gateDone": g["done"], "gateTotal": g["total"], "blocking": blocking,
            "items": [{"label": i["label"], "done": i["done"], "required": i["required"]}
                      for i in g["items"]],
        })

    # Параллельные треки — привязаны к своему этапу, чтобы схема группировалась
    # по этапам проекта: земля → трек «право», реализация → ТП и оборудование,
    # эксплуатация → объект в сети.
    tracks: list[dict[str, Any]] = []
    if tc:
        tc_state = ("done" if tc["status"] == "done" else
                    "failed" if tc["status"] == "rejected" else
                    "overdue" if tc["overdue"] else
                    "current" if tc["status"] != "draft" else "waiting")
        tracks.append({
            "key": "tp", "kind": "track", "phase": "build", "label": "Техприсоединение",
            "state": tc_state, "status": tc["statusLabel"],
            "date": tc["doneDate"] or tc["dueDate"],
            "detail": " · ".join(x for x in [
                tc["gridOperator"],
                f"ТУ {tc['specsNo']}" if tc["specsNo"] else None,
                f"{tc['powerKwt']:.0f} кВт" if tc["powerKwt"] else None,
            ] if x) or None,
            "note": "срок мероприятий прошёл" if tc["overdue"] else None,
        })
    else:
        tracks.append({"key": "tp", "kind": "track", "phase": "build", "label": "Техприсоединение",
                       "state": "empty", "status": "не заведено", "date": None,
                       "detail": None, "note": "заявка в сетевую не подана"})

    if eq["items"]:
        eq_state = ("done" if eq["allInstalled"] else
                    "overdue" if any(i["overdue"] for i in eq["items"]) else
                    "current" if eq["allSupplied"] else "waiting")
        tracks.append({
            "key": "equipment", "kind": "track", "phase": "build", "label": "Оборудование",
            "state": eq_state,
            "status": f"{sum(1 for i in eq['items'] if i['status'] in ('supplied', 'installed'))}"
                      f" из {len(eq['items'])} поставлено",
            "date": None,
            "detail": ", ".join(i["title"] for i in eq["items"][:2] if i["title"]) or None,
            "note": "просрочена поставка" if any(i["overdue"] for i in eq["items"]) else None,
        })
    else:
        tracks.append({"key": "equipment", "kind": "track", "phase": "build", "label": "Оборудование",
                       "state": "empty", "status": "потребность не заведена",
                       "date": None, "detail": None, "note": None})

    contract = None
    if site.contract_id:
        c = (await db.execute(select(Contract).where(Contract.id == site.contract_id))).scalar_one_or_none()
        if c:
            contract = f"№ {c.number} от {c.date}"
    tracks.append({
        "key": "land", "kind": "track", "phase": "land", "label": "Право на землю",
        "state": ("done" if site.contract_start and contract else
                  "current" if site.contract_start else "empty"),
        "status": (site.control_form or "форма контроля не определена"),
        "date": site.contract_end,
        "detail": contract or (f"договор с {site.contract_start}" if site.contract_start else None),
        "note": ("подписан, но не заведён в учёте" if site.contract_start and not contract else
                 "срок договора истёк" if site.contract_end and site.contract_end < today else None),
    })

    location = None
    if site.location_id:
        loc = (await db.execute(select(ServiceLocation).where(
            ServiceLocation.id == str(site.location_id)))).scalar_one_or_none()
        location = loc.name if loc else None
    tracks.append({
        "key": "network", "kind": "track", "phase": "operate", "label": "Объект в сети",
        "state": "done" if location else "empty",
        "status": location or "не связан",
        "date": site.commissioned_on, "detail": None,
        "note": None if location else "после ввода проект должен стать станцией сети",
    })

    return {
        "stage": site.stage, "stageLabel": STAGE_LABELS.get(site.stage, site.stage),
        "phase": STAGE_PHASE.get(site.stage),
        "offPath": off_path, "unknownProgress": unknown_progress,
        "stoppedAt": STAGE_LABELS.get(site.prev_stage or "", None) if off_path else None,
        "phases": [{"key": p["key"], "label": p["label"], "hint": p["hint"]} for p in PHASES],
        "steps": steps,
        "tracks": tracks,
        "docs": {"count": len(docs), "kinds": sorted({d["kindLabel"] for d in docs})},
        "subsidy": subsidy_check(site),
        "progress": None if unknown_progress
                    else round(sum(1 for s in steps if s["state"] in ("done", "stopped"))
                               / len(steps) * 100),
    }


async def project_context(db: AsyncSession, company_id, site: EzsSite) -> dict[str, Any]:
    """Всё, что показывает карточка проекта сверх паспорта."""
    contract = None
    if site.contract_id:
        c = (await db.execute(select(Contract).where(Contract.id == site.contract_id))).scalar_one_or_none()
        if c:
            contract = {"id": str(c.id), "number": c.number, "date": c.date,
                        "basis": c.basis, "validUntil": c.valid_until, "type": c.type}
    location = None
    if site.location_id:
        loc = (await db.execute(select(ServiceLocation).where(
            ServiceLocation.id == str(site.location_id)))).scalar_one_or_none()
        if loc:
            location = {"id": str(loc.id), "name": loc.name, "code": loc.code,
                        "status": loc.operational_status}
    return {
        "phase": STAGE_PHASE.get(site.stage, "select"),
        "phases": [{"key": p["key"], "label": p["label"], "hint": p["hint"],
                    "stages": [{"stage": s, "label": STAGE_LABELS[s]} for s in p["stages"]]}
                   for p in PHASES],
        "techConnection": await get_tech_connection(db, company_id, site.id),
        "equipment": await list_equipment(db, company_id, site.id),
        "costs": await list_costs(db, company_id, site.id),
        "subsidy": subsidy_check(site),
        "contract": contract,
        "location": location,
        "docKinds": DOC_KINDS,
        "tcStatuses": TC_STATUSES,
        "eqStatuses": EQ_STATUSES,
        "costKinds": COST_KINDS,
    }


# ── Выгрузки экранов ───────────────────────────────────────────────────────
# Каждая аналитика обязана уметь «дать таблицу»: совещание идёт по Excel, и
# если выгрузки нет, цифры руками переписывают в свой файл — а он тут же
# расходится с системой. Один сборщик на все экраны, чтобы формат был общим.

def build_xlsx(sheets: list[dict[str, Any]]) -> bytes:
    """Книга из листов: `{title, headers, rows, widths?}`.

    Первая строка жирная и закреплена, ширины по месту — иначе выгрузку
    открывают и первым делом растягивают колонки.
    """
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    for i, sh in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = sh["title"][:31]
        ws.append(list(sh["headers"]))
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for r in sh["rows"]:
            ws.append(list(r))
        for idx, w in enumerate(sh.get("widths") or [], start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w
        ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def export_funnel_xlsx(db: AsyncSession, company_id) -> bytes:
    """Воронка: сколько стоит на стадии, сколько идёт дольше нормы, по регионам."""
    from app.services.ezs_sites import sites_overview

    ov = await sites_overview(db, company_id)
    pf = await portfolio_overview(db, company_id)
    by_stage = {f["stage"]: f for f in pf.get("funnel", [])}
    rows = []
    for f in ov["funnel"]:
        p = by_stage.get(f["stage"], {})
        rows.append([
            f["label"], f["hint"], f["count"],
            p.get("medianDays"), p.get("normDays"),
            p.get("stuck"), p.get("passRate"),
        ])
    regions = [[r["region"], r["count"]] for r in ov["byRegion"]]
    return build_xlsx([
        {"title": "Воронка",
         "headers": ["Стадия", "Смысл", "Сейчас", "Медиана, дн", "Норматив, дн",
                     "Дольше нормы", "Проходят дальше"],
         "rows": rows, "widths": [22, 40, 10, 14, 14, 14, 16]},
        {"title": "Регионы", "headers": ["Регион", "Проектов"],
         "rows": regions, "widths": [34, 12]},
    ])


async def export_matrix_xlsx(db: AsyncSession, company_id) -> bytes:
    """Приоритеты: обе оси, уверенность, решение и чего не хватает."""
    from app.services.ezs_site_analysis import priority_matrix

    m = await priority_matrix(db, company_id)
    # Подписи квадрантов берём из самого ответа: там они уже посчитаны вместе со
    # счётчиками, и второй источник рано или поздно с ним разойдётся.
    labels = {q["key"]: q["label"] for q in m.get("quadrants", [])}
    rows = [[
        i["projectNo"], i["title"], i["region"], i["city"], i["address"],
        i["stageLabel"], i["attract"], i["feasible"], i["confidence"],
        i["nearestStationKm"], "да" if i["cannibalization"] else "",
        labels.get(i["quadrant"], i["quadrant"]), "; ".join(i["unknown"]),
    ] for i in m["items"]]
    return build_xlsx([{
        "title": "Приоритеты",
        "headers": ["Проект", "Название", "Регион", "Город", "Адрес", "Стадия",
                    "Привлекательность", "Исполнимость", "Уверенность, %",
                    "До сети, км", "Каннибализация", "Решение", "Чего не хватает"],
        "rows": rows,
        "widths": [16, 24, 22, 18, 40, 16, 18, 16, 15, 13, 15, 22, 46],
    }])


async def export_budget_xlsx(db: AsyncSession, company_id) -> bytes:
    """Бюджет портфеля: статьи с судьбой денег и корзины по состоянию проектов."""
    rep = await costs_report(db, company_id)
    rows = [[
        i["label"], "капвложение" if i.get("capital") else "расход периода",
        i.get("sites"), i.get("plan"), i.get("fact"),
        (i.get("fact") or 0) - (i.get("plan") or 0),
    ] for i in rep["items"]]
    buckets = [[b["label"], b["plan"], b["fact"]] for b in rep["buckets"]]
    return build_xlsx([
        {"title": "По статьям",
         "headers": ["Статья", "Судьба", "Проектов", "План, ₽", "Факт, ₽", "Отклонение, ₽"],
         "rows": rows, "widths": [34, 18, 12, 16, 16, 16]},
        {"title": "Состояние", "headers": ["Состояние проектов", "План, ₽", "Факт, ₽"],
         "rows": buckets, "widths": [34, 16, 16]},
    ])


async def export_accounting_xlsx(db: AsyncSession, company_id) -> bytes:
    """Ждёт учёта: где проект ушёл вперёд, а бухгалтерия об этом не знает."""
    rep = await awaiting_accounting(db, company_id)
    # Расхождения приходят тремя списками по виду — в таблице это одна колонка,
    # иначе на листе три пустых блока вместо одного читаемого перечня.
    groups = {
        "contractMissing": "Нет договора на землю",
        "locationMissing": "Нет объекта сети",
        "supplyMissing": "Оборудование не поставлено",
    }
    rows = []
    for key, label in groups.items():
        for i in rep.get(key) or []:
            rows.append([
                i.get("projectNo"), i.get("title") or i.get("address"),
                i.get("stageLabel"), label, i.get("hint") or "",
            ])
    return build_xlsx([{
        "title": "Ждёт учёта",
        "headers": ["Проект", "Название", "Стадия", "Расхождение", "Подробности"],
        "rows": rows, "widths": [16, 30, 18, 34, 46],
    }])


async def export_tech_connections_xlsx(db: AsyncSession, company_id) -> bytes:
    """Присоединения: заявки, ТУ, договоры, сроки и просрочки по сетевым."""
    rep = await tech_connections_report(db, company_id)
    rows = [[
        i["projectNo"], i["title"], i["region"], i["city"],
        i["statusLabel"], i["gridOperator"], i["applicationNo"], i["applicationDate"],
        i["specsNo"], i["specsDate"], i["contractNo"], i["contractDate"],
        i["powerKwt"], i["cost"], i["dueDate"], i["doneDate"],
        "да" if i["overdue"] else "",
    ] for i in rep["items"]]
    return build_xlsx([{
        "title": "Присоединение",
        "headers": ["Проект", "Название", "Регион", "Город", "Статус", "Сетевая",
                    "№ заявки", "Дата заявки", "№ ТУ", "Дата ТУ", "№ договора",
                    "Дата договора", "Мощность, кВт", "Стоимость, ₽",
                    "Срок мероприятий", "Исполнено", "Просрочено"],
        "rows": rows,
        "widths": [16, 26, 22, 18, 18, 26, 18, 13, 18, 13, 18, 13, 14, 14, 16, 13, 12],
    }])


async def export_equipment_xlsx(db: AsyncSession, company_id) -> bytes:
    """Оборудование проектов: что заказано, что пришло, что смонтировано."""
    rep = await equipment_report(db, company_id)
    rows = [[
        i["projectNo"], i["projectTitle"], i["city"], i["title"], i["manufacturer"],
        i["powerKwt"], i["qty"], i["supplier"], i["price"],
        i["statusLabel"], i["dueDate"], i["suppliedDate"], "да" if i["overdue"] else "",
    ] for i in rep["items"]]
    return build_xlsx([{
        "title": "Оборудование",
        "headers": ["Проект", "Название проекта", "Город", "Оборудование", "Производитель",
                    "Мощность, кВт", "Кол-во", "Поставщик", "Стоимость, ₽",
                    "Статус", "Плановая поставка", "Поставлено", "Просрочено"],
        "rows": rows,
        "widths": [16, 26, 18, 30, 20, 14, 10, 24, 14, 16, 16, 13, 12],
    }])
