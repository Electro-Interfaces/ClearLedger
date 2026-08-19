"""
Нормализация НОВОГО входного потока реестров РусГидро (июль 2026) — три xlsx,
которые контрагент ведёт вручную и присылает файлами. Канал тот же —
`reestr_contracts_payments` (manual_table): формат загруженного файла
определяется автоматически (detect_reestr_format), пользователь грузит файлы
по очереди в один канал.

Форматы:
  svodnaya — «Инф-я по Дог. аренды_электроэнергии», лист «Сводная» (725×146):
             паспорт станции + договоры э/э (контрагент/тариф/оплачено) +
             аренда (контрагент/суммы) + сервисный договор + ПОМЕСЯЧНЫЕ объёмы
             входящей э/э (кВт·ч, колонки «<месяц> <год> электроэнергия»);
  arenda   — «ЭЗС_Договоры_Аренда», лист «Договоры_Аренда» (498×29) + лист
             «Контакты»: актуальная постоянная часть аренды, ИНН контрагентов,
             переменная часть, сроки, реквизиты договора;
  tariffs  — «Тарифы Электроэнергия_Входящие», «Лист1» (437×29+): входящие
             тарифы э/э руб/кВт·ч с НДС помесячно (колонки прирастают);
  svod     — старый «Общий свод» → reestr_normalize.ingest_reestr (совместимость).

Сопряжение с уже загруженными объектами (620 service_locations, тип ev_charging):
  «№ по БУ» → extra_metadata.number (главный ключ, ~440 прямых) → code;
  зав. номер / OCPP ID → serialNumber → code (~350);
  координаты (±150 м) — добор для «Сводной»;
  «№ ZOI-1» в БД изначально НЕТ — сквозной ключ МЕЖДУ файлами; при первом
  успешном резолве пишется в extra_metadata.zoi1 (и buNumber), чтобы следующие
  файлы резолвились и по нему.

L1 RAW — DataEntry(layer='raw') пер-строку (идемпотентно по source_label);
L2 — Counterparty (+ИНН/контакты), Contract, StationContractSettlement
     (role=energy|rent|service, суммы/сроки), StationEnergyPeriod (объёмы+тарифы).
"""
from __future__ import annotations

import io
import re
from typing import Any

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.station_passport import (
    passport_value, stations_by_location, write_value)
from app.models import (
    Contract,
    ContractLocation,
    Counterparty,
    DataEntry,
    Organization,
    ServiceLocation,
    StationContractSettlement,
    StationDispensePeriod,
    StationEnergyPeriod,
)
from app.services.reestr_normalize import (
    SHEET as OLD_SHEET,
    _blank,
    _clean_cp_name,
    _cp_type,
    _detect_basis,
    _normname,
    _parse_num_date,
)

TAG_SVODNAYA = "reestr-rh-svodnaya"
TAG_ARENDA = "reestr-rh-arenda"
TAG_TARIFFS = "reestr-rh-tariffs"
TAG_OBSHAYA = "reestr-rh-obshaya"

RU_MONTHS = {
    "январ": 1, "феврал": 2, "март": 3, "апрел": 4, "ма": 5, "июн": 6,
    "июл": 7, "август": 8, "сентябр": 9, "октябр": 10, "ноябр": 11, "декабр": 12,
}
_MONTH_RE = re.compile(
    r"(январ|феврал|март|апрел|ма[йя]|июн|июл|август|сентябр|октябр|ноябр|декабр)\w*\s*[-,./ ]*\s*(\d{2,4})",
    re.I,
)


# --------------------------------------------------------------------------- helpers
def _s(v) -> str:
    """Значение → строка без хвостов ('123.0' → '123')."""
    if v is None:
        return ""
    s = str(v).strip()
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def _num(v) -> float | None:
    """Число из ячейки (умеет '1 234,5'); мусор → None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _iso_date(v) -> str | None:
    """Ячейка (datetime | 'дд.мм.гггг' | ISO) → 'YYYY-MM-DD'."""
    if _blank(v):
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    s = str(v).strip()
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return m.group(0)
    m = re.search(r"(\d{1,2})[.](\d{1,2})[.](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        return f"{y:04d}-{mo:02d}-{d:02d}"
    return None


def _month_from_header(text: str) -> str | None:
    """Шапка вида 'сентябрь 22 электроэнергия' / 'Тариф…, июнь 2026 …' → '2022-09-01'."""
    m = _MONTH_RE.search(str(text))
    if not m:
        return None
    stem = m.group(1).lower()
    mo = 5 if stem in ("май", "мая") else RU_MONTHS.get(stem)
    if mo is None:
        for k, n in RU_MONTHS.items():
            if stem.startswith(k):
                mo = n
                break
    if mo is None:
        return None
    y = int(m.group(2))
    if y < 100:
        y += 2000
    if not (2015 <= y <= 2040):
        return None
    return f"{y:04d}-{mo:02d}-01"


_RU_MONTH_ABBR = {
    "янв": 1, "фев": 2, "мар": 3, "апр": 4, "май": 5, "мая": 5, "июн": 6,
    "июл": 7, "авг": 8, "сен": 9, "окт": 10, "ноя": 11, "дек": 12,
}
_MONTH_SHORT_RE = re.compile(r"(янв|фев|мар|апр|ма[йя]|июн|июл|авг|сен|окт|ноя|дек)\w*\s*[.\-,/ ]*\s*(\d{2,4})", re.I)


def _month_short(v) -> str | None:
    """Шапка месяца сводной выработки → '2024-03-01'.
    Умеет datetime (лист «Сумма»), 'мар.24', 'мар.2024', 'июнь.25', 'июл.26'."""
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        s = v.isoformat()[:10]
        return s[:8] + "01" if s[:4].isdigit() else None
    m = _MONTH_SHORT_RE.search(str(v))
    if not m:
        return None
    mo = _RU_MONTH_ABBR.get(m.group(1).lower()[:3])
    if mo is None:
        return None
    y = int(m.group(2))
    if y < 100:
        y += 2000
    if not (2015 <= y <= 2040):
        return None
    return f"{y:04d}-{mo:02d}-01"


def _ru_month_paid(v) -> tuple[str, str | None, str | None]:
    """«Последний оплаченный месяц» → (status, paid_through_ISO, note).
    Умеет datetime, 'февраль' (год неизвестен → note), 'не оплачено'."""
    if _blank(v):
        return "unknown", None, None
    if hasattr(v, "isoformat"):
        return "paid", v.isoformat()[:7] + "-01", None
    s = str(v).strip()
    low = s.lower()
    if "не оплач" in low or "не произв" in low or "не поступ" in low:
        return "unpaid", None, s
    iso = _month_from_header(low)
    if iso:
        return "paid", iso, None
    # голый русский месяц без года — оплату фиксируем, месяц в примечание
    for stem in RU_MONTHS:
        if low.startswith(stem):
            return "paid", None, s
    return "special", None, s


def _parse_contract_req(v) -> tuple[str | None, str | None]:
    """Реквизиты договора: '№ N от дд.мм.гггг' ЛИБО 'дд.мм.гггг № N' → (номер, дата)."""
    if _blank(v):
        return None, None
    s = str(v).strip()
    m = re.match(r"(\d{1,2}[.]\d{1,2}[.]\d{2,4})\s*(.+)", s)
    if m:
        return m.group(2).strip().lstrip("№").strip()[:100], _iso_date(m.group(1))
    number, date = _parse_num_date(s)
    return (number.lstrip("№").strip()[:100] if number else None), date


def _jsonable_row(d: dict) -> dict:
    out = {}
    for k, v in d.items():
        out[k] = v.isoformat() if hasattr(v, "isoformat") else v
    return out


def _compact(d: dict) -> dict:
    """Убрать пустые значения из extra/meta."""
    return {k: v for k, v in d.items() if v not in (None, "", [], {})}


# --------------------------------------------------------------------------- detect
def detect_reestr_format(content: bytes) -> str:
    """Формат xlsx по листам/сигнатуре шапки: svodnaya | arenda | tariffs | svod."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        names = set(wb.sheetnames)
        if "Сводная" in names:
            return "svodnaya"
        if "Договоры_Аренда" in names:
            return "arenda"
        if "Коннектор_общая" in names or "Общие итоги" in names:
            return "obshaya"
        if OLD_SHEET in names:
            return "svod"
        # «Тарифы…»: лист может называться «Лист1» — ищем сигнатуру шапки
        for ws in wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            head = next(ws.iter_rows(values_only=True, max_row=1), ()) or ()
            joined = " | ".join(str(c) for c in head if c is not None)
            if "тариф электроэнерг" in joined.lower():
                return "tariffs"
        raise ValueError(
            f"Не удалось определить формат реестра; листы: {sorted(names)}"
        )
    finally:
        wb.close()


# --------------------------------------------------------------------------- resolver
class StationResolver:
    """Сопряжение строк реестров с существующими service_locations компании.

    Каскад: № по БУ → metadata.number → code; зав.№/OCPP → serialNumber → code;
    zoi1 (запомненный ранее) ; координаты (±150 м). Успешный резолв дописывает
    buNumber/zoi1 в extra_metadata (новым dict — иначе JSONB-изменение не увидится).
    """

    def __init__(self) -> None:
        self.by_number: dict[str, list[ServiceLocation]] = {}
        self.by_code: dict[str, list[ServiceLocation]] = {}
        self.by_serial: dict[str, list[ServiceLocation]] = {}
        self.by_zoi: dict[str, list[ServiceLocation]] = {}
        self.coords: list[tuple[float, float, ServiceLocation]] = []
        self.stats = {"bu": 0, "serial": 0, "zoi": 0, "coords": 0, "miss": 0, "ambiguous": 0}

    async def build(self, db: AsyncSession, company_id) -> None:
        locs = (await db.execute(
            select(ServiceLocation).where(ServiceLocation.company_id == company_id)
        )).scalars().all()
        for l in locs:
            md = l.extra_metadata or {}

            def put(index: dict, val) -> None:
                v = _s(val).lower()
                if v:
                    index.setdefault(v, []).append(l)

            put(self.by_number, md.get("number"))
            put(self.by_number, md.get("buNumber"))
            put(self.by_code, l.code)
            put(self.by_serial, md.get("serialNumber"))
            put(self.by_serial, l.serial_number)
            put(self.by_zoi, md.get("zoi1"))
            lat = _num(md.get("latitude") if md.get("latitude") is not None else l.latitude)
            lon = _num(md.get("longitude") if md.get("longitude") is not None else l.longitude)
            if lat and lon:
                self.coords.append((lat, lon, l))

    def _pick(self, index: dict, key: str) -> ServiceLocation | None:
        cand = index.get(_s(key).lower()) if key else None
        if not cand:
            return None
        if len(set(c.id for c in cand)) > 1:
            self.stats["ambiguous"] += 1
            return None
        return cand[0]

    def resolve(self, *, bu=None, zoi=None, serial=None, ocpp=None,
                lat=None, lon=None) -> ServiceLocation | None:
        loc = self._pick(self.by_zoi, zoi) if zoi else None
        if loc is not None:
            self.stats["zoi"] += 1
            return self._remember(loc, bu, zoi)
        for key in (bu,):
            if not _s(key):
                continue
            loc = self._pick(self.by_number, key) or self._pick(self.by_code, key)
            if loc is not None:
                self.stats["bu"] += 1
                return self._remember(loc, bu, zoi)
        for key in (serial, ocpp):
            if not _s(key):
                continue
            loc = self._pick(self.by_serial, key) or self._pick(self.by_code, key)
            if loc is not None:
                self.stats["serial"] += 1
                return self._remember(loc, bu, zoi)
        la, lo = _num(lat), _num(lon)
        if la and lo:
            near = [l for (a, b, l) in self.coords
                    if abs(a - la) < 0.0015 and abs(b - lo) < 0.0015]
            if len(set(l.id for l in near)) == 1:
                self.stats["coords"] += 1
                return self._remember(near[0], bu, zoi)
        self.stats["miss"] += 1
        return None

    def _remember(self, loc: ServiceLocation, bu, zoi) -> ServiceLocation:
        """Дописать buNumber/zoi1 в метаданные объекта + горячие индексы."""
        md = dict(loc.extra_metadata or {})
        changed = False
        if _s(bu) and not _s(md.get("buNumber")):
            md["buNumber"] = _s(bu)
            self.by_number.setdefault(_s(bu).lower(), []).append(loc)
            changed = True
        if _s(zoi) and not _s(md.get("zoi1")):
            md["zoi1"] = _s(zoi)
            self.by_zoi.setdefault(_s(zoi).lower(), []).append(loc)
            changed = True
        if changed:
            loc.extra_metadata = md
        return loc


# --------------------------------------------------------------------------- L1 RAW
async def _upsert_raw(db: AsyncSession, company_id, tag: str, key: str,
                      title: str, meta: dict, existing: dict[str, DataEntry]) -> int:
    """DataEntry(layer='raw') идемпотентно по source_label; префетч в existing."""
    label = f"{tag}-{key}"
    e = existing.get(label)
    if e is not None:
        e.meta = meta
        e.status = "recognized"
        return 0
    e = DataEntry(
        company_id=company_id, title=title,
        category_id="energy", subcategory_id="contracts_payments",
        doc_type_id="reestr_row", status="recognized",
        source="upload", source_label=label, source_id=label,
        layer="raw", meta=meta,
    )
    db.add(e)
    existing[label] = e
    return 1


async def _prefetch_raw(db: AsyncSession, company_id, tag: str) -> dict[str, DataEntry]:
    rows = (await db.execute(
        select(DataEntry).where(
            DataEntry.company_id == company_id,
            DataEntry.layer == "raw",
            DataEntry.source_label.like(f"{tag}-%"),
        )
    )).scalars().all()
    return {e.source_label: e for e in rows}


# --------------------------------------------------------------------------- L2 helpers
class L2Cache:
    """Префетч контрагентов/договоров/расчётов/периодов компании для upsert без N+1."""

    def __init__(self) -> None:
        self.cp_by_norm: dict[str, Counterparty] = {}
        self.contr_by_key: dict[tuple, Contract] = {}
        self.contr_by_cp_type: dict[tuple, list[Contract]] = {}
        self.settl: dict[tuple, StationContractSettlement] = {}
        self.periods: dict[tuple, StationEnergyPeriod] = {}
        self.dispense: dict[tuple, StationDispensePeriod] = {}
        self.contract_locs: set[tuple] = set()
        self.org_id: str = ""

    async def build(self, db: AsyncSession, company_id) -> None:
        for c in (await db.execute(
            select(Counterparty).where(Counterparty.company_id == company_id)
        )).scalars().all():
            nn = _normname(c.name)
            if nn and nn not in self.cp_by_norm:
                self.cp_by_norm[nn] = c
        contracts = (await db.execute(
            select(Contract).where(Contract.company_id == company_id)
        )).scalars().all()
        for c in contracts:
            self.contr_by_key[(c.counterparty_id, str(c.number), c.type)] = c
            self.contr_by_cp_type.setdefault((c.counterparty_id, c.type), []).append(c)
        for s in (await db.execute(
            select(StationContractSettlement).where(
                StationContractSettlement.company_id == company_id)
        )).scalars().all():
            self.settl[(s.location_id, s.role)] = s
        for p in (await db.execute(
            select(StationEnergyPeriod).where(StationEnergyPeriod.company_id == company_id)
        )).scalars().all():
            self.periods[(p.location_id, p.period)] = p
        for d in (await db.execute(
            select(StationDispensePeriod).where(StationDispensePeriod.company_id == company_id)
        )).scalars().all():
            self.dispense[(d.location_id, d.period, d.connector_type or "")] = d
        for cl in (await db.execute(
            select(ContractLocation).where(ContractLocation.company_id == company_id)
        )).scalars().all():
            self.contract_locs.add((str(cl.contract_id), cl.location_id))
        if contracts:
            from collections import Counter
            top = Counter(c.organization_id for c in contracts if c.organization_id).most_common(1)
            self.org_id = top[0][0] if top else ""
        if not self.org_id:
            org = (await db.execute(
                select(Organization).where(Organization.company_id == company_id)
            )).scalars().first()
            self.org_id = str(org.id) if org else ""

    async def counterparty(self, db: AsyncSession, company_id, raw_name,
                           inn: str | None = None, alias: str | None = None,
                           res: dict | None = None) -> Counterparty | None:
        """Контрагент по имени (норм-форма): существующий (+дозаполнить ИНН) или новый."""
        if _blank(raw_name) or "информац" in str(raw_name).lower():
            return None
        nn = _normname(raw_name)
        if not nn:
            return None
        cp = self.cp_by_norm.get(nn)
        if cp is None:
            cp = Counterparty(
                company_id=company_id, inn=_s(inn)[:20], name=_clean_cp_name(raw_name)[:500],
                type=_cp_type(raw_name), aliases=[alias] if alias else [], kind="external",
            )
            db.add(cp)
            await db.flush()
            self.cp_by_norm[nn] = cp
            if res is not None:
                res["counterparties"] = res.get("counterparties", 0) + 1
        else:
            if _s(inn) and not _s(cp.inn):
                cp.inn = _s(inn)[:20]
            if alias and alias not in (cp.aliases or []):
                cp.aliases = sorted(set((cp.aliases or []) + [alias]))
        return cp

    async def contract(self, db: AsyncSession, company_id, cp: Counterparty,
                       ctype: str, number, cdate, basis, loc: ServiceLocation | None,
                       res: dict) -> Contract | None:
        """Договор: привязка к существующему (номер+тип у контрагента), иначе создать."""
        ids = [x for x in (cp.external_ref, str(cp.id)) if x]
        contract = None
        if number:
            for cid in ids:
                contract = self.contr_by_key.get((cid, str(number), ctype))
                if contract:
                    break
        if contract is None:
            for cid in ids:
                same = self.contr_by_cp_type.get((cid, ctype))
                if same:
                    contract = same[0]
                    break
        if contract is None:
            # Без номера обязательство всё равно есть: часть площадок стоит на
            # муниципальной земле по разрешению или сервитуту, и раньше такие строки
            # реестра оставались вне договорного контура (settlement.contract_id пустой).
            # Заводим «б/н» — основание видно в `basis`, номер проставляется вручную.
            num = str(number)[:100] if number else "б/н"
            contract = Contract(
                company_id=company_id, number=num, date=cdate or "",
                counterparty_id=str(cp.id), organization_id=self.org_id or str(cp.id),
                type=ctype, kind="СПоставщиком", basis=basis, scope_type="locations",
            )
            db.add(contract)
            await db.flush()
            self.contr_by_key[(str(cp.id), num, ctype)] = contract
            self.contr_by_cp_type.setdefault((str(cp.id), ctype), []).append(contract)
            res["contracts"] = res.get("contracts", 0) + 1
        if contract is not None and loc is not None:
            key = (str(contract.id), loc.id)
            if key not in self.contract_locs:
                db.add(ContractLocation(
                    company_id=company_id, contract_id=contract.id, location_id=loc.id))
                self.contract_locs.add(key)
        return contract

    def settlement(self, db: AsyncSession, company_id, loc: ServiceLocation,
                   role: str, fields: dict, res: dict) -> StationContractSettlement:
        """Upsert StationContractSettlement (company+location+role); None-поля не затирают."""
        s = self.settl.get((loc.id, role))
        if s is None:
            s = StationContractSettlement(company_id=company_id, location_id=loc.id, role=role)
            db.add(s)
            self.settl[(loc.id, role)] = s
            res["settlements"] = res.get("settlements", 0) + 1
        for k, v in fields.items():
            if k == "extra":
                merged = dict(s.extra or {})
                merged.update(_compact(v or {}))
                if merged:
                    s.extra = merged
            elif v is not None:
                setattr(s, k, v)
        return s

    def energy_period(self, db: AsyncSession, company_id, loc: ServiceLocation,
                      period: str, *, intake_kwh: float | None = None,
                      tariff: float | None = None, source: str | None = None,
                      res: dict | None = None) -> None:
        """Upsert StationEnergyPeriod: intake/tariff приходят из разных файлов."""
        p = self.periods.get((loc.id, period))
        if p is None:
            p = StationEnergyPeriod(
                company_id=company_id, location_id=loc.id, period=period, source=source)
            db.add(p)
            self.periods[(loc.id, period)] = p
            if res is not None:
                res["periods"] = res.get("periods", 0) + 1
        if intake_kwh is not None:
            p.intake_kwh = intake_kwh
        if tariff is not None:
            p.tariff_rub_kwh = tariff
        if source:
            p.source = source

    def dispense_period(self, db: AsyncSession, company_id, loc: ServiceLocation,
                        period: str, connector: str | None, *,
                        kwh: float | None = None, rub: float | None = None,
                        source: str | None = None, res: dict | None = None) -> None:
        """Upsert StationDispensePeriod: кВт·ч и ₽ приходят из разных листов."""
        key = (loc.id, period, connector or "")
        p = self.dispense.get(key)
        if p is None:
            p = StationDispensePeriod(
                company_id=company_id, location_id=loc.id, period=period,
                connector_type=connector or None, source=source)
            db.add(p)
            self.dispense[key] = p
            if res is not None:
                res["dispense_periods"] = res.get("dispense_periods", 0) + 1
        if kwh is not None:
            p.dispense_kwh = kwh
        if rub is not None:
            p.amount_rub = rub
        if source:
            p.source = source


# --------------------------------------------------------------------------- parse
def _load_sheet(content: bytes, title: str) -> list[tuple]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        if title not in wb.sheetnames:
            raise ValueError(f"Лист «{title}» не найден; есть: {wb.sheetnames}")
        return list(wb[title].iter_rows(values_only=True))
    finally:
        wb.close()


# «Сводная»: фиксированные колонки (см. шапку r1) — при изменении структуры
# файла контрагентом сигнатуры проверяются в _assert_header.
SVOD_COLS = {
    "zoi": 1, "serial": 3, "lat": 5, "lon": 6, "installed": 8, "removed": 9,
    "bu": 32, "bu_name": 33, "status1": 34, "comment1": 35, "inv": 36,
    "status2": 37, "remark": 38,
    "svc_cp": 39, "svc_num": 40, "svc_start": 41, "svc_end": 42, "svc_contacts": 43,
    "svc_fix_net": 44, "svc_var_net": 45,
    "en_cp": 68, "en_contract": 69, "en_max_kw": 70, "en_contacts": 71,
    "en_tariff_avg": 72, "en_paid": 73,
    "rent_cp": 74, "rent_contract": 75, "rent_start": 76, "rent_end": 77,
    "rent_contacts": 78, "rent_paid": 79, "rent_gross": 80, "rent_net": 81,
    "note": 83, "ocpp": 140,
}


def _assert_header(hdr: tuple, checks: dict[int, str], fmt: str) -> None:
    for idx, expect in checks.items():
        got = str(hdr[idx] if idx < len(hdr) else "").lower()
        if expect not in got:
            raise ValueError(
                f"Формат «{fmt}»: в колонке {idx} ожидалось «{expect}», найдено «{got[:60]}» — "
                "структура файла изменилась, обновите маппинг колонок"
            )


def parse_svodnaya(content: bytes) -> dict[str, Any]:
    rows = _load_sheet(content, "Сводная")
    hdr = rows[1]
    _assert_header(hdr, {1: "zoi-1", 32: "по бу", 68: "э/энерг", 74: "аренд"}, "Сводная")
    month_cols: dict[int, str] = {}
    for i, h in enumerate(hdr):
        if h is None or "электроэнерг" not in str(h).lower():
            continue
        iso = _month_from_header(h)
        if iso:
            month_cols[i] = iso
    out = []
    for r in rows[3:]:
        row = {k: (r[i] if i < len(r) else None) for k, i in SVOD_COLS.items()}
        if _blank(row["bu"]) and _blank(row["serial"]) and _blank(row["zoi"]):
            continue
        row["_months"] = {iso: _num(r[i]) for i, iso in month_cols.items()
                          if i < len(r) and _num(r[i]) is not None}
        out.append(row)
    return {"rows": out, "months": sorted(set(month_cols.values()))}


ARENDA_COLS = {
    "zoi": 1, "bu": 2, "region": 3, "city": 4, "object": 5, "inv": 6,
    "cp": 7, "inn": 8, "kind": 9, "dek": 10, "paid_month": 11,
    "fix_gross": 12, "fix_net": 13, "pay_sum": 14, "vat_pct": 15, "note": 16,
    "req": 17, "start": 18, "end": 19, "pay_from": 20, "lump": 21, "pay": 22,
    "docs_bu": 23, "docs_nu": 24, "fix_terms": 25, "var_gross": 26,
    "var_terms": 27, "activity": 28,
}


def parse_arenda(content: bytes) -> dict[str, Any]:
    rows = _load_sheet(content, "Договоры_Аренда")
    _assert_header(rows[3], {1: "zoi-1", 2: "по бу", 7: "контрагент", 8: "инн"},
                   "Договоры_Аренда")
    out = []
    for r in rows[5:]:
        row = {k: (r[i] if i < len(r) else None) for k, i in ARENDA_COLS.items()}
        if _blank(row["bu"]) and _blank(row["zoi"]):
            continue
        out.append(row)
    # лист «Контакты» (опционален)
    contacts = []
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        if "Контакты" in wb.sheetnames:
            for r in list(wb["Контакты"].iter_rows(values_only=True))[2:]:
                if len(r) > 3 and not _blank(r[3]):
                    contacts.append({
                        "bu": r[1], "object": r[2], "cp": r[3],
                        "contacts": r[4] if len(r) > 4 else None,
                        "exchange": r[5] if len(r) > 5 else None,
                        "comment": r[6] if len(r) > 6 else None,
                    })
    finally:
        wb.close()
    return {"rows": out, "contacts": contacts}


TARIFF_COLS = {"zoi": 1, "bu": 22, "bu_name": 23, "en_cp": 24, "en_contract": 25}


def parse_tariffs(content: bytes) -> dict[str, Any]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = None
        for cand in wb.worksheets:
            head = next(cand.iter_rows(values_only=True, max_row=1), ()) or ()
            if any("тариф электроэнерг" in str(c).lower() for c in head if c):
                ws = cand
                break
        if ws is None:
            raise ValueError("Лист с колонками «Тариф электроэнерги, <месяц>» не найден")
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    hdr = rows[0]
    _assert_header(hdr, {1: "zoi-1", 22: "по бу", 24: "э/энерг"}, "Тарифы")
    month_cols: dict[int, str] = {}
    for i, h in enumerate(hdr):
        if h is not None and "тариф электроэнерг" in str(h).lower():
            iso = _month_from_header(h)
            if iso:
                month_cols[i] = iso
    out = []
    for r in rows[1:]:
        row = {k: (r[i] if i < len(r) else None) for k, i in TARIFF_COLS.items()}
        if _blank(row["bu"]) and _blank(row["zoi"]):
            continue
        row["_tariffs"] = {iso: _num(r[i]) for i, iso in month_cols.items()
                           if i < len(r) and _num(r[i]) is not None}
        out.append(row)
    return {"rows": out, "months": sorted(set(month_cols.values()))}


# --------------------------------------------------------------------------- ingest
async def ingest_svodnaya(db: AsyncSession, company_id, parsed: dict) -> dict[str, Any]:
    resolver = StationResolver()
    await resolver.build(db, company_id)
    l2 = L2Cache()
    await l2.build(db, company_id)
    raw_existing = await _prefetch_raw(db, company_id, TAG_SVODNAYA)

    res: dict[str, Any] = {"rows": 0, "raw": 0, "unmatched": 0}
    intake: dict[tuple[str, str], float] = {}  # (loc_id, period) → кВт·ч (сумма строк объекта)
    loc_by_id: dict[str, ServiceLocation] = {}

    for row in parsed["rows"]:
        res["rows"] += 1
        bu_key = _s(row["bu"]) or _s(row["zoi"]) or _s(row["serial"])
        # На одном № БУ бывает несколько станций (705 строк / 673 БУ) — сырьё
        # различаем зав. номером, иначе вторая строка затёрла бы первую в L1.
        key = f"{bu_key}-{_s(row['serial'])}" if _s(row["serial"]) else bu_key
        loc = resolver.resolve(
            bu=row["bu"], zoi=row["zoi"], serial=row["serial"],
            ocpp=row["ocpp"], lat=row["lat"], lon=row["lon"],
        )
        # клеймо сопряжения — в L1-сырьё (единый принцип: строка знает свой объект;
        # по нему считаются «сироты» в linkage/нормализации)
        res["raw"] += await _upsert_raw(
            db, company_id, TAG_SVODNAYA, key,
            f"Сводная ЭЗС №{bu_key} — договоры/объёмы э/э",
            _jsonable_row({k: v for k, v in row.items() if k != "_months"} |
                          {"_months": row["_months"]}) |
            {"_match": {"resolved": loc is not None,
                        "locationId": loc.id if loc else None}},
            raw_existing,
        )
        if loc is None:
            res["unmatched"] += 1
            continue
        loc_by_id[loc.id] = loc

        # объёмы входящей э/э: на объекте может стоять несколько станций → суммируем
        for period, kwh in (row["_months"] or {}).items():
            intake[(loc.id, period)] = intake.get((loc.id, period), 0.0) + kwh

        note = None if _blank(row["note"]) else str(row["note"]).strip()[:1000]

        # — энергоснабжение —
        cp = await l2.counterparty(db, company_id, row["en_cp"], alias="energy", res=res)
        if cp is not None:
            number, cdate = _parse_contract_req(row["en_contract"])
            basis = _detect_basis(row["en_contract"], row["en_cp"])
            contract = await l2.contract(db, company_id, cp, "Энергоснабжение",
                                         number, cdate, basis, loc, res)
            pstatus, paid_through, pnote = _ru_month_paid(row["en_paid"])
            l2.settlement(db, company_id, loc, "energy", dict(
                counterparty_id=str(cp.id),
                contract_id=contract.id if contract else None,
                payment_status=pstatus, paid_through=paid_through,
                basis=basis, comment=note, period="2026-06-01", source=TAG_SVODNAYA,
                extra={"avgTariff": _num(row["en_tariff_avg"]),
                       "maxPowerKw": _num(row["en_max_kw"]),
                       "contacts": _s(row["en_contacts"]) or None,
                       "paidNote": pnote},
            ), res)

        # — аренда (в «Сводной» — базовый срез; файл «Договоры_Аренда» актуальнее) —
        cp = await l2.counterparty(db, company_id, row["rent_cp"], alias="rent", res=res)
        if cp is not None:
            number, cdate = _parse_contract_req(row["rent_contract"])
            basis = _detect_basis(row["rent_contract"], row["rent_cp"])
            contract = await l2.contract(db, company_id, cp, "Аренда",
                                         number, cdate, basis, loc, res)
            pstatus, paid_through, pnote = _ru_month_paid(row["rent_paid"])
            l2.settlement(db, company_id, loc, "rent", dict(
                counterparty_id=str(cp.id),
                contract_id=contract.id if contract else None,
                payment_status=pstatus, paid_through=paid_through,
                basis=basis, comment=note,
                amount_gross=_num(row["rent_gross"]), amount_net=_num(row["rent_net"]),
                contract_start=_iso_date(row["rent_start"]),
                contract_end=_iso_date(row["rent_end"]),
                period="2026-06-01", source=TAG_SVODNAYA,
                extra={"contacts": _s(row["rent_contacts"]) or None, "paidNote": pnote},
            ), res)

        # — сервисный договор (обслуживание ЭЗС) —
        cp = await l2.counterparty(db, company_id, row["svc_cp"], alias="service", res=res)
        if cp is not None:
            number, cdate = _parse_contract_req(row["svc_num"])
            contract = await l2.contract(db, company_id, cp, "Сервис",
                                         number, cdate or _iso_date(row["svc_start"]),
                                         "договор", loc, res)
            l2.settlement(db, company_id, loc, "service", dict(
                counterparty_id=str(cp.id),
                contract_id=contract.id if contract else None,
                payment_status="unknown",
                basis="договор",
                amount_net=_num(row["svc_fix_net"]),
                contract_start=_iso_date(row["svc_start"]),
                contract_end=_iso_date(row["svc_end"]),
                period="2026-06-01", source=TAG_SVODNAYA,
                extra={"variableNet": _num(row["svc_var_net"]),
                       "contacts": _s(row["svc_contacts"]) or None},
            ), res)

    for (loc_id, period), kwh in intake.items():
        l2.energy_period(db, company_id, loc_by_id[loc_id], period,
                         intake_kwh=round(kwh, 3), source=TAG_SVODNAYA, res=res)

    await db.flush()
    res["resolve"] = resolver.stats
    return {"status": "success", "kind": "reestr_svodnaya",
            "shifts": res["rows"],
            "created": res.get("settlements", 0) + res.get("periods", 0),
            "updated": 0, "skipped_kinds": [],
            "message": (f"Сводная: строк {res['rows']}, объектов не сопоставлено "
                        f"{res['unmatched']}; периодов э/э {res.get('periods', 0)}"),
            **res}


async def ingest_arenda(db: AsyncSession, company_id, parsed: dict) -> dict[str, Any]:
    resolver = StationResolver()
    await resolver.build(db, company_id)
    l2 = L2Cache()
    await l2.build(db, company_id)
    raw_existing = await _prefetch_raw(db, company_id, TAG_ARENDA)

    res: dict[str, Any] = {"rows": 0, "raw": 0, "unmatched": 0, "inn_filled": 0}

    for row in parsed["rows"]:
        res["rows"] += 1
        bu_key = _s(row["bu"]) or _s(row["zoi"])
        key = f"{bu_key}-{_s(row['zoi'])}" if _s(row["zoi"]) else bu_key
        loc = resolver.resolve(bu=row["bu"], zoi=row["zoi"])
        res["raw"] += await _upsert_raw(
            db, company_id, TAG_ARENDA, key,
            f"Аренда ЭЗС №{bu_key} — договор/постоянная часть",
            _jsonable_row(row) | {"_match": {"resolved": loc is not None,
                                             "locationId": loc.id if loc else None}},
            raw_existing,
        )
        if loc is None:
            res["unmatched"] += 1
            continue

        had_inn = bool(_s(row["inn"]))
        cp = await l2.counterparty(db, company_id, row["cp"], inn=_s(row["inn"]),
                                   alias="rent", res=res)
        if cp is None:
            continue
        if had_inn and _s(cp.inn) == _s(row["inn"]):
            res["inn_filled"] += 1

        number, cdate = _parse_contract_req(row["req"])
        kind_txt = _s(row["kind"]).lower()
        basis = _detect_basis(row["req"], row["kind"]) if kind_txt != "аренда" else "договор"
        contract = await l2.contract(db, company_id, cp, "Аренда",
                                     number, cdate or _iso_date(row["start"]),
                                     basis, loc, res)
        pstatus, paid_through, pnote = _ru_month_paid(row["paid_month"])
        note = None if _blank(row["note"]) else str(row["note"]).strip()[:1000]
        l2.settlement(db, company_id, loc, "rent", dict(
            counterparty_id=str(cp.id),
            contract_id=contract.id if contract else None,
            payment_status=pstatus, paid_through=paid_through,
            basis=basis, comment=note,
            amount_gross=_num(row["fix_gross"]), amount_net=_num(row["fix_net"]),
            vat_pct=_num(row["vat_pct"]),
            contract_start=_iso_date(row["start"]), contract_end=_iso_date(row["end"]),
            period="2026-06-01", source=TAG_ARENDA,
            extra={"variableGross": _num(row["var_gross"]),
                   "fixTerms": _s(row["fix_terms"]) or None,
                   "varTerms": _s(row["var_terms"]) or None,
                   "lumpSum": _s(row["lump"]) or None,
                   "activity": _s(row["activity"]) or None,
                   "dek": _s(row["dek"]) or None,
                   "paySum": _num(row["pay_sum"]),
                   "paidNote": pnote,
                   "contractKind": _s(row["kind"]) or None},
        ), res)

    # лист «Контакты» → карточки контрагентов. У контрагента бывает несколько
    # строк (разные ЭЗС) — аккумулируем уникальные варианты и пишем ОДИН раз,
    # иначе последняя строка флапала бы значение при каждом прогоне.
    by_cp: dict[str, list[dict]] = {}
    for c in parsed.get("contacts", []):
        nn = _normname(c["cp"])
        if nn in l2.cp_by_norm:
            item = _compact({
                "contacts": _s(c.get("contacts")) or None,
                "exchange": _s(c.get("exchange")) or None,
                "comment": _s(c.get("comment")) or None,
            })
            if item and item not in by_cp.setdefault(nn, []):
                by_cp[nn].append(item)
    for nn, items in by_cp.items():
        cp = l2.cp_by_norm[nn]
        if not _s(cp.phone):
            for it in items:
                m = re.search(r"(\+?[78][\s(]?\d{3}[\s)]?[\s-]?\d{3}[\s-]?\d{2}[\s-]?\d{2})",
                              it.get("contacts") or "")
                if m:
                    cp.phone = m.group(1)[:100]
                    break
        raw = dict(cp.raw or {})
        if raw.get("reestrContacts") != items:
            raw["reestrContacts"] = items
            cp.raw = raw
            res["contacts"] = res.get("contacts", 0) + 1

    await db.flush()
    res["resolve"] = resolver.stats
    return {"status": "success", "kind": "reestr_arenda",
            "shifts": res["rows"],
            "created": res.get("settlements", 0),
            "updated": 0, "skipped_kinds": [],
            "message": (f"Аренда: строк {res['rows']}, не сопоставлено {res['unmatched']}; "
                        f"ИНН заполнено {res['inn_filled']}, контактов {res.get('contacts', 0)}"),
            **res}


async def ingest_tariffs(db: AsyncSession, company_id, parsed: dict) -> dict[str, Any]:
    resolver = StationResolver()
    await resolver.build(db, company_id)
    l2 = L2Cache()
    await l2.build(db, company_id)
    raw_existing = await _prefetch_raw(db, company_id, TAG_TARIFFS)

    res: dict[str, Any] = {"rows": 0, "raw": 0, "unmatched": 0, "tariffs": 0}
    # тариф на объект: строка = станция; на объекте берём последнее непустое значение
    per_loc: dict[tuple[str, str], float] = {}
    loc_by_id: dict[str, ServiceLocation] = {}

    for row in parsed["rows"]:
        res["rows"] += 1
        key = _s(row["zoi"]) or _s(row["bu"])
        loc = resolver.resolve(bu=row["bu"], zoi=row["zoi"])
        res["raw"] += await _upsert_raw(
            db, company_id, TAG_TARIFFS, key,
            f"Тарифы э/э ЭЗС №{key} — входящие",
            _jsonable_row({k: v for k, v in row.items() if k != "_tariffs"} |
                          {"_tariffs": row["_tariffs"]}) |
            {"_match": {"resolved": loc is not None,
                        "locationId": loc.id if loc else None}},
            raw_existing,
        )
        if loc is None:
            res["unmatched"] += 1
            continue
        loc_by_id[loc.id] = loc
        for period, val in (row["_tariffs"] or {}).items():
            per_loc[(loc.id, period)] = val

        # контрагент э/э как добор (если «Сводная» ещё не грузилась)
        cp = await l2.counterparty(db, company_id, row["en_cp"], alias="energy", res=res)
        if cp is not None and (loc.id, "energy") not in l2.settl:
            number, cdate = _parse_contract_req(row["en_contract"])
            contract = await l2.contract(db, company_id, cp, "Энергоснабжение",
                                         number, cdate, "договор", loc, res)
            l2.settlement(db, company_id, loc, "energy", dict(
                counterparty_id=str(cp.id),
                contract_id=contract.id if contract else None,
                payment_status="unknown", basis="договор",
                period="2026-06-01", source=TAG_TARIFFS,
            ), res)

    for (loc_id, period), val in per_loc.items():
        l2.energy_period(db, company_id, loc_by_id[loc_id], period,
                         tariff=val, source=TAG_TARIFFS, res=res)
        res["tariffs"] += 1

    await db.flush()
    res["resolve"] = resolver.stats
    return {"status": "success", "kind": "reestr_tariffs",
            "shifts": res["rows"],
            "created": res["tariffs"],
            "updated": 0, "skipped_kinds": [],
            "message": (f"Тарифы: строк {res['rows']}, не сопоставлено {res['unmatched']}; "
                        f"тарифных значений {res['tariffs']} за {len(parsed['months'])} мес."),
            **res}


# --------------------------------------------------------------------------- obshaya
# Сводная выработка «ОБЩАЯ_2024-2026» — четвёртый слот канала реестров.
# кВт·ч: «Коннектор_общая» (коннектор×месяц, ~472 станции) + добор станций без
# коннекторной детализации из станционного листа «Общие итоги» (478; «Медленные» —
# его подмножество, не читается). ₽: скрытый лист «Сумма» (коннектор×месяц,
# 01.2024–07.2025). Сумма коннекторов сверена со станционным листом (0 расхожде-
# ний) — поэтому кВт·ч храним ТОЛЬКО на одной гранулярности на станцию (без
# двойного счёта: SUM по станции всегда корректен). Паспортные атрибуты станции
# (трасса/город, Быстрая/Медленная, марка, даты жизненного цикла, инвентарный
# номер, статус корп) обогащают service_locations из станционного листа.

def _pv(v) -> str:
    """Значение ячейки сводной: '(пусто)' (подпись Excel-pivot) → ''."""
    s = _s(v)
    return "" if s.lower() in ("(пусто)", "пусто") else s


def _hdr_index(hdr: tuple, want: dict[str, tuple[str, ...]]) -> dict[str, int]:
    """Индексы колонок по подстрокам заголовков (первая подошедшая колонка)."""
    out: dict[str, int] = {}
    for i, h in enumerate(hdr):
        if h is None:
            continue
        low = str(h).strip().lower()
        for key, pats in want.items():
            if key not in out and any(p in low for p in pats):
                out[key] = i
                break
    return out


_OBSH_STATION_COLS = {
    "bu": ("б/у",), "zoi": ("zoi",), "region": ("регион",),
    "loc_class": ("трасса",), "city": ("город",), "address": ("адрес",),
    "speed": ("быстрая",), "brand": ("марка", "производитель", "пр-тель"),
    "power": ("мощность",), "lat": ("широта",), "lon": ("долгота",),
    "corp": ("корп",), "installed": ("дата установки",),
    "decommissioned": ("вывода",), "inv": ("инвентарный",),
    "connector": ("connector",),
}


def _parse_obsh_sheet(rows: list[tuple], fmt: str, *, need_connector: bool) -> list[dict]:
    """Общий разбор листа сводной: шапка r2 (или r1), месяцы после паспортных
    колонок. Дубли ключа (станция/коннектор) агрегируются суммой месяцев."""
    hrow = 1 if rows and rows[0] and any(c is not None for c in rows[0]) else 1
    # шапка — первая строка, где есть «zoi»
    hdr_i = next((i for i, r in enumerate(rows[:4])
                  if any(c is not None and "zoi" in str(c).lower() for c in r)), None)
    if hdr_i is None:
        raise ValueError(f"«{fmt}»: не найдена шапка с колонкой ZOI-1")
    hdr = rows[hdr_i]
    idx = _hdr_index(hdr, _OBSH_STATION_COLS)
    if "zoi" not in idx and "bu" not in idx:
        raise ValueError(f"«{fmt}»: нет ключевых колонок Б/У / № ZOI-1")
    mstart = max(idx.values()) + 1
    month_cols = {i: iso for i in range(mstart, len(hdr))
                  if (iso := _month_short(hdr[i])) is not None}
    if not month_cols:
        raise ValueError(f"«{fmt}»: не найдены колонки месяцев")

    agg: dict[tuple, dict] = {}
    for r in rows[hdr_i + 1:]:
        get = lambda k: (r[idx[k]] if k in idx and idx[k] < len(r) else None)  # noqa: E731
        bu, zoi = _pv(get("bu")), _pv(get("zoi"))
        if not bu and not zoi:
            continue
        region = _pv(get("region"))
        if region.lower().endswith("сумма"):
            continue  # агрегатные строки Excel-свода
        ct = _pv(get("connector"))[:40]
        if need_connector and not ct:
            continue  # в коннекторном листе строка без коннектора = агрегат
        key = (zoi or bu, ct if need_connector else "")
        row = agg.get(key)
        if row is None:
            row = {k: get(k) for k in _OBSH_STATION_COLS}
            row["connector"] = ct or None
            row["_months"] = {}
            agg[key] = row
        for i, iso in month_cols.items():
            v = _num(r[i]) if i < len(r) else None
            if v is not None:
                row["_months"][iso] = row["_months"].get(iso, 0.0) + v
    return list(agg.values())


def parse_obshaya(content: bytes) -> dict[str, Any]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        names = set(wb.sheetnames)

        def sheet(title: str) -> list[tuple]:
            return list(wb[title].iter_rows(values_only=True)) if title in names else []

        stations = _parse_obsh_sheet(sheet("Общие итоги"), "Общие итоги",
                                     need_connector=False) if "Общие итоги" in names else []
        connectors = _parse_obsh_sheet(sheet("Коннектор_общая"), "Коннектор_общая",
                                       need_connector=True) if "Коннектор_общая" in names else []
        rub = _parse_obsh_sheet(sheet("Сумма"), "Сумма (₽)",
                                need_connector=True) if "Сумма" in names else []
        if not stations and not connectors:
            raise ValueError("Сводная выработка: нет листов «Общие итоги» / «Коннектор_общая»")
        return {"stations": stations, "connectors": connectors, "rub": rub}
    finally:
        wb.close()


def _enrich_passport(loc: ServiceLocation, row: dict, unit=None) -> bool:
    """Обогащение паспорта станции из станционной строки сводной.
    Классификаторы (city/highway, fast/slow, корп) переписываются свежим файлом;
    факты (даты, инвентарный, марка, мощность) — только дозаполняются."""
    changed = False

    def setattr_if(attr: str, value) -> None:
        nonlocal changed
        if value is not None and getattr(loc, attr) != value:
            setattr(loc, attr, value)
            changed = True

    lc = _pv(row.get("loc_class")).lower()
    if lc.startswith("город"):
        setattr_if("location_class", "city")
    elif lc.startswith("трасс"):
        setattr_if("location_class", "highway")
    # Класс скорости — характеристика железа, поэтому едет владельцу графы, а не
    # в колонку точки: `setattr_if` пишет только в `loc` и здесь не подходит.
    sp = _pv(row.get("speed")).lower()
    speed = "fast" if sp.startswith("быстр") else "slow" if sp.startswith("медлен") else None
    if speed and write_value("speedClass", loc, unit, speed):
        changed = True
    if _pv(row.get("corp")).lower().startswith("корп") and not loc.is_corp:
        loc.is_corp = True
        changed = True
    inst = _iso_date(row.get("installed"))
    if inst and not passport_value("installedOn", loc, unit):
        write_value("installedOn", loc, unit, inst)
        changed = True
    dec_raw = row.get("decommissioned")
    dec = _iso_date(dec_raw)
    if dec and not passport_value("decommissionedOn", loc, unit):
        write_value("decommissionedOn", loc, unit, dec)
        changed = True
    elif dec is None and not _blank(dec_raw):
        s = str(dec_raw).strip()
        # «Установлена» = работает (не вывод); прочий текст без даты — в примечание
        if "установлен" not in s.lower() and s != "(пусто)":
            md = dict(loc.extra_metadata or {})
            if md.get("decommissionNote") != s[:200]:
                md["decommissionNote"] = s[:200]
                loc.extra_metadata = md
                changed = True
    inv = _pv(row.get("inv"))
    if (inv and not passport_value("inventoryNumber", loc, unit)
            and "нет" not in inv.lower() and "не извест" not in inv.lower()):
        write_value("inventoryNumber", loc, unit, inv[:60])
        changed = True
    brand = _pv(row.get("brand"))
    if brand and not passport_value("brand", loc, unit):
        write_value("brand", loc, unit, brand[:120])
        changed = True
    pw = _num(row.get("power"))
    if pw and not passport_value("powerKwt", loc, unit):
        write_value("powerKwt", loc, unit, pw)
        changed = True
    return changed


async def ingest_obshaya(db: AsyncSession, company_id, parsed: dict) -> dict[str, Any]:
    resolver = StationResolver()
    await resolver.build(db, company_id)
    l2 = L2Cache()
    await l2.build(db, company_id)
    raw_existing = await _prefetch_raw(db, company_id, TAG_OBSHAYA)

    res: dict[str, Any] = {"rows": 0, "raw": 0, "unmatched": 0, "enriched": 0}

    def skey(row: dict) -> str:
        return _pv(row.get("zoi")) or _pv(row.get("bu"))

    def resolve(row: dict) -> ServiceLocation | None:
        return resolver.resolve(bu=_pv(row.get("bu")), zoi=_pv(row.get("zoi")),
                                lat=row.get("lat"), lon=row.get("lon"))

    # станции, покрытые коннекторным листом кВт·ч, — станционный ряд для них не пишем
    conn_covered = {skey(r) for r in parsed["connectors"] if r.get("_months")}

    # Паспорт железа принадлежит станции (СТО, docs/OBJECTS.md); карта по всем
    # точкам компании — какие встретятся в реестре, заранее неизвестно.
    stations = await stations_by_location(db, company_id, None)

    # 1) станционный лист: L1 + паспорт + кВт·ч только для непокрытых коннекторами
    for row in parsed["stations"]:
        res["rows"] += 1
        loc = resolve(row)
        res["raw"] += await _upsert_raw(
            db, company_id, TAG_OBSHAYA, f"st-{skey(row)}",
            f"Выработка ЭЗС №{skey(row)} — сводная помесячно (станция)",
            _jsonable_row({k: v for k, v in row.items() if k != "_months"}) |
            {"_months": row["_months"], "_grain": "station",
             "_match": {"resolved": loc is not None,
                        "locationId": loc.id if loc else None}},
            raw_existing,
        )
        if loc is None:
            res["unmatched"] += 1
            continue
        if _enrich_passport(loc, row, stations.get(loc.id)):
            res["enriched"] += 1
        if skey(row) not in conn_covered:
            for period, kwh in (row["_months"] or {}).items():
                l2.dispense_period(db, company_id, loc, period, None,
                                   kwh=round(kwh, 3), source=TAG_OBSHAYA, res=res)

    # 2) коннекторный лист кВт·ч (основная гранулярность)
    for row in parsed["connectors"]:
        res["rows"] += 1
        loc = resolve(row)
        ct = row.get("connector")
        res["raw"] += await _upsert_raw(
            db, company_id, TAG_OBSHAYA, f"conn-{skey(row)}-{ct or 'na'}",
            f"Выработка ЭЗС №{skey(row)} · {ct or '—'} — кВт·ч помесячно",
            _jsonable_row({k: v for k, v in row.items() if k != "_months"}) |
            {"_months": row["_months"], "_grain": "connector",
             "_match": {"resolved": loc is not None,
                        "locationId": loc.id if loc else None}},
            raw_existing,
        )
        if loc is None:
            res["unmatched"] += 1
            continue
        for period, kwh in (row["_months"] or {}).items():
            l2.dispense_period(db, company_id, loc, period, ct,
                               kwh=round(kwh, 3), source=TAG_OBSHAYA, res=res)

    # 3) ₽-лист «Сумма» (01.2024–07.2025): рубли ложатся в те же ключи периодов
    for row in parsed["rub"]:
        res["rows"] += 1
        loc = resolve(row)
        ct = row.get("connector")
        res["raw"] += await _upsert_raw(
            db, company_id, TAG_OBSHAYA, f"rub-{skey(row)}-{ct or 'na'}",
            f"Выручка ЭЗС №{skey(row)} · {ct or '—'} — ₽ помесячно",
            _jsonable_row({k: v for k, v in row.items() if k != "_months"}) |
            {"_months": row["_months"], "_grain": "rub",
             "_match": {"resolved": loc is not None,
                        "locationId": loc.id if loc else None}},
            raw_existing,
        )
        if loc is None:
            res["unmatched"] += 1
            continue
        for period, rub_v in (row["_months"] or {}).items():
            l2.dispense_period(db, company_id, loc, period, ct,
                               rub=round(rub_v, 2), source=TAG_OBSHAYA, res=res)

    await db.flush()
    res["resolve"] = resolver.stats
    return {"status": "success", "kind": "reestr_obshaya",
            "shifts": res["rows"],
            "created": res.get("dispense_periods", 0),
            "updated": 0, "skipped_kinds": [],
            "message": (f"Сводная выработка: строк {res['rows']}, не сопоставлено "
                        f"{res['unmatched']}; периодов отпуска {res.get('dispense_periods', 0)}, "
                        f"паспортов обогащено {res['enriched']}"),
            **res}


# --------------------------------------------------------------------------- dispatch
# Порядок обработки слотов при полном прогоне: сначала якорная «Сводная» (она
# прописывает станциям buNumber/zoi1), затем уточняющие аренда и тарифы.
REESTR_SLOT_ORDER = ("svod", "svodnaya", "obshaya", "arenda", "tariffs")
REESTR_SLOT_LABEL = {
    "svod": "Общий свод (старый формат)",
    "svodnaya": "Сводная: договоры + объёмы э/э",
    "obshaya": "Сводная выработка 2024–2026 (кВт·ч/₽)",
    "arenda": "Договоры аренды (актуальные)",
    "tariffs": "Тарифы э/э входящие",
}


async def run_reestr_file(db: AsyncSession, company_id, content: bytes) -> dict[str, Any]:
    """Определить формат загруженного файла и прогнать соответствующий ингест.
    В результате всегда есть поле `format` — по нему оркестратор ведёт слоты."""
    fmt = detect_reestr_format(content)
    if fmt == "svod":
        from app.services.reestr_normalize import ingest_reestr, parse_reestr_xlsx
        res = await ingest_reestr(db, company_id, parse_reestr_xlsx(content))
    elif fmt == "svodnaya":
        res = await ingest_svodnaya(db, company_id, parse_svodnaya(content))
    elif fmt == "obshaya":
        res = await ingest_obshaya(db, company_id, parse_obshaya(content))
    elif fmt == "arenda":
        res = await ingest_arenda(db, company_id, parse_arenda(content))
    else:
        res = await ingest_tariffs(db, company_id, parse_tariffs(content))
    res["format"] = fmt
    return res
