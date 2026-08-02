"""Импорт рабочего реестра работ по парку (переносы, замены, демонтажи).

ЗАЧЕМ. Заказчик ведёт эти работы в отдельном файле по регионам: что за станция,
что с ней делаем, кто подрядчик, к какому сроку. Экран «План работ» отвечает на тот
же вопрос, но пока он пуст, файл остаётся правдой, а система — витриной. Импорт
переносит реестр разом; дальше работа ведётся в системе, а файл больше не нужен.

ГРАНИЦА, КОТОРУЮ ИМПОРТ ДЕРЖИТ. Лист «Ремонты» проектами НЕ становится: замена
автомата в щите — заявка поддержки, а не капвложение (решение МАГа 28.07.2026).
Такие строки импорт не заводит, а возвращает списком: их место — в заявках по
объекту, кнопка есть в карточке.

ИДЕНТИЧНОСТЬ строки — лист плюс нормализованный адрес (`dedup_key`). Повторная
загрузка того же файла обновляет комментарий, срок и подрядчика, но не плодит
вторую работу и не откатывает стадию: стадию ведёт система, а не файл.
"""
from __future__ import annotations

import io
import re
import uuid
from datetime import date, datetime, timezone
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import EzsProject, EzsSite, ServiceLocation
from app.services.ezs_sites import format_project_no, parse_project_seq, project_no_prefix

# Лист файла → что с ним делать. `kind=None` — вид работы берём из строки.
# `stage` — стадия, с которой работа встаёт в систему.
SHEETS: dict[str, dict[str, Any]] = {
    "объекты все регионы": {"kind": None, "stage": "construction"},
    "комплексные _дв":     {"kind": None, "stage": "decision"},
    "пир":                 {"kind": "retrofit", "stage": "dd"},
    "разобрать":           {"kind": "retrofit", "stage": "decision"},
    "выполненные":         {"kind": None, "stage": "commissioning"},
    "закрытые работы":     {"kind": None, "stage": "live", "closed": True},
    # Ремонт — заявка поддержки, а не проект. Строки собираем и отдаём отдельно.
    "ремонты":             {"tickets": True},
    # Справочник выпадающих списков — не данные.
    "данные для списка":   {"skip": True},
}

# Слово из файла → вид работы. Порядок важен: «перемещение на склад (утилизация)»
# это демонтаж, а не перенос, поэтому утилизацию проверяем раньше перемещения.
# Корни, а не слова целиком: в файле пишут «демонтирована», «заменена», «замене» —
# это те же работы. «демонти» проверяем раньше «монтаж», иначе демонтаж станет стройкой.
KIND_WORDS: list[tuple[str, str]] = [
    ("утилизац", "decommission"),
    ("демонт", "decommission"),
    ("замен", "retrofit"),
    ("ремонт", "retrofit"),
    ("модерниз", "retrofit"),
    ("перемещ", "relocation"),
    ("перенос", "relocation"),
    ("монтаж", "new_build"),
]

# Заголовки, которые нам нужны. Ключ — поле нашей строки, значения — что искать
# в шапке листа (нормализованно, по вхождению).
HEADERS: list[tuple[str, tuple[str, ...]]] = [
    ("region", ("регион",)),
    ("address", ("адрес",)),
    ("supplier", ("производитель",)),
    ("contractor", ("подрядчик",)),
    ("comment", ("комментарий",)),
    ("work", ("новый монтаж", "следующее действие")),
    ("due", ("срок",)),
    ("status", ("статус",)),
    ("priority", ("приоритет",)),
    ("counterparty", ("контрагент",)),
]

# Листы «Закрытые работы» и «РАЗОБРАТЬ» шапки не имеют вовсе — данные начинаются
# сразу, но раскладка колонок у файла общая. Берём её, иначе эти листы теряются.
DEFAULT_COLS: dict[str, int] = {
    "region": 0, "address": 1, "supplier": 2, "counterparty": 5,
    "comment": 11, "contractor": 12, "priority": 13, "due": 14,
}

DATE_RE = re.compile(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})")
EZS_NO_RE = re.compile(r"эзс\s*№?\s*(\d+)", re.I)


def _norm(v: Any) -> str:
    return " ".join(str(v or "").strip().lower().split())


def _text(v: Any) -> str:
    return " ".join(str(v).strip().split()) if v is not None else ""


def _kind_of(text: str) -> str | None:
    low = _norm(text)
    for word, kind in KIND_WORDS:
        if word in low:
            return kind
    return None


def _due(v: Any) -> str | None:
    """Срок из ячейки: датой, текстом «срок работ до 26.07.2026» или пусто."""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    m = DATE_RE.search(str(v or ""))
    if not m:
        return None
    d, mo, y = (int(x) for x in m.groups())
    try:
        return date(y, mo, d).isoformat()
    except ValueError:
        return None


def _addr_key(address: str) -> str:
    """Ключ адреса: номер ЭЗС, если он есть, иначе сам адрес без мусора.

    Номер станции надёжнее адреса: один и тот же объект в файле записан то
    «ЭЗС №206, ТЦ Столица», то «ТЦ Столица, г. Чита» — а номер один.
    """
    m = EZS_NO_RE.search(address)
    if m:
        return f"ezs{m.group(1)}"
    return re.sub(r"[^а-яa-z0-9]+", "", _norm(address))[:120]


async def import_park_plan_xlsx(
    db: AsyncSession, company_id, content: bytes, dry_run: bool,
) -> dict[str, Any]:
    """Загрузить реестр работ по парку. UPSERT по (лист, адрес), стадию не трогаем."""
    import openpyxl

    existing = (await db.execute(
        select(EzsSite).where(EzsSite.company_id == company_id))).scalars().all()
    by_key = {s.dedup_key: s for s in existing if s.dedup_key}

    # Объекты сети — чтобы работа встала на существующую станцию, а не завела
    # рядом вторую карточку того же места.
    locations = (await db.execute(
        select(ServiceLocation).where(
            ServiceLocation.company_id == company_id))).scalars().all()
    loc_by_no: dict[str, str] = {}
    for loc in locations:
        for src in (loc.code or "", loc.name or ""):
            m = EZS_NO_RE.search(src) or re.fullmatch(r"\D*(\d{1,4})\D*", src or "")
            if m:
                loc_by_no.setdefault(m.group(1), loc.id)

    prefix = project_no_prefix()
    seq = parse_project_seq((await db.execute(
        select(EzsSite.project_no).where(
            EzsSite.company_id == company_id,
            EzsSite.project_no.like(f"{prefix}%"))
        .order_by(EzsSite.project_no.desc()).limit(1))).scalar())

    now = datetime.now(timezone.utc)
    today = date.today().isoformat()
    report: dict[str, Any] = {
        "dryRun": dry_run, "sheets": [], "created": 0, "updated": 0,
        "linkedToObject": 0, "toFunnel": 0, "skipped": [], "tickets": [],
        "unknownSheets": [],
    }

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    for sn in wb.sheetnames:
        rule = SHEETS.get(_norm(sn))
        if rule is None:
            report["unknownSheets"].append(sn)
            continue
        if rule.get("skip"):
            continue

        rows = list(wb[sn].iter_rows(values_only=True))
        colmap, header_idx = {}, None
        for ri, row in enumerate(rows[:6]):
            cm: dict[str, int] = {}
            for ci, cell in enumerate(row):
                h = _norm(cell)
                if not h:
                    continue
                for field, words in HEADERS:
                    if field not in cm and any(w in h for w in words):
                        cm[field] = ci
                        break
            if "address" in cm and len(cm) >= 4:
                colmap, header_idx = cm, ri
                break
        if header_idx is None:
            # Шапки нет — читаем по общей раскладке файла с первой строки.
            colmap, header_idx = dict(DEFAULT_COLS), -1

        def cell(row: tuple, field: str) -> Any:
            ci = colmap.get(field)
            return row[ci] if ci is not None and ci < len(row) else None

        cnt = 0
        for ri, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            address = _text(cell(row, "address"))
            if not address:
                continue
            cnt += 1
            comment = _text(cell(row, "comment"))
            work = _text(cell(row, "work"))
            status = _text(cell(row, "status"))

            if rule.get("tickets"):
                # Ремонт проектом не заводим — отдаём строкой, чтобы завести заявкой.
                report["tickets"].append({
                    "sheet": sn, "row": ri, "region": _text(cell(row, "region")),
                    "address": address, "comment": comment,
                    "due": _due(cell(row, "due")),
                })
                continue

            if "не трогаем" in _norm(work):
                # Решение уже принято в файле: работы нет — значит и работы в системе нет.
                report["skipped"].append({
                    "sheet": sn, "row": ri, "address": address,
                    "reason": "в файле помечено «Не трогаем»",
                })
                continue
            kind = rule.get("kind") or _kind_of(work) or _kind_of(comment)
            if kind is None:
                report["skipped"].append({
                    "sheet": sn, "row": ri, "address": address,
                    "reason": f"вид работы не распознан: «{(work or comment or '—')[:120]}»",
                })
                continue
            if kind == "new_build":
                # Новая стройка — не работа по парку, её ведёт воронка проектов.
                report["toFunnel"] += 1

            key = f"park:{_norm(sn)}:{_addr_key(address)}"
            m = EZS_NO_RE.search(address)
            location_id = loc_by_no.get(m.group(1)) if m else None
            due = _due(cell(row, "due"))
            region = _text(cell(row, "region"))
            note = " · ".join(x for x in (comment, status) if x)[:2000]

            site = by_key.get(key)
            if site is not None:
                # Обновляем то, что ведёт файл: срок, комментарий, подрядчика.
                # Стадию и ответственного не трогаем — их ведёт система.
                site.next_action = (work or comment or site.next_action or "")[:300] or None
                site.next_action_due = due or site.next_action_due
                site.comment = note or site.comment
                site.contractor = _text(cell(row, "contractor")) or site.contractor
                site.updated_at = now
                report["updated"] += 1
                continue

            seq += 1
            site = EzsSite(
                company_id=company_id,
                project_no=format_project_no(prefix, seq),
                title=f"{address[:200]}",
                kind=kind,
                stage="archive" if rule.get("closed") else rule["stage"],
                stage_since=today,
                region=region or None, region_norm=region or None,
                address=address[:500], full_address=address[:500],
                supplier=_text(cell(row, "supplier")) or None,
                contractor=_text(cell(row, "contractor")) or None,
                owner=_text(cell(row, "counterparty")) or None,
                comment=note or None,
                next_action=(work or comment or "")[:300] or None,
                next_action_due=due,
                location_id=location_id,
                archive_reason="Работа выполнена (импорт реестра)" if rule.get("closed") else None,
                dedup_key=key, source_sheet=sn, row_no=ri,
                manual_fields=[],
                created_at=now, updated_at=now,
            )
            db.add(site)
            await db.flush()
            db.add(EzsProject(
                company_id=company_id, site_id=site.id, location_id=location_id,
                kind=kind, project_no=site.project_no, title=site.title,
                stage=site.stage, stage_since=site.stage_since,
                next_action=site.next_action, next_action_due=site.next_action_due,
                created_at=now, updated_at=now,
            ))
            by_key[key] = site
            report["created"] += 1
            if location_id:
                report["linkedToObject"] += 1

        report["sheets"].append({"sheet": sn, "rows": cnt})

    if dry_run:
        await db.rollback()
    else:
        await db.commit()
    return report


def demo() -> None:
    """Проверка разбора: срок, вид работы, ключ адреса."""
    assert _due(datetime(2026, 7, 30)) == "2026-07-30"
    assert _due("срок выполнения работ до 26.07.2026") == "2026-07-26"
    assert _due("05.02.2026.") == "2026-02-05"
    assert _due("") is None and _due("без срока") is None
    assert _due("31.02.2026") is None, "несуществующая дата не должна проходить"

    assert _kind_of("Перемещение на склад (утилизация)") == "decommission"
    assert _kind_of("Станция демонтирована и перевезена") == "decommission"
    # «демонтаж» содержит «монтаж» — корень демонтажа обязан проверяться раньше.
    assert _kind_of("демонтаж") == "decommission"
    assert _kind_of("Станция ПСС заменена на ФОРУ") == "retrofit"
    assert _kind_of("замена") == "retrofit"
    assert _kind_of("перемещение") == "relocation"
    assert _kind_of("новый монтаж") == "new_build"
    assert _kind_of("станция работает") is None

    # Один объект, две записи адреса — ключ должен совпасть.
    assert _addr_key("ЭЗС №206, ТЦ Столица г. Чита") == _addr_key("ТЦ Столица, ЭЗС № 206")
    assert _addr_key("Кафе Отдых, п.Кировский") != _addr_key("СТО Маслёнка, п.Лучегорск")
    print("ok")


if __name__ == "__main__":
    demo()
