"""
Нормализация реестра «Договоры и оплаты ЭЗС» (energy/РусГидро) — SERVER-SIDE,
в идеологии Источник → Канал → L1 RAW → Нормализация → L2 → Разрез.

Канал `reestr_contracts_payments` (источник типа `manual_table`):
  fetch  — загруженный xlsx (SourceFile) парсится здесь (parse_reestr_xlsx);
  L1 RAW — DataEntry(layer='raw') пер-строку (сырьё как пришло, идемпотентно);
  normalize→L2 — контрагенты/договоры/StationContractSettlement (типизир. L2,
                 аналог FuelShift) + DataEntry(layer='clean', derived_from raw);
  разрез — записи несут role=energy|rent → разрезы «Поставщики э/э»/«Аренда».

Вся бизнес-логика (резолв станции по buNumber, разбор оплаты «оплачено по»,
основание договор/разрешение) — ЗДЕСЬ, на сервере, не на клиенте.
Опора: SOURCE_CONTRACTS_PAYMENTS_RUSHYDRO.md.
"""
from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Contract,
    ContractLocation,
    Counterparty,
    DataEntry,
    Organization,
    ServiceLocation,
    StationContractSettlement,
)

SHEET = "Общий свод"
SOURCE_TAG = "reestr-ezs-oplaty-dogovory"
COL = {
    "bu": 0, "name": 1,
    "energy_cp": 2, "energy_contract": 3, "energy_pay": 4,
    "rent_cp": 5, "rent_contract": 6, "rent_pay": 7,
    "comment": 8,
}
ROLE_DEF = {
    "energy": ("Энергоснабжение", "СПоставщиком"),
    "rent": ("Аренда", "СПоставщиком"),
}
ROLE_COLS = {
    "energy": ("energy_cp", "energy_contract", "energy_pay"),
    "rent": ("rent_cp", "rent_contract", "rent_pay"),
}


# --------------------------------------------------------------------------- helpers
def _blank(v) -> bool:
    return v in (None, "", "-") or (isinstance(v, str) and not v.strip())


def _normname(s) -> str:
    if s is None:
        return ""
    low = str(s).strip().lower().replace('"', "").replace("«", "").replace("»", "").replace("ё", "е")
    low = re.sub(r"\(.*?\)", "", low)
    low = re.sub(r"\b(ооо|оао|пао|ао|зао|ип|ано|нко|мо)\b", "", low)
    low = re.sub(r"[^\w]+", "", low)
    return low


def _clean_cp_name(s) -> str:
    return re.sub(r"\s*\(.*?\)\s*$", "", str(s).strip()).strip()


def _cp_type(name) -> str:
    n = str(name).strip().lower()
    return "ИП" if n.startswith("ип ") or n.startswith("ип.") else "ЮЛ"


def _detect_basis(*texts) -> str:
    t = " ".join(str(x).lower() for x in texts if not _blank(x))
    if "разрешен" in t:
        return "разрешение"
    if "постановлен" in t:
        return "постановление"
    if "приказ" in t:
        return "приказ"
    if "сервитут" in t:
        return "сервитут"
    return "договор"


def _parse_num_date(cell):
    if _blank(cell):
        return None, None
    s = str(cell).strip()
    m = re.search(r"(.*?)\s+от\s+(\d{1,2})[.,](\d{1,2})[.,](\d{2,4})", s, re.I)
    if m:
        number = m.group(1).strip() or s
        d, mo, y = int(m.group(2)), int(m.group(3)), int(m.group(4))
        if y < 100:
            y += 2000
        try:
            return number[:100], f"{y:04d}-{mo:02d}-{d:02d}"
        except ValueError:
            return number[:100], None
    return s[:100], None


def _iso_month(v):
    if hasattr(v, "isoformat"):
        return v.isoformat()[:7] + "-01"
    if isinstance(v, str):
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})", v)
        if m:
            return f"{m.group(1)}-{m.group(2)}-01"
        m = re.search(r"(\d{1,2})[.](\d{1,2})[.](\d{4})", v)
        if m:
            return f"{m.group(3)}-{int(m.group(2)):02d}-01"
    return None


def _parse_payment(cell):
    """-> (payment_status, paid_through, special_note)."""
    if _blank(cell):
        return "unknown", None, None
    pt = _iso_month(cell)
    if pt:
        return "paid", pt, None
    s = str(cell).strip().lower()
    if "не произв" in s or "не поступ" in s or "не оплач" in s:
        return "unpaid", None, None
    if "отсут" in s or "нет информ" in s:
        return "unknown", None, None
    return "special", None, str(cell).strip()


def _jsonable(d: dict) -> dict:
    """Сырьё строки → JSON-сериализуемое (datetime → ISO) для DataEntry.meta (L1)."""
    out = {}
    for k, v in d.items():
        if hasattr(v, "isoformat"):
            out[k] = v.isoformat()
        else:
            out[k] = v
    return out


def _cp_ids(cp: Counterparty) -> list[str]:
    return [x for x in (cp.external_ref, str(cp.id)) if x]


# --------------------------------------------------------------------------- parse (fetch stage)
def parse_reestr_xlsx(content: bytes) -> list[dict[str, Any]]:
    """xlsx (bytes) → список СЫРЫХ строк реестра (L1, без трансформации)."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"Лист «{SHEET}» не найден; есть: {wb.sheetnames}")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    out: list[dict[str, Any]] = []
    for r in rows[2:]:  # r0/r1 — двухуровневая шапка
        bu = r[0] if len(r) > 0 else None
        if _blank(bu):
            continue
        out.append({k: (r[i] if i < len(r) else None) for k, i in COL.items()})
    return out


# --------------------------------------------------------------------------- normalize → L2
async def ingest_reestr(
    db: AsyncSession, company_id, rows: list[dict[str, Any]], channel_id=None
) -> dict[str, Any]:
    """L1 RAW (DataEntry) → нормализация → L2 (контрагенты/договоры/settlements +
    DataEntry clean). Идемпотентно. Возвращает сводку для лога канала."""
    # --- индексы прода ---
    locs = (await db.execute(
        select(ServiceLocation).where(ServiceLocation.company_id == company_id)
    )).scalars().all()
    loc_by_bu: dict[str, ServiceLocation] = {}
    for l in locs:
        md = l.extra_metadata or {}
        for k in (md.get("buNumber"), l.code):
            if k is not None and str(k).strip():
                loc_by_bu[str(k).strip()] = l

    cps = (await db.execute(
        select(Counterparty).where(Counterparty.company_id == company_id)
    )).scalars().all()
    cp_by_norm: dict[str, Counterparty] = {}
    for c in cps:
        nn = _normname(c.name)
        if nn:
            cp_by_norm[nn] = c

    contracts = (await db.execute(
        select(Contract).where(Contract.company_id == company_id)
    )).scalars().all()
    contr_by_key: dict[tuple, Contract] = {}
    contr_by_cp_type: dict[tuple, list[Contract]] = {}
    for c in contracts:
        contr_by_key[(c.counterparty_id, str(c.number), c.type)] = c
        contr_by_cp_type.setdefault((c.counterparty_id, c.type), []).append(c)

    # организация-владелец договоров (для новых): наиболее частая у существующих, иначе первая Organization
    org_id = ""
    if contracts:
        from collections import Counter
        org_id = Counter(c.organization_id for c in contracts if c.organization_id).most_common(1)
        org_id = org_id[0][0] if org_id else ""
    if not org_id:
        org = (await db.execute(
            select(Organization).where(Organization.company_id == company_id)
        )).scalars().first()
        org_id = str(org.id) if org else ""

    res = {"rows": 0, "raw": 0, "clean": 0, "settlements": 0,
           "counterparties": 0, "contracts": 0, "unmatched": 0}

    for row in rows:
        if _blank(row.get("bu")):
            continue
        res["rows"] += 1
        bukey = str(row["bu"]).strip()
        if re.fullmatch(r"\d+\.0", bukey):
            bukey = bukey[:-2]

        # --- L1 RAW: DataEntry(layer='raw'), идемпотентно по source_label ---
        raw_label = f"{SOURCE_TAG}-{bukey}"
        raw_entry = (await db.execute(
            select(DataEntry).where(
                DataEntry.company_id == company_id,
                DataEntry.source_label == raw_label,
                DataEntry.layer == "raw",
            )
        )).scalar_one_or_none()
        meta_raw = _jsonable(row)
        if raw_entry:
            raw_entry.meta = meta_raw
            raw_entry.status = "recognized"
        else:
            raw_entry = DataEntry(
                company_id=company_id,
                title=f"Реестр ЭЗС №{bukey} — договоры и оплаты",
                category_id="energy", subcategory_id="contracts_payments",
                doc_type_id="reestr_row", status="recognized",
                source="upload", source_label=raw_label, source_id=raw_label,
                layer="raw", meta=meta_raw,
            )
            db.add(raw_entry)
            res["raw"] += 1
        await db.flush()

        loc = loc_by_bu.get(bukey)
        if not loc:
            res["unmatched"] += 1
            continue

        comment = None if _blank(row.get("comment")) else str(row["comment"]).strip()
        normalized: dict[str, Any] = {"buNumber": bukey, "locationId": loc.id, "roles": {}}

        for role, (cp_k, contr_k, pay_k) in ROLE_COLS.items():
            cp_cell = row.get(cp_k)
            if _blank(cp_cell) or "информац" in str(cp_cell).lower():
                continue
            ctype, ckind = ROLE_DEF[role]
            cpn = _normname(cp_cell)
            disp = _clean_cp_name(cp_cell)

            # контрагент: существующий или создать
            cp = cp_by_norm.get(cpn)
            if cp is None and cpn:
                cp = Counterparty(
                    company_id=company_id, inn="", name=disp[:500],
                    type=_cp_type(cp_cell), aliases=[role], kind="external",
                )
                db.add(cp)
                await db.flush()
                cp_by_norm[cpn] = cp
                res["counterparties"] += 1
            elif cp is not None and role not in (cp.aliases or []):
                cp.aliases = sorted(set((cp.aliases or []) + [role]))
            if cp is None:
                continue

            # договор: привязка к существующему; создаём только если у cp нет договора типа
            number, cdate = _parse_num_date(row.get(contr_k))
            basis = _detect_basis(row.get(contr_k), cp_cell)
            ids = _cp_ids(cp)
            contract = None
            if number:
                for cid in ids:
                    contract = contr_by_key.get((cid, str(number), ctype))
                    if contract:
                        break
            if contract is None:
                for cid in ids:
                    same = contr_by_cp_type.get((cid, ctype))
                    if same:
                        contract = same[0]
                        break
            if contract is None and number:
                contract = Contract(
                    company_id=company_id, number=str(number)[:100], date=cdate or "",
                    counterparty_id=str(cp.id), organization_id=org_id or str(cp.id),
                    type=ctype, kind=ckind, basis=basis, scope_type="locations",
                )
                db.add(contract)
                await db.flush()
                contr_by_cp_type.setdefault((str(cp.id), ctype), []).append(contract)
                res["contracts"] += 1
                # привязка договора к станции (scope=locations)
                db.add(ContractLocation(company_id=company_id, contract_id=contract.id, location_id=loc.id))

            # оплата → статус
            pstatus, paid_through, special = _parse_payment(row.get(pay_k))
            s_comment = comment
            if special:
                s_comment = (special + ("; " + comment if comment else ""))[:1000]

            # --- L2: StationContractSettlement (upsert company+location+role) ---
            existing = (await db.execute(
                select(StationContractSettlement).where(
                    StationContractSettlement.company_id == company_id,
                    StationContractSettlement.location_id == loc.id,
                    StationContractSettlement.role == role,
                )
            )).scalar_one_or_none()
            fields = dict(
                contract_id=contract.id if contract else None,
                counterparty_id=str(cp.id),
                paid_through=paid_through, payment_status=pstatus,
                basis=basis, comment=s_comment, period="2026-06-01", source=SOURCE_TAG,
            )
            if existing:
                for k, v in fields.items():
                    setattr(existing, k, v)
            else:
                db.add(StationContractSettlement(
                    company_id=company_id, location_id=loc.id, role=role, **fields))
                res["settlements"] += 1

            normalized["roles"][role] = {
                "counterparty": disp, "contractNumber": number, "basis": basis,
                "paymentStatus": pstatus, "paidThrough": paid_through, "comment": s_comment,
            }

        # --- L2 CLEAN: DataEntry(layer='clean', derived_from raw) ---
        clean_label = f"{SOURCE_TAG}-clean-{bukey}"
        clean_entry = (await db.execute(
            select(DataEntry).where(
                DataEntry.company_id == company_id,
                DataEntry.source_label == clean_label,
                DataEntry.layer == "clean",
            )
        )).scalar_one_or_none()
        if clean_entry:
            clean_entry.meta = normalized
            clean_entry.derived_from_entry_id = raw_entry.id
        else:
            db.add(DataEntry(
                company_id=company_id,
                title=f"Реестр ЭЗС №{bukey} — нормализовано",
                category_id="energy", subcategory_id="contracts_payments",
                doc_type_id="reestr_row", status="verified",
                source="upload", source_label=clean_label, source_id=clean_label,
                layer="clean", meta=normalized, derived_from_entry_id=raw_entry.id,
            ))
            res["clean"] += 1

    await db.flush()
    # ключи для лога канала (run_channel ожидает shifts/created/updated/skipped_kinds)
    return {
        "status": "success", "kind": "reestr",
        "shifts": res["rows"],
        "created": res["counterparties"] + res["contracts"] + res["settlements"],
        "updated": 0, "skipped_kinds": [],
        **res,
    }
