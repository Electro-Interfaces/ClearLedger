"""Отчёт доводит до предмета, а нумератор показывает число.

Цифра в обзоре ценна ровно тем, что за ней открывается список. Плитки «Без
номера», «Возвращены», «Просрочены» вели в реестр без отбора — на все документы
периода, — и человек, нажавший «Просрочены 3», получал восемнадцать. Проверяем
не наличие параметра, а то, что он реально сужает: в ответе есть искомый
документ и нет соседнего.

Отдельно — отказ на незнакомое слово. Молча вернуть весь реестр на опечатку в
адресе хуже, чем ответить ошибкой: молчание читается как «просроченных много».
"""
from datetime import datetime, timedelta, timezone

import pytest
from httpx import AsyncClient

from tests.helpers import seed_company_id

pytestmark = pytest.mark.asyncio(loop_scope="session")


async def _company(client: AsyncClient) -> str:
    me = (await client.get("/api/auth/me")).json()
    assert me["companies"], "сид-суперадмин не состоит ни в одной компании"
    return seed_company_id(me)


async def _kind(client: AsyncClient, cid: str) -> dict:
    await client.post(f"/api/docs/kinds/starter?company_id={cid}")
    kinds = (await client.get("/api/docs/kinds", params={"company_id": cid})).json()["kinds"]
    return next(k for k in kinds if k["code"] == "doc_in")


async def _ids(client: AsyncClient, cid: str, attention: str) -> set[str]:
    r = await client.get("/api/docs", params={
        "company_id": cid, "attention": attention, "limit": 500})
    assert r.status_code == 200, r.text
    return {d["id"] for d in r.json()["docs"]}


async def test_отбор_отчёта_сужает_реестр(auth_client: AsyncClient):
    cid = await _company(auth_client)
    kind = await _kind(auth_client, cid)
    вчера = datetime.now(timezone.utc) - timedelta(days=2)

    async def завести(title: str, *, register: bool, due=None) -> str:
        r = await auth_client.post("/api/docs", json={
            "company_id": cid, "kind_id": kind["id"], "title": title,
            **({"due_at": due.isoformat()} if due else {})})
        assert r.status_code == 201, r.text
        doc_id = r.json()["id"]
        if register:
            r = await auth_client.post(f"/api/docs/{doc_id}/register",
                                       json={"company_id": cid})
            assert r.status_code == 200, r.text
        return doc_id

    черновик = await завести("[отчёт] без номера", register=False)
    просрочен = await завести("[отчёт] срок прошёл", register=True, due=вчера)
    в_сроке = await завести("[отчёт] срок не наступил", register=True,
                            due=datetime.now(timezone.utc) + timedelta(days=30))

    без_номера = await _ids(auth_client, cid, "unnumbered")
    assert черновик in без_номера
    assert просрочен not in без_номера, "зарегистрированный попал в «без номера»"

    просроченные = await _ids(auth_client, cid, "overdue")
    assert просрочен in просроченные
    assert в_сроке not in просроченные, "документ в сроке попал в просроченные"
    assert черновик not in просроченные, "документ без срока попал в просроченные"

    # Без отбора реестр по-прежнему отдаёт всё: сужение не должно протекать
    # в обычный вызов.
    r = await auth_client.get("/api/docs", params={"company_id": cid, "limit": 500})
    все = {d["id"] for d in r.json()["docs"]}
    assert {черновик, просрочен, в_сроке} <= все


async def test_незнакомый_отбор_это_отказ(auth_client: AsyncClient):
    cid = await _company(auth_client)
    r = await auth_client.get("/api/docs", params={
        "company_id": cid, "attention": "просроченные"})
    assert r.status_code == 400, r.text


async def test_нумератор_отдаёт_текущее_значение(auth_client: AsyncClient):
    """Ради этого числа экран «Нумераторы» и существует отдельно от «Видов»."""
    cid = await _company(auth_client)
    kind = await _kind(auth_client, cid)

    def счёт(ответ: dict) -> int:
        строка = next(k for k in ответ["counters"] if k["kind_id"] == kind["id"])
        for область in строка["scopes"]:
            assert область["next"] == область["issued"] + 1, \
                "следующий номер обязан быть на единицу больше выданного"
        return строка["issued"]

    до = счёт((await auth_client.get("/api/docs/counters",
                                     params={"company_id": cid})).json())

    r = await auth_client.post("/api/docs", json={
        "company_id": cid, "kind_id": kind["id"], "title": "[отчёт] в счёт"})
    assert r.status_code == 201, r.text
    # Счётчик двигает регистрация, а не заведение карточки: у черновика номера нет.
    середина = счёт((await auth_client.get("/api/docs/counters",
                                           params={"company_id": cid})).json())
    assert середина == до, "черновик сдвинул счётчик"

    r = await auth_client.post(f"/api/docs/{r.json()['id']}/register",
                               json={"company_id": cid})
    assert r.status_code == 200, r.text
    после = счёт((await auth_client.get("/api/docs/counters",
                                        params={"company_id": cid})).json())
    assert после == до + 1, f"счётчик не сдвинулся: было {до}, стало {после}"


async def test_выгрузка_отчёта_открывается_книгой(auth_client: AsyncClient):
    """Файл должен открываться Excel и называть себя.

    Проверяем не «ответ 200», а то, что книга читается и первый лист говорит,
    что это за отчёт и за какой период: файл без шапки через неделю в почте
    неотличим от соседнего, и спор о цифре начинается с «а это по чему?».
    """
    import io as _io

    from openpyxl import load_workbook

    cid = await _company(auth_client)
    for report in ("docs", "discipline", "errands", "calendar"):
        r = await auth_client.get("/api/docs/reports/export", params={
            "report": report, "company_id": cid,
            "date_from": "2026-08-01", "date_to": "2026-08-31"})
        assert r.status_code == 200, f"{report}: {r.text[:300]}"
        assert "spreadsheetml" in r.headers["content-type"], report
        # Имя по RFC 5987: без него кириллица не доедет до диска.
        assert "filename*=UTF-8''" in r.headers["content-disposition"], report

        wb = load_workbook(_io.BytesIO(r.content))
        assert wb.sheetnames[0] == "Отчёт", f"{report}: шапки нет: {wb.sheetnames}"
        подписи = [строка[0] for строка in wb["Отчёт"].iter_rows(values_only=True)]
        assert "Период" in подписи and "Выгрузил" in подписи, report
        assert len(wb.sheetnames) > 1, f"{report}: книга без данных"


async def test_выгрузка_реестра_держит_отбор(auth_client: AsyncClient):
    """Выгрузка «просроченных» не должна отдавать весь период.

    Отбор из обзора приезжает в адрес выгрузки, и если ручка его не примет,
    человек получит книгу на восемнадцать строк там, где на экране было три, —
    и заметит это в лучшем случае на совещании.
    """
    import io as _io

    from openpyxl import load_workbook

    cid = await _company(auth_client)
    kind = await _kind(auth_client, cid)
    r = await auth_client.post("/api/docs", json={
        "company_id": cid, "kind_id": kind["id"], "title": "[отчёт] выгрузка без номера"})
    assert r.status_code == 201, r.text

    def строк(content: bytes) -> int:
        wb = load_workbook(_io.BytesIO(content))
        # Лист «Реестр» без строки заголовков.
        return wb["Реестр"].max_row - 1

    общая = await auth_client.get("/api/docs/export", params={
        "company_id": cid, "format": "xlsx"})
    узкая = await auth_client.get("/api/docs/export", params={
        "company_id": cid, "format": "xlsx", "attention": "unnumbered"})
    assert общая.status_code == 200 and узкая.status_code == 200, узкая.text
    все, без_номера = строк(общая.content), строк(узкая.content)
    assert 0 < без_номера < все, f"отбор не сузил выгрузку: {без_номера} из {все}"
